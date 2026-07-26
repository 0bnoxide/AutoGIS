"""compliance_summary.py — cross-event compliance matrix with trend analysis."""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill

from ..common.config import ND_QUALIFIERS
from ..common.qa import QACollector, QARecord, SEV_INFO

_FILL_EXCEED = PatternFill(fill_type="solid", fgColor="FF9999")
_FILL_DETECT = PatternFill(fill_type="solid", fgColor="FFFF99")


@dataclass
class ComplianceRecord:
    location_id: str
    analyte_name: str
    screening_level: Optional[float]
    units: str
    ever_detected: bool
    ever_exceeded: bool
    detection_count: int
    exceedance_count: int
    total_event_count: int
    max_value: Optional[float]
    max_event_date: str
    max_ratio: Optional[float]
    trend: str


@dataclass
class ComplianceSummaryResult:
    records: list
    well_count: int
    analyte_count: int
    locations_with_exceedances: int
    qa: QACollector


def compute_trend(values: list, dates: list, min_points: int = 3) -> str:
    if len(values) < min_points:
        return "insufficient_data"
    # Ordinal from ISO date prefix (first 7 chars YYYY-MM)
    try:
        xs = [float(d[:4]) * 12 + float(d[5:7]) for d in dates]
    except (ValueError, IndexError):
        xs = list(range(len(values)))
    n = len(xs)
    sum_x = sum(xs)
    sum_y = sum(values)
    sum_xy = sum(x * y for x, y in zip(xs, values))
    sum_xx = sum(x * x for x in xs)
    denom = n * sum_xx - sum_x ** 2
    if denom == 0:
        return "stable"
    slope = (n * sum_xy - sum_x * sum_y) / denom
    if slope > 0.01:
        return "worsening"
    if slope < -0.01:
        return "improving"
    return "stable"


def _parse_float(v: str) -> Optional[float]:
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def build_compliance_summary(
    result_rows: list,
    *,
    screening_levels: Optional[dict] = None,
    analytes: Optional[list] = None,
    wells: Optional[list] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    nd_qualifiers: frozenset = ND_QUALIFIERS,
    min_events_for_trend: int = 3,
    qa: Optional[QACollector] = None,
) -> ComplianceSummaryResult:
    if qa is None:
        qa = QACollector()
    sl = screening_levels or {}

    rows = result_rows
    if analytes:
        rows = [r for r in rows if r.get("AnalyteName") in analytes]
    if wells:
        rows = [r for r in rows if r.get("LocationID") in wells]
    if date_from:
        rows = [r for r in rows if r.get("SampleDate", "") >= date_from]
    if date_to:
        rows = [r for r in rows if r.get("SampleDate", "") <= date_to]

    groups: dict[tuple, list] = defaultdict(list)
    for r in rows:
        groups[(r.get("LocationID", ""), r.get("AnalyteName", ""))].append(r)

    records = []
    exceed_locations: set = set()

    for (loc, analyte), grp in sorted(groups.items()):
        qual_detected = [
            r for r in grp
            if r.get("ResultQualifier", "").upper().strip() not in nd_qualifiers
            and _parse_float(r.get("ResultValue", "")) is not None
        ]
        ever_detected = len(qual_detected) > 0
        screening = sl.get(analyte)
        exceedances = [
            r for r in qual_detected
            if screening is not None
            and (_parse_float(r.get("ResultValue", "")) or 0) > screening
        ]
        ever_exceeded = len(exceedances) > 0
        if ever_exceeded:
            exceed_locations.add(loc)

        det_values = [_parse_float(r.get("ResultValue", "")) for r in qual_detected
                      if _parse_float(r.get("ResultValue", "")) is not None]
        det_dates  = [r.get("SampleDate", "") for r in qual_detected
                      if _parse_float(r.get("ResultValue", "")) is not None]

        if det_values:
            max_val = max(det_values)
            max_idx = det_values.index(max_val)
            max_date = det_dates[max_idx] if max_idx < len(det_dates) else ""
            max_ratio = (max_val / screening) if screening else None
        else:
            max_val = None
            max_date = ""
            max_ratio = None

        trend = compute_trend(det_values, det_dates, min_points=min_events_for_trend)
        units = grp[0].get("ReportedUnits", "") if grp else ""

        records.append(ComplianceRecord(
            location_id=loc, analyte_name=analyte,
            screening_level=screening, units=units,
            ever_detected=ever_detected, ever_exceeded=ever_exceeded,
            detection_count=len(qual_detected),
            exceedance_count=len(exceedances),
            total_event_count=len(grp),
            max_value=max_val, max_event_date=max_date,
            max_ratio=round(max_ratio, 4) if max_ratio is not None else None,
            trend=trend,
        ))

    all_wells = sorted({r.location_id for r in records})
    all_analytes = sorted({r.analyte_name for r in records})

    qa.add(QARecord(SEV_INFO, "compliance_built",
                    f"{len(all_wells)} wells, {len(all_analytes)} analytes, "
                    f"{len(exceed_locations)} with exceedances"))

    return ComplianceSummaryResult(
        records=records, well_count=len(all_wells),
        analyte_count=len(all_analytes),
        locations_with_exceedances=len(exceed_locations), qa=qa,
    )


def write_compliance_workbook(
    result: ComplianceSummaryResult,
    out_path: Path,
    *,
    group_map: Optional[dict] = None,
) -> None:
    wb = openpyxl.Workbook()

    # Matrix sheet
    matrix_ws = wb.active
    matrix_ws.title = "Compliance Matrix"
    all_wells = sorted({r.location_id for r in result.records})
    all_analytes = sorted({r.analyte_name for r in result.records})

    # Index records
    rec_idx = {(r.location_id, r.analyte_name): r for r in result.records}

    # Header row
    matrix_ws.cell(1, 1, "Location").font = Font(bold=True)
    for col_idx, analyte in enumerate(all_analytes, start=2):
        matrix_ws.cell(1, col_idx, analyte).font = Font(bold=True)

    for row_idx, well in enumerate(all_wells, start=2):
        matrix_ws.cell(row_idx, 1, well).font = Font(bold=True)
        for col_idx, analyte in enumerate(all_analytes, start=2):
            rec = rec_idx.get((well, analyte))
            cell = matrix_ws.cell(row_idx, col_idx)
            if rec is None:
                cell.value = "—"
            elif not rec.ever_detected:
                cell.value = "ND"
            elif rec.ever_exceeded:
                cell.value = f"{rec.max_value}**"
                cell.fill = _FILL_EXCEED
            else:
                cell.value = rec.max_value
                cell.fill = _FILL_DETECT

    # Detail sheet
    detail_ws = wb.create_sheet("Detail")
    det_headers = [
        "LocationID", "AnalyteName", "ScreeningLevel", "Units",
        "EverDetected", "EverExceeded", "DetectionCount", "ExceedanceCount",
        "TotalEvents", "MaxValue", "MaxRatio", "Trend",
    ]
    detail_ws.append(det_headers)
    for rec in result.records:
        detail_ws.append([
            rec.location_id, rec.analyte_name, rec.screening_level, rec.units,
            rec.ever_detected, rec.ever_exceeded, rec.detection_count,
            rec.exceedance_count, rec.total_event_count,
            rec.max_value, rec.max_ratio, rec.trend,
        ])

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
