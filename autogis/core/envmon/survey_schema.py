"""survey_schema.py — XLSForm reader, validator, and schema-drift classifier.

Survey123 add-on roadmap Phase 1 (ADR-0112; decision record ADR-0115).
openpyxl + stdlib only — no arcpy, no arcgis. The form-vs-layer leg reuses
autogis.core.agol.audit_schema.diff_schema (itself arcgis-free).
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Optional

import openpyxl

from ..common.qa import QACollector, QARecord, SEV_ERROR, SEV_WARNING, SEV_INFO
from .sample_id import xform_sample_id_calc

CLASS_SAFE = "safe"
CLASS_REVIEW = "review-required"
CLASS_DESTRUCTIVE = "destructive"
_CLASS_RANK = {CLASS_SAFE: 1, CLASS_REVIEW: 2, CLASS_DESTRUCTIVE: 3}

_FIELD_BASES = {"text", "integer", "decimal", "date", "datetime", "time",
                "calculate", "geopoint", "barcode",
                "select_one", "select_multiple"}
_STRUCT_BASES = {"begin_group", "end_group", "begin_repeat", "end_repeat",
                 "note"}
#: xlsform.org: "Names have to start with a letter or an underscore. Names
#: can only contain letters, digits, hyphens, underscores, and periods."
#: Applies to the survey sheet's name column only -- choice VALUES are far
#: more permissive (see _check_choices).
_NAME_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_.\-]*$")
_REF_RE = re.compile(r"\$\{([^}]+)\}")
_TRUTHY_REQUIRED = {"yes", "true", "true()"}
_REQUIRED_VALUES = {"", "yes", "no", "true", "false", "true()", "false()"}


@dataclass
class SurveyQuestion:
    type: str = ""
    name: str = ""
    label: str = ""
    hint: str = ""
    required: str = ""
    calculation: str = ""
    relevant: str = ""
    constraint: str = ""
    appearance: str = ""
    default: str = ""
    row: int = 0

    @property
    def base(self) -> str:
        return self.type.split()[0].lower() if self.type.strip() else ""

    @property
    def list_name(self) -> str:
        parts = self.type.split()
        if self.base.startswith("select_") and len(parts) > 1:
            return parts[1]
        return ""


@dataclass
class FormSchema:
    questions: list = field(default_factory=list)
    choices: dict = field(default_factory=dict)
    settings: dict = field(default_factory=dict)


def _header_map(ws) -> dict:
    return {str(c.value).strip().lower(): i
            for i, c in enumerate(ws[1], start=1) if c.value}


def _cell(ws, r, col: Optional[int]) -> str:
    if not col:
        return ""
    v = ws.cell(r, col).value
    return str(v).strip() if v is not None else ""


def _sheet(wb, name: str):
    """Case-insensitive sheet lookup. Excel round-trips capitalize sheet
    names, and an XLSForm with a 'Survey' sheet is still an XLSForm."""
    for sn in wb.sheetnames:
        if sn.lower() == name:
            return wb[sn]
    return None


def read_xlsform(path) -> FormSchema:
    """Read an XLSForm .xlsx. Columns map by header row, not position;
    absent optional columns/sheets yield empty values — validation, not the
    reader, decides what is an error.

    A workbook with no 'survey' sheet is not an XLSForm and raises. Returning
    an empty schema instead would make diff_forms report every question in the
    *other* form as a safe addition, so pointing --baseline-form at the wrong
    workbook would produce a clean bill of health on a publication gate.
    """
    wb = openpyxl.load_workbook(str(path), data_only=True)
    schema = FormSchema()
    ws = _sheet(wb, "survey")
    if ws is None:
        wb.close()
        raise ValueError(
            f"{path}: no 'survey' sheet -- not an XLSForm "
            f"(sheets: {', '.join(wb.sheetnames) or 'none'})")
    if ws is not None:
        h = _header_map(ws)
        for r in range(2, ws.max_row + 1):
            q = SurveyQuestion(
                type=_cell(ws, r, h.get("type")),
                name=_cell(ws, r, h.get("name")),
                label=_cell(ws, r, h.get("label")),
                hint=_cell(ws, r, h.get("hint")),
                required=_cell(ws, r, h.get("required")),
                calculation=_cell(ws, r, h.get("calculation")),
                relevant=_cell(ws, r, h.get("relevant")),
                constraint=_cell(ws, r, h.get("constraint")),
                appearance=_cell(ws, r, h.get("appearance")),
                default=_cell(ws, r, h.get("default")),
                row=r,
            )
            if q.type or q.name or q.label or q.calculation:
                schema.questions.append(q)
    cs = _sheet(wb, "choices")
    if cs is not None:
        h = _header_map(cs)
        for r in range(2, cs.max_row + 1):
            ln = _cell(cs, r, h.get("list_name"))
            nm = _cell(cs, r, h.get("name"))
            lb = _cell(cs, r, h.get("label"))
            if ln or nm:
                schema.choices.setdefault(ln, []).append((nm, lb))
    st = _sheet(wb, "settings")
    if st is not None:
        h = _header_map(st)
        for key, col in h.items():
            schema.settings[key] = _cell(st, 2, col)
    wb.close()
    return schema


_SELECTED_RE = re.compile(r'selected\(\s*\$\{(\w+)\}\s*,\s*"([^"]*)"\s*\)')
_SLUG_RE = re.compile(r"^[a-z0-9_]+$")


def _err(qa, cat, msg):
    qa.add(QARecord(SEV_ERROR, cat, msg))


def _warn(qa, cat, msg):
    qa.add(QARecord(SEV_WARNING, cat, msg))


def validate_form(schema: FormSchema, qa: QACollector, *,
                  event_config: Optional[dict] = None,
                  site_config: Optional[dict] = None,
                  analyte_dict: Optional[dict] = None) -> None:
    """Static XLSForm validation (ADR-0115 checks 1-13). Appends QARecords."""
    site_id = (site_config or {}).get("site_id", "")
    ctx = f"[{site_id}] " if site_id else ""

    # 1. sheets
    if not schema.questions:
        _err(qa, "missing_sheet", ctx + "survey sheet missing or empty")
    if not schema.choices:
        _err(qa, "missing_sheet", ctx + "choices sheet missing or empty")
    if not schema.settings:
        _warn(qa, "settings_incomplete", ctx + "settings sheet missing")

    names: dict = {}
    fields = [q for q in schema.questions if q.base in _FIELD_BASES]

    for q in schema.questions:
        base = q.base
        # 3. type recognized
        if base and base not in _FIELD_BASES and base not in _STRUCT_BASES:
            _err(qa, "unknown_type",
                 f"row {q.row}: unknown question type {q.type!r}")
        if base.startswith("select_") and not q.list_name:
            _err(qa, "unknown_type",
                 f"row {q.row}: {base} without a choice list name")
        # 2. names
        if base in _FIELD_BASES:
            if not q.name:
                _err(qa, "missing_name", f"row {q.row}: {base} without a name")
            else:
                if not _NAME_RE.match(q.name):
                    _err(qa, "invalid_name",
                         f"row {q.row}: invalid name {q.name!r}")
                key = q.name.lower()
                if key in names:
                    _err(qa, "duplicate_name",
                         f"row {q.row}: duplicate name {q.name!r} "
                         f"(first at row {names[key]})")
                else:
                    names[key] = q.row
        # 7. required sanity
        if q.required and q.required.lower() not in _REQUIRED_VALUES:
            _warn(qa, "odd_required_value",
                  f"row {q.row}: required={q.required!r} on {q.name!r}")

    # 4. select_* list resolution + choices hygiene
    for q in fields:
        if q.base.startswith("select_") and q.list_name:
            if q.list_name not in schema.choices:
                _err(qa, "unknown_choice_list",
                     f"row {q.row}: {q.name!r} references missing choice "
                     f"list {q.list_name!r}")
            elif not [c for c, _ in schema.choices[q.list_name] if c]:
                _err(qa, "empty_choice_list",
                     f"row {q.row}: choice list {q.list_name!r} is empty")
    # Choice VALUES are not identifiers. xlsform.org constrains them exactly
    # once: "Choice names for select_multiple must not contain spaces because
    # spaces are used as a separator"; select_one "may contain spaces".
    # Applying the survey-sheet name rule here rejected Likert codes 1/2 and
    # -- because build_xlsform writes location_ids verbatim as choice codes --
    # made validate-survey-form fail on build-survey-form's own output.
    multi_lists = {q.list_name for q in fields
                   if q.base == "select_multiple" and q.list_name}
    allow_dupes = str(
        schema.settings.get("allow_choice_duplicates", "")).strip().lower() \
        in ("yes", "true")
    for ln, pairs in schema.choices.items():
        seen = set()
        for cname, _lbl in pairs:
            if not cname:
                _err(qa, "invalid_choice_name",
                     f"list {ln!r}: empty choice name")
                continue
            if ln in multi_lists and " " in cname:
                _err(qa, "invalid_choice_name",
                     f"list {ln!r}: select_multiple choice {cname!r} "
                     "contains a space")
                continue
            if cname in seen and not allow_dupes:
                _err(qa, "duplicate_choice",
                     f"list {ln!r}: duplicate choice {cname!r}")
            seen.add(cname)

    # 5. ${ref} resolution (order-independent by design — ADR-0113 emits
    # SampleID after QAFlags; XLSForm resolves calculates by dependency)
    # Every named row is addressable, not just field-type ones: count(${rep})
    # over a begin_repeat, indexed-repeat(...), and a relevant that references
    # a note are all standard XLSForm. Only the end_* rows carry no name.
    all_names = {q.name for q in schema.questions
                 if q.name and q.base not in ("end_group", "end_repeat")}
    for q in schema.questions:
        for expr in (q.calculation, q.relevant, q.constraint):
            for ref in _REF_RE.findall(expr or ""):
                if ref not in all_names:
                    _err(qa, "unresolved_reference",
                         f"row {q.row}: ${{{ref}}} does not resolve to a "
                         f"question name")

    # 6. group/repeat balance
    stack = []
    for q in schema.questions:
        if q.base in ("begin_group", "begin_repeat"):
            stack.append((q.base, q.row))
        elif q.base in ("end_group", "end_repeat"):
            want = "begin_" + q.base.split("_")[1]
            if not stack or stack[-1][0] != want:
                _err(qa, "unbalanced_group",
                     f"row {q.row}: {q.base} without matching {want}")
            else:
                stack.pop()
    for base, row in stack:
        _err(qa, "unbalanced_group", f"row {row}: {base} never closed")

    # 8-9. SampleID contract (ADR-0113) + duplicate-leg dependency
    sid = next((q for q in fields if q.name == "SampleID"), None)
    if sid is None:
        _err(qa, "sample_id_missing",
             ctx + "no calculate question named 'SampleID'")
    else:
        if sid.base != "calculate":
            _err(qa, "sample_id_contract_mismatch",
                 f"row {sid.row}: SampleID must be a calculate, got "
                 f"{sid.type!r}")
        expected = " ".join(xform_sample_id_calc().split())
        actual = " ".join((sid.calculation or "").split())
        if actual != expected:
            _err(qa, "sample_id_contract_mismatch",
                 f"row {sid.row}: SampleID calculation diverges from the "
                 f"ADR-0113 contract; expected {expected!r}")
        for m in _SELECTED_RE.finditer(sid.calculation or ""):
            flag_q, flag_val = m.group(1), m.group(2)
            src = next((q for q in fields if q.name == flag_q), None)
            if src is not None and src.list_name:
                have = {c for c, _ in schema.choices.get(src.list_name, [])}
                if flag_val not in have:
                    _err(qa, "sample_id_dup_leg",
                         f"choice list {src.list_name!r} lacks the "
                         f"{flag_val!r} choice the SampleID duplicate leg "
                         f"reads")

    # 10. settings
    if schema.settings:
        fid = schema.settings.get("form_id", "")
        if not fid or not _SLUG_RE.match(fid):
            _warn(qa, "settings_incomplete",
                  f"settings: form_id {fid!r} missing or not slug-shaped")
        if not schema.settings.get("version"):
            _warn(qa, "settings_incomplete", "settings: version missing")

    _cross_reference_checks(schema, qa, event_config, analyte_dict, ctx)

    qa.add(QARecord(SEV_INFO, "validation_complete",
                    f"{ctx}validated {len(fields)} questions, "
                    f"{len(schema.choices)} choice lists"))


def _list_for(schema: FormSchema, qname: str) -> Optional[str]:
    q = next((q for q in schema.questions
              if q.name == qname and q.base.startswith("select_")), None)
    return q.list_name if q else None


def _set_check(schema, qa, qname, expected, miss_cat, extra_cat, what, ctx):
    ln = _list_for(schema, qname)
    if ln is None:
        if expected:
            _err(qa, miss_cat,
                 f"{ctx}form has no {qname!r} select question to carry "
                 f"{what}: {', '.join(sorted(set(expected)))}")
        return
    have = {c for c, _ in schema.choices.get(ln, [])}
    missing = sorted(set(expected) - have)
    extra = sorted(have - set(expected))
    if missing:
        _err(qa, miss_cat, f"{ctx}{what} missing from form list {ln!r}: "
                           f"{', '.join(missing)}")
    if extra:
        _warn(qa, extra_cat, f"{ctx}extra {what} in form list {ln!r}: "
                             f"{', '.join(extra)}")


def _cross_reference_checks(schema, qa, event_config, analyte_dict, ctx):
    """Spec checks 11-13 — each only when its config is supplied."""
    if not event_config:
        # Say so. Returning silently let `validate-survey-form form.xlsx` with
        # no config flags report PASS having run none of 11-13 -- a green
        # light for an unverified form on what is meant to be a gate.
        qa.add(QARecord(SEV_INFO, "cross_checks_skipped",
                        f"{ctx}no event config supplied -- location, matrix, "
                        "crew and analyte cross-checks were not run"))
        return
    from .survey123_form_builder import _field_name, _slug

    _set_check(schema, qa, "WellID", event_config.get("location_ids", []),
               "missing_location_choice", "extra_location_choice",
               "planned locations", ctx)
    _set_check(schema, qa, "Matrix", event_config.get("matrices", []),
               "missing_matrix_choice", "extra_matrix_choice",
               "matrices", ctx)
    _set_check(schema, qa, "SampledBy",
               [_slug(m) for m in event_config.get("crew_list", [])],
               "missing_crew_choice", "extra_crew_choice", "crew", ctx)

    groups = event_config.get("analyte_groups", {})
    if analyte_dict is not None:
        known = set((analyte_dict.get("analytes") or {}))
        unknown = sorted({a for names in groups.values()
                          if isinstance(names, list) for a in names} - known)
        if unknown:
            _warn(qa, "unknown_analyte",
                  f"{ctx}analyte_groups name(s) not in the analyte "
                  f"dictionary: {', '.join(unknown)}")

    used: set = set()
    expected = {}
    for names in groups.values():
        if not isinstance(names, list):
            continue
        for analyte in names:
            expected[_field_name(analyte, used)] = analyte
    have = {q.name for q in schema.questions
            if q.base == "decimal" and q.name}
    missing = sorted(expected[n] for n in expected if n not in have)
    if missing:
        _err(qa, "missing_analyte_question",
             f"{ctx}no decimal question for analyte(s): "
             f"{', '.join(missing)}")

    # decimals inside grp_* groups that no analyte accounts for
    gstack, unexpected = [], []
    for q in schema.questions:
        if q.base == "begin_group":
            gstack.append(q.name)
        elif q.base == "end_group":
            gstack and gstack.pop()
        elif (q.base == "decimal" and q.name and q.name not in expected
              and any(g.startswith("grp_") for g in gstack)):
            unexpected.append(q.name)
    if unexpected:
        _warn(qa, "unexpected_analyte_question",
              f"{ctx}decimal question(s) in analyte groups with no "
              f"analyte behind them: {', '.join(sorted(unexpected))}")


# ------------------------------------------------------------------ diffing

@dataclass
class SchemaChange:
    kind: str
    classification: str
    name: str
    detail: str


def worst_classification(changes) -> Optional[str]:
    worst = None
    for c in changes:
        if worst is None or _CLASS_RANK[c.classification] > _CLASS_RANK[worst]:
            worst = c.classification
    return worst


def _is_required(v: str) -> bool:
    return (v or "").strip().lower() in _TRUTHY_REQUIRED


def _scopes(questions) -> dict:
    """name -> (repeat_path, group_path) for field-type questions."""
    out, rstack, gstack = {}, [], []
    for q in questions:
        b = q.base
        if b == "begin_repeat":
            rstack.append(q.name)
        elif b == "end_repeat":
            rstack and rstack.pop()
        elif b == "begin_group":
            gstack.append(q.name)
        elif b == "end_group":
            gstack and gstack.pop()
        elif b in _FIELD_BASES and q.name:
            out[q.name] = (tuple(rstack), tuple(gstack))
    return out


def diff_forms(old: FormSchema, new: FormSchema) -> list:
    """Classify every change between two forms (ADR-0115 taxonomy table)."""
    changes: list = []

    def add(kind, cls, name, detail):
        changes.append(SchemaChange(kind, cls, name, detail))

    oldq = {q.name: q for q in old.questions
            if q.base in _FIELD_BASES and q.name}
    newq = {q.name: q for q in new.questions
            if q.base in _FIELD_BASES and q.name}
    oscope, nscope = _scopes(old.questions), _scopes(new.questions)

    orep = {q.name for q in old.questions if q.base == "begin_repeat"}
    nrep = {q.name for q in new.questions if q.base == "begin_repeat"}
    for name in sorted(orep - nrep):
        add("repeat_removed", CLASS_DESTRUCTIVE, name,
            f"repeat {name!r} removed — submission data shape changes")
    for name in sorted(nrep - orep):
        add("repeat_added", CLASS_DESTRUCTIVE, name,
            f"repeat {name!r} added — submission data shape changes")

    for name in sorted(oldq.keys() - newq.keys()):
        add("question_removed", CLASS_DESTRUCTIVE, name,
            f"question {name!r} removed (a rename is a removal)")
    for name in sorted(newq.keys() - oldq.keys()):
        q = newq[name]
        if _is_required(q.required):
            add("question_added_required", CLASS_REVIEW, name,
                f"new required question {name!r}")
        else:
            add("question_added", CLASS_SAFE, name,
                f"new optional question {name!r}")

    for name in sorted(oldq.keys() & newq.keys()):
        o, n = oldq[name], newq[name]
        if o.base != n.base:
            add("type_changed", CLASS_DESTRUCTIVE, name,
                f"type {o.base!r} -> {n.base!r}")
        elif o.base.startswith("select_") and o.list_name != n.list_name:
            add("list_repointed", CLASS_REVIEW, name,
                f"choice list {o.list_name!r} -> {n.list_name!r}")
        if oscope.get(name, ((), ()))[0] != nscope.get(name, ((), ()))[0]:
            add("repeat_scope_changed", CLASS_DESTRUCTIVE, name,
                f"repeat scope {oscope.get(name)!r} -> {nscope.get(name)!r}")
        elif oscope.get(name, ((), ()))[1] != nscope.get(name, ((), ()))[1]:
            add("group_changed", CLASS_REVIEW, name,
                f"group {oscope.get(name)!r} -> {nscope.get(name)!r}")
        if _is_required(o.required) != _is_required(n.required):
            if _is_required(n.required):
                add("required_tightened", CLASS_REVIEW, name,
                    "optional -> required")
            else:
                add("required_relaxed", CLASS_SAFE, name,
                    "required -> optional")
        oc = " ".join((o.calculation or "").split())
        nc = " ".join((n.calculation or "").split())
        if oc != nc:
            if name == "SampleID":
                add("sample_id_calculation_changed", CLASS_DESTRUCTIVE, name,
                    "the ADR-0113 identity calculation changed")
            else:
                add("calculation_changed", CLASS_REVIEW, name,
                    f"calculation {oc!r} -> {nc!r}")
        cosmetic = [c for c in ("label", "hint", "appearance", "default")
                    if getattr(o, c) != getattr(n, c)]
        if cosmetic:
            add("cosmetic_changed", CLASS_SAFE, name,
                f"changed: {', '.join(cosmetic)}")

    # choices, per list
    for ln in sorted(old.choices.keys() | new.choices.keys()):
        omap = dict(old.choices.get(ln, []))
        nmap = dict(new.choices.get(ln, []))
        removed = {c: omap[c] for c in omap.keys() - nmap.keys()}
        added = {c: nmap[c] for c in nmap.keys() - omap.keys()}
        # same label on a removed+added pair = a code change
        for rc, rl in sorted(removed.items()):
            match = next((ac for ac, al in sorted(added.items())
                          if al == rl and rl != ""), None)
            if match is not None:
                add("choice_code_changed", CLASS_DESTRUCTIVE, ln,
                    f"list {ln!r}: code {rc!r} -> {match!r} (label {rl!r})")
                del added[match]
            else:
                add("choice_removed", CLASS_REVIEW, ln,
                    f"list {ln!r}: choice {rc!r} removed")
        for ac in sorted(added):
            add("choice_added", CLASS_SAFE, ln,
                f"list {ln!r}: choice {ac!r} added")
        for c in sorted(omap.keys() & nmap.keys()):
            if omap[c] != nmap[c]:
                add("choice_label_changed", CLASS_SAFE, ln,
                    f"list {ln!r}: {c!r} label {omap[c]!r} -> {nmap[c]!r}")

    # settings
    of, nf = old.settings.get("form_id", ""), new.settings.get("form_id", "")
    if of != nf:
        add("form_id_changed", CLASS_DESTRUCTIVE, "form_id",
            f"form_id {of!r} -> {nf!r} rebinds the portal item")
    for key in ("form_title", "version", "instance_name"):
        if old.settings.get(key, "") != new.settings.get(key, ""):
            add("settings_changed", CLASS_SAFE, key,
                f"{key} {old.settings.get(key, '')!r} -> "
                f"{new.settings.get(key, '')!r}")
    return changes


# ------------------------------------------------------------ form vs layer

_FIELD_TYPE_TO_ESRI = {
    "text": "esriFieldTypeString", "calculate": "esriFieldTypeString",
    "barcode": "esriFieldTypeString", "time": "esriFieldTypeString",
    "select_one": "esriFieldTypeString",
    "select_multiple": "esriFieldTypeString",
    "integer": "esriFieldTypeInteger", "decimal": "esriFieldTypeDouble",
    "date": "esriFieldTypeDate", "datetime": "esriFieldTypeDate",
}

_DRIFT_CLASS = {
    "TYPE_MISMATCH": CLASS_DESTRUCTIVE,
    "EXTRA_FIELD": CLASS_REVIEW,        # form question with no layer field
    "DOMAIN_DRIFT": CLASS_REVIEW,
    "NULLABLE_MISMATCH": CLASS_REVIEW,
    "MISSING_FIELD": CLASS_SAFE,        # layer field the form doesn't collect
}


def form_layer_fields(schema: FormSchema) -> list:
    """Map form questions to AGOL-REST-shaped field dicts (the 'fetched'
    side of audit_schema.diff_schema). select_multiple gets no domain —
    Survey123 stores it comma-joined.

    ponytail: no `nullable` is emitted, so audit_schema's NULLABLE_MISMATCH
    cannot fire from the form side (the _DRIFT_CLASS row is forward-compat
    only); upgrade path is mapping required -> nullable if a saved spec
    ever needs the check."""
    fields = []
    for q in schema.questions:
        esri = _FIELD_TYPE_TO_ESRI.get(q.base)
        if not esri or not q.name:
            continue
        f = {"name": q.name, "type": esri}
        if q.base == "select_one" and q.list_name:
            pairs = schema.choices.get(q.list_name, [])
            if pairs:
                f["domain"] = {
                    "name": q.list_name,
                    "codedValues": [{"code": c, "name": l}
                                    for c, l in pairs],
                }
        fields.append(f)
    return fields


def diff_form_vs_layer(schema: FormSchema, layer_spec: dict) -> list:
    """Compatibility of a form against a saved feature-layer definition
    (audit_schema local-spec format). Reuses audit_schema.diff_schema; the
    form plays the fetched side."""
    from ..agol.audit_schema import diff_schema

    form_fields = form_layer_fields(schema)
    spec_by_name = {f.get("name"): f for f in layer_spec.get("fields", [])}
    for f in form_fields:
        sf = spec_by_name.get(f["name"])
        if sf and "domain" in f and isinstance(sf.get("domain"), dict):
            # the form has no portal domain name — adopt the spec's so only
            # coded-value drift surfaces, not a guaranteed name mismatch
            f["domain"]["name"] = sf["domain"].get("name",
                                                   f["domain"]["name"])
    report = diff_schema({"fields": form_fields}, layer_spec)
    return [
        SchemaChange(
            kind=item.drift_type.lower(),
            classification=_DRIFT_CLASS.get(item.drift_type, CLASS_REVIEW),
            name=item.field_name,
            detail=item.message,
        )
        for item in report.drift_items
    ]
