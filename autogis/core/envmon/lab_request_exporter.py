"""lab_request_exporter.py — lab analytical request workbook from sampling event plan."""
from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from ..common.qa import QACollector

_DEFAULT_COLUMNS = [
    "SampleID", "LocationID", "Matrix", "AnalyteGroup",
    "AnalyteList", "ContainerType", "Preservative",
    "HoldTimeDays", "TurnaroundDays", "ProjectCode",
    "CollectionDate", "Notes",
]


@dataclass
class LabRequestRow:
    sample_id: str
    location_id: str
    matrix: str
    analyte_group: str
    analyte_list: str
    container_type: str
    preservative: str
    hold_time_days: int
    turnaround_days: int
    project_code: str
    collection_date: str
    notes: str


@dataclass
class LabRequestResult:
    workbook_path: Path
    sample_count: int
    analyte_group_count: int
    qa: QACollector


def _parse_int(v) -> int:
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def build_lab_request_rows(
    plan_rows: list,
    analyte_groups: dict,
    *,
    project_code: str = "",
    turnaround_days: int = 5,
) -> list:
    rows = []
    for p in plan_rows:
        group_name = p.get("AnalyteGroup", "")
        group_cfg = analyte_groups.get(group_name, {})
        analytes = group_cfg.get("analytes", [])
        analyte_list = ", ".join(analytes)
        hold_days = _parse_int(p.get("HoldTimeDays") or group_cfg.get("hold_time_days", 0))
        rows.append(LabRequestRow(
            sample_id=p.get("SampleID", ""),
            location_id=p.get("LocationID", ""),
            matrix=p.get("Matrix", group_cfg.get("matrix", "")),
            analyte_group=group_name,
            analyte_list=analyte_list,
            container_type=p.get("Container", group_cfg.get("container", "")),
            preservative=p.get("Preservative", group_cfg.get("preservative", "")),
            hold_time_days=hold_days,
            turnaround_days=turnaround_days,
            project_code=project_code,
            collection_date=p.get("CollectionDate", ""),
            notes=p.get("Notes", ""),
        ))
    return rows


def write_lab_request_workbook(
    rows: list,
    out_path: Path,
    *,
    site_id: str = "",
    event_date: str = "",
    column_map: Optional[dict] = None,
) -> LabRequestResult:
    import openpyxl
    from openpyxl.styles import Font

    cm = column_map or {}
    headers = [cm.get(c, c) for c in _DEFAULT_COLUMNS]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sample Request"

    # Site header row — only written when site/date metadata is provided
    if site_id or event_date:
        ws.cell(1, 1, f"Site: {site_id}  Event Date: {event_date}")
        ws.cell(1, 1).font = Font(bold=True)

    # Column headers
    ws.append(headers)
    # Bold the header row (last appended row)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for r in rows:
        ws.append([
            r.sample_id, r.location_id, r.matrix, r.analyte_group,
            r.analyte_list, r.container_type, r.preservative,
            r.hold_time_days, r.turnaround_days, r.project_code,
            r.collection_date, r.notes,
        ])

    # Sheet 2: analyte groups expanded
    analyte_ws = wb.create_sheet("Analyte Groups")
    analyte_ws.append(["AnalyteGroup", "Analyte"])
    analyte_ws[1][0].font = Font(bold=True)
    analyte_ws[1][1].font = Font(bold=True)
    seen_groups: set = set()
    for r in rows:
        if r.analyte_group in seen_groups:
            continue
        seen_groups.add(r.analyte_group)
        for analyte in r.analyte_list.split(", "):
            if analyte.strip():
                analyte_ws.append([r.analyte_group, analyte.strip()])

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))

    group_count = len({r.analyte_group for r in rows})
    return LabRequestResult(
        workbook_path=out_path, sample_count=len(rows),
        analyte_group_count=group_count, qa=QACollector(),
    )


def write_lab_request_csv(rows: list, out_path: Path) -> None:
    fields = [
        "SampleID", "LocationID", "Matrix", "AnalyteGroup",
        "AnalyteList", "ContainerType", "Preservative",
        "HoldTimeDays", "TurnaroundDays", "ProjectCode",
        "CollectionDate", "Notes",
    ]
    with Path(out_path).open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fields)
        w.writeheader()
        for r in rows:
            w.writerow({
                "SampleID": r.sample_id, "LocationID": r.location_id,
                "Matrix": r.matrix, "AnalyteGroup": r.analyte_group,
                "AnalyteList": r.analyte_list, "ContainerType": r.container_type,
                "Preservative": r.preservative, "HoldTimeDays": r.hold_time_days,
                "TurnaroundDays": r.turnaround_days, "ProjectCode": r.project_code,
                "CollectionDate": r.collection_date, "Notes": r.notes,
            })
