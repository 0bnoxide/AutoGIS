"""ingest_reviewer_comments.py — headless ingest of reviewer map comments/redlines.

Parses review comment exports (CSV, GeoJSON FeatureCollection, XLSX) into a tracked
``ReviewerComment`` table with status lifecycle management.

No arcpy. No arcgis. Live AGOL fetch is NOT implemented here — the caller
exports to a supported format first (headless boundary).

Public API
----------
ReviewerComment          : dataclass (13 fields)
parse_comments_csv()     : parse flat CSV with flexible column aliases
parse_comments_geojson() : parse AGOL or generic GeoJSON FeatureCollection
parse_comments_xlsx()    : parse XLSX (openpyxl; prefers "Comments" sheet)
ingest_comments()        : auto-detect format by extension, emit QA records
merge_tracker()          : merge incoming into existing, preserving status
write_tracker_csv()      : write tracker to UTF-8 CSV
read_tracker_csv()       : read tracker CSV; returns [] if file absent
format_comment_summary() : human-readable status count + open comment preview
"""
from __future__ import annotations

import csv
import hashlib
import json
from collections import Counter
from dataclasses import dataclass, asdict, fields as _fields
from pathlib import Path
from typing import Optional

from autogis.core.common.qa import QACollector, SEV_INFO, SEV_WARNING, SEV_ERROR

# Status lifecycle: OPEN → IN_REVIEW → RESOLVED | WONT_FIX
VALID_STATUSES: set[str] = {"OPEN", "IN_REVIEW", "RESOLVED", "WONT_FIX"}
DEFAULT_STATUS: str = "OPEN"


@dataclass
class ReviewerComment:
    """A single reviewer markup comment with location and status tracking."""
    comment_id: str
    source_file: str
    source_format: str
    figure_ref: str
    comment_text: str
    reviewer: str
    status: str = DEFAULT_STATUS
    x: Optional[float] = None
    y: Optional[float] = None
    page_or_sheet: str = ""
    assigned_to: str = ""
    resolved_date: str = ""
    resolution_note: str = ""


_TRACKER_FIELDS: list[str] = [f.name for f in _fields(ReviewerComment)]


def _make_comment_id(
    source_file: str, figure_ref: str, comment_text: str, reviewer: str
) -> str:
    """Generate a stable, deterministic comment ID from key identifying fields.

    Note: ``comment_text`` is part of the hash, so re-exporting a comment whose
    wording was edited yields a *new* id and ``merge_tracker`` treats it as a new
    record (the prior one persists). To keep a stable identity across edits,
    supply a ``comment_id`` column in the source (the alias maps accept one) so
    downstream tracking systems own the id.
    """
    key = f"{source_file}|{figure_ref}|{comment_text}|{reviewer}"
    return "rc-" + hashlib.sha256(key.encode("utf-8")).hexdigest()[:12]


def _normalize_status(raw: str) -> str:
    """Normalize a raw status string to VALID_STATUSES, defaulting to OPEN."""
    upper = raw.strip().upper().replace(" ", "_").replace("-", "_")
    return upper if upper in VALID_STATUSES else DEFAULT_STATUS


def _to_float(v: object) -> Optional[float]:
    """Convert a value to float; return None for empty/None/non-numeric."""
    s = str(v).strip() if v is not None else ""
    try:
        return float(s) if s else None
    except ValueError:
        return None


_CSV_FIELD_MAP: dict[str, str] = {
    "figure": "figure_ref", "figure_ref": "figure_ref", "figure_name": "figure_ref",
    "comment": "comment_text", "comment_text": "comment_text",
    "reviewer": "reviewer", "author": "reviewer", "created_by": "reviewer",
    "status": "status",
    "x": "x", "longitude": "x", "lon": "x",
    "y": "y", "latitude": "y", "lat": "y",
    "page": "page_or_sheet", "page_or_sheet": "page_or_sheet", "sheet": "page_or_sheet",
    "assigned_to": "assigned_to", "assigned to": "assigned_to", "assignee": "assigned_to",
    "resolved_date": "resolved_date", "resolved date": "resolved_date",
    "resolution_note": "resolution_note", "resolution note": "resolution_note",
    "comment_id": "comment_id",
}

_GEOJSON_PROP_MAP: dict[str, str] = {
    "content": "comment_text", "comment": "comment_text", "comment_text": "comment_text",
    "author": "reviewer", "reviewer": "reviewer", "created_by": "reviewer",
    "figure": "figure_ref", "figure_ref": "figure_ref", "figure_name": "figure_ref",
    "status": "status",
    "assigned_to": "assigned_to", "assigned to": "assigned_to", "assignee": "assigned_to",
    "resolved_date": "resolved_date", "resolved date": "resolved_date",
    "resolution_note": "resolution_note", "resolution note": "resolution_note",
    "page": "page_or_sheet", "sheet": "page_or_sheet", "page_or_sheet": "page_or_sheet",
    "comment_id": "comment_id",
}

_XLSX_FIELD_MAP: dict[str, str] = {
    "figure": "figure_ref", "figure_ref": "figure_ref", "figure_name": "figure_ref",
    "comment": "comment_text", "comment_text": "comment_text",
    "reviewer": "reviewer", "author": "reviewer", "created_by": "reviewer",
    "status": "status",
    "x": "x", "longitude": "x", "lon": "x",
    "y": "y", "latitude": "y", "lat": "y",
    "page": "page_or_sheet", "sheet": "page_or_sheet", "page_or_sheet": "page_or_sheet",
    "assigned_to": "assigned_to", "assigned to": "assigned_to", "assignee": "assigned_to",
    "resolved_date": "resolved_date", "resolved date": "resolved_date",
    "resolution_note": "resolution_note", "resolution note": "resolution_note",
    "comment_id": "comment_id",
}


def parse_comments_csv(path: Path) -> list[ReviewerComment]:
    """Parse a flat reviewer comment CSV into ``ReviewerComment`` records.

    Accepts flexible column names (``_CSV_FIELD_MAP``); unmapped columns are
    ignored. UTF-8-sig safe (Excel BOM). Generates a deterministic
    ``comment_id`` from key fields when absent.
    """
    src = Path(path).name
    out: list[ReviewerComment] = []
    with Path(path).open(newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            mapped: dict[str, object] = {}
            for col, val in row.items():
                target = _CSV_FIELD_MAP.get((col or "").strip().lower())
                if target:
                    mapped[target] = (val or "").strip()
            existing_id = str(mapped.pop("comment_id", "")).strip()
            cid = existing_id or _make_comment_id(
                src, str(mapped.get("figure_ref", "")),
                str(mapped.get("comment_text", "")), str(mapped.get("reviewer", "")),
            )
            out.append(ReviewerComment(
                comment_id=cid, source_file=src, source_format="csv",
                figure_ref=str(mapped.get("figure_ref", "")),
                comment_text=str(mapped.get("comment_text", "")),
                reviewer=str(mapped.get("reviewer", "")),
                status=_normalize_status(str(mapped.get("status", ""))),
                x=_to_float(mapped.get("x")), y=_to_float(mapped.get("y")),
                page_or_sheet=str(mapped.get("page_or_sheet", "")),
                assigned_to=str(mapped.get("assigned_to", "")),
                resolved_date=str(mapped.get("resolved_date", "")),
                resolution_note=str(mapped.get("resolution_note", "")),
            ))
    return out


def write_tracker_csv(comments: list[ReviewerComment], path: Path) -> Path:
    """Write the full comment tracker to a UTF-8 CSV file (creates parent dirs)."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=_TRACKER_FIELDS)
        w.writeheader()
        for c in comments:
            w.writerow(asdict(c))
    return path


def read_tracker_csv(path: Path) -> list[ReviewerComment]:
    """Read a comment tracker CSV. Returns ``[]`` if the file does not exist."""
    p = Path(path)
    if not p.exists():
        return []
    out: list[ReviewerComment] = []
    with p.open(newline="", encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            out.append(ReviewerComment(
                comment_id=row.get("comment_id", ""),
                source_file=row.get("source_file", ""),
                source_format=row.get("source_format", "csv"),
                figure_ref=row.get("figure_ref", ""),
                comment_text=row.get("comment_text", ""),
                reviewer=row.get("reviewer", ""),
                status=_normalize_status(row.get("status", "")),
                x=_to_float(row.get("x")), y=_to_float(row.get("y")),
                page_or_sheet=row.get("page_or_sheet", ""),
                assigned_to=row.get("assigned_to", ""),
                resolved_date=row.get("resolved_date", ""),
                resolution_note=row.get("resolution_note", ""),
            ))
    return out


def parse_comments_geojson(path: Path, *, qa: QACollector) -> list[ReviewerComment]:
    """Parse an AGOL comment export or generic GeoJSON FeatureCollection.

    Each Point feature becomes one ``ReviewerComment``. Features without a Point
    geometry are ingested with ``x=None, y=None`` plus a QA WARNING. Returns
    ``[]`` with a QA ERROR if the root ``type`` is not ``"FeatureCollection"``.
    """
    src = Path(path).name
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    if data.get("type") != "FeatureCollection":
        qa.add(SEV_ERROR, "geojson_format",
               f"{src}: root type is not FeatureCollection — got {data.get('type')!r}",
               source_workbook=src)
        return []
    out: list[ReviewerComment] = []
    for i, feat in enumerate(data.get("features") or []):
        props = feat.get("properties") or {}
        geom = feat.get("geometry") or {}
        mapped: dict[str, str] = {}
        for k, v in props.items():
            target = _GEOJSON_PROP_MAP.get((k or "").lower().strip())
            if target:
                mapped[target] = str(v).strip() if v is not None else ""
        x = y = None
        if geom.get("type") == "Point":
            coords = geom.get("coordinates") or []
            if len(coords) >= 2:
                try:
                    x, y = float(coords[0]), float(coords[1])
                except (TypeError, ValueError):
                    qa.add(SEV_WARNING, "geojson_bad_coordinates",
                           f"{src}: feature {i} has non-numeric coordinates "
                           f"{coords[:2]!r} — x/y will be null",
                           source_workbook=src)
        else:
            qa.add(SEV_WARNING, "geojson_no_geometry",
                   f"{src}: feature {i} has no Point geometry — x/y will be null",
                   source_workbook=src)
        existing_id = mapped.pop("comment_id", "")
        cid = existing_id or _make_comment_id(
            src, mapped.get("figure_ref", ""),
            mapped.get("comment_text", ""), mapped.get("reviewer", ""),
        )
        out.append(ReviewerComment(
            comment_id=cid, source_file=src, source_format="geojson",
            figure_ref=mapped.get("figure_ref", ""),
            comment_text=mapped.get("comment_text", ""),
            reviewer=mapped.get("reviewer", ""),
            status=_normalize_status(mapped.get("status", "")),
            x=x, y=y,
            page_or_sheet=mapped.get("page_or_sheet", ""),
            assigned_to=mapped.get("assigned_to", ""),
            resolved_date=mapped.get("resolved_date", ""),
            resolution_note=mapped.get("resolution_note", ""),
        ))
    return out


def parse_comments_xlsx(
    path: Path, *, sheet: Optional[str] = None, qa: QACollector,
) -> list[ReviewerComment]:
    """Parse a reviewer comment spreadsheet with openpyxl.

    Sheet priority: ``sheet`` arg → a sheet named "Comments" → first sheet.
    Row 1 is the header; unmapped columns ignored; blank rows skipped.
    """
    import openpyxl
    src = Path(path).name
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet and sheet in wb.sheetnames:
        ws = wb[sheet]
    elif "Comments" in wb.sheetnames:
        ws = wb["Comments"]
    else:
        ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.values)
    if not all_rows:
        qa.add(SEV_WARNING, "xlsx_empty",
               f"{src}: selected sheet has no rows", source_workbook=src)
        return []
    header = [str(h).strip().lower() if h is not None else "" for h in all_rows[0]]
    out: list[ReviewerComment] = []
    for row_vals in all_rows[1:]:
        if all(v is None for v in row_vals):
            continue
        mapped: dict[str, object] = {}
        for col_idx, val in enumerate(row_vals):
            if col_idx < len(header):
                target = _XLSX_FIELD_MAP.get(header[col_idx])
                if target:
                    mapped[target] = str(val).strip() if val is not None else ""
        existing_id = str(mapped.pop("comment_id", "")).strip()
        cid = existing_id or _make_comment_id(
            src, str(mapped.get("figure_ref", "")),
            str(mapped.get("comment_text", "")), str(mapped.get("reviewer", "")),
        )
        out.append(ReviewerComment(
            comment_id=cid, source_file=src, source_format="xlsx",
            figure_ref=str(mapped.get("figure_ref", "")),
            comment_text=str(mapped.get("comment_text", "")),
            reviewer=str(mapped.get("reviewer", "")),
            status=_normalize_status(str(mapped.get("status", ""))),
            x=_to_float(mapped.get("x")), y=_to_float(mapped.get("y")),
            page_or_sheet=str(mapped.get("page_or_sheet", "")),
            assigned_to=str(mapped.get("assigned_to", "")),
            resolved_date=str(mapped.get("resolved_date", "")),
            resolution_note=str(mapped.get("resolution_note", "")),
        ))
    return out


def ingest_comments(path: Path, *, qa: QACollector) -> list[ReviewerComment]:
    """Ingest reviewer comments from ``path``, auto-detecting format by extension.

    Supported: .csv .geojson .json .xlsx. Emits SEV_ERROR and returns []
    for unrecognized extensions; SEV_WARNING per comment with empty
    comment_text/figure_ref; SEV_INFO "ingest_complete" on success.

    Legacy binary ``.xls`` is NOT supported — openpyxl cannot read it; convert
    to ``.xlsx`` first.
    """
    p = Path(path)
    ext = p.suffix.lower()
    if ext == ".csv":
        comments = parse_comments_csv(p)
    elif ext in (".geojson", ".json"):
        comments = parse_comments_geojson(p, qa=qa)
    elif ext == ".xlsx":
        comments = parse_comments_xlsx(p, qa=qa)
    else:
        detail = (" Legacy binary .xls is not readable by openpyxl; convert "
                  "to .xlsx." if ext == ".xls" else "")
        qa.add(SEV_ERROR, "unknown_format",
               f"Unsupported file extension {ext!r} for {p.name}. "
               f"Supported: .csv, .geojson, .json, .xlsx.{detail}",
               source_workbook=p.name)
        return []
    for c in comments:
        if not c.comment_text.strip():
            qa.add(SEV_WARNING, "empty_comment_text",
                   f"comment_id={c.comment_id}: empty comment_text",
                   source_workbook=c.source_file)
        if not c.figure_ref.strip():
            qa.add(SEV_WARNING, "empty_figure_ref",
                   f"comment_id={c.comment_id}: empty figure_ref",
                   source_workbook=c.source_file)
    qa.add(SEV_INFO, "ingest_complete",
           f"Ingested {len(comments)} comment(s) from {p.name} [{ext}]")
    return comments


def merge_tracker(
    existing: list[ReviewerComment],
    incoming: list[ReviewerComment],
    *,
    qa: QACollector,
) -> list[ReviewerComment]:
    """Merge ``incoming`` into ``existing``, preserving existing records by id.

    A ``comment_id`` already in ``existing`` is not updated (its status,
    assigned_to, resolved_date, resolution_note are preserved); new ids are
    appended; nothing is deleted. ``incoming`` is de-duplicated by
    ``comment_id`` (first occurrence wins) so a source export with repeated rows
    cannot inflate the tracker. Returns existing (original order) + new.
    """
    existing_ids = {c.comment_id for c in existing}
    new_comments: list[ReviewerComment] = []
    seen_new: set[str] = set()
    for c in incoming:
        if c.comment_id in existing_ids or c.comment_id in seen_new:
            continue
        seen_new.add(c.comment_id)
        new_comments.append(c)
    qa.add(SEV_INFO, "merge_complete",
           f"Merge: {len(existing)} existing + {len(incoming)} incoming "
           f"→ {len(new_comments)} new, {len(existing) + len(new_comments)} total")
    return list(existing) + new_comments


def format_comment_summary(comments: list[ReviewerComment]) -> str:
    """Return a human-readable status summary + a preview of up to 5 open comments."""
    counts = Counter(c.status for c in comments)
    lines = [f"Reviewer Comment Tracker — {len(comments)} total comment(s)", ""]
    for status in ("OPEN", "IN_REVIEW", "RESOLVED", "WONT_FIX"):
        lines.append(f"  {status:<12} {counts.get(status, 0):>4}")
    other = sum(v for k, v in counts.items() if k not in VALID_STATUSES)
    if other:
        lines.append(f"  OTHER        {other:>4}")
    lines.append("")
    open_comments = [c for c in comments if c.status == "OPEN"]
    if open_comments:
        lines.append(f"Open comments ({len(open_comments)}):")
        for c in open_comments[:5]:
            prefix = f"[{c.figure_ref}] " if c.figure_ref else ""
            preview = c.comment_text[:70]
            if len(c.comment_text) > 70:
                preview += "..."
            lines.append(f"  {prefix}{preview!r}  ({c.reviewer})")
        if len(open_comments) > 5:
            lines.append(f"  ... and {len(open_comments) - 5} more open comment(s).")
    return "\n".join(lines)
