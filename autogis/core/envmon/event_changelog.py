"""Generate structured event changelog from two results CSV sets (headless).

Diffs prior-event and current-event result rows keyed on (LocationID, AnalyteName)
and classifies each pair into one of eight ChangeType values.

No arcpy dependency. stdlib + openpyxl (lazy) for optional Excel output.
"""
from __future__ import annotations

import csv
import dataclasses
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from autogis.core.common.qa import QACollector, SEV_ERROR, SEV_INFO, SEV_WARNING
from autogis.core.envmon.canonical_read import canonical_result_rows


class ChangeType:
    NEW_LOCATION = "NEW_LOCATION"
    DROPPED_LOCATION = "DROPPED_LOCATION"
    NEW_ANALYTE = "NEW_ANALYTE"
    DROPPED_ANALYTE = "DROPPED_ANALYTE"
    NEW_EXCEEDANCE = "NEW_EXCEEDANCE"
    CLEARED_EXCEEDANCE = "CLEARED_EXCEEDANCE"
    VALUE_CHANGE = "VALUE_CHANGE"
    NO_CHANGE = "NO_CHANGE"


@dataclasses.dataclass
class ChangeRecord:
    location_id: str
    analyte_name: str
    change_type: str
    prior_value: Optional[float]
    current_value: Optional[float]
    prior_exceeds: Optional[bool]
    current_exceeds: Optional[bool]
    delta_pct: Optional[float]
    notes: str


@dataclasses.dataclass
class EventChangeResult:
    prior_event_id: str
    current_event_id: str
    changes: List[ChangeRecord]
    new_location_count: int
    dropped_location_count: int
    new_exceedance_count: int
    cleared_exceedance_count: int
    qa: QACollector


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _parse_float(v) -> Optional[float]:
    if v is None or str(v).strip() == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _parse_exceed(v) -> Optional[bool]:
    if v is None or str(v).strip() == "":
        return None
    try:
        return int(v) == 1
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Core logic
# ---------------------------------------------------------------------------

def generate_event_changelog(
    prior_rows: List[dict],
    current_rows: List[dict],
    *,
    prior_event_id: str = "prior",
    current_event_id: str = "current",
    delta_pct_threshold: float = 10.0,
) -> EventChangeResult:
    """Diff two result sets and classify each (LocationID, AnalyteName) pair.

    Args:
        prior_rows: List of dict rows from the prior event CSV.
        current_rows: List of dict rows from the current event CSV.
        prior_event_id: Label for the prior event (default ``"prior"``).
        current_event_id: Label for the current event (default ``"current"``).
        delta_pct_threshold: Minimum absolute percent change required to
            classify a row as VALUE_CHANGE rather than NO_CHANGE (default 10.0).

    Returns:
        EventChangeResult containing the classified ChangeRecord list and
        summary counts. The embedded ``qa`` collector holds any warnings.
    """
    qa = QACollector()

    # Canonical-read before the (LocationID, AnalyteName) diff maps: drop QC rows
    # and resolve fraction pairs so a Step-2 export carrying Total/Dissolved
    # splits or QC rows can't let an arbitrary row win the last-wins map and
    # corrupt the change classification (ADR-0075). No-op on legacy/pre-2.2 CSVs
    # (no QCType/ResultFraction columns -> nothing dropped or resolved).
    prior_rows = canonical_result_rows(prior_rows, qa)
    current_rows = canonical_result_rows(current_rows, qa)

    def _key(row: dict) -> Tuple[str, str]:
        return row.get("LocationID", ""), row.get("AnalyteName", "")

    prior_map: Dict[Tuple[str, str], dict] = {_key(r): r for r in prior_rows}
    current_map: Dict[Tuple[str, str], dict] = {_key(r): r for r in current_rows}

    prior_locs = {k[0] for k in prior_map}
    current_locs = {k[0] for k in current_map}
    new_locs = current_locs - prior_locs
    dropped_locs = prior_locs - current_locs

    all_keys = sorted(set(prior_map) | set(current_map))
    changes: List[ChangeRecord] = []

    for loc, analyte in all_keys:
        key = (loc, analyte)
        in_prior = key in prior_map
        in_current = key in current_map

        # --- Entirely new or dropped ---
        if not in_prior:
            cr = current_map[key]
            change_type = (
                ChangeType.NEW_LOCATION if loc in new_locs else ChangeType.NEW_ANALYTE
            )
            changes.append(ChangeRecord(
                location_id=loc,
                analyte_name=analyte,
                change_type=change_type,
                prior_value=None,
                current_value=_parse_float(cr.get("ResultNumeric")),
                prior_exceeds=None,
                current_exceeds=_parse_exceed(cr.get("ExceedsScreeningLevel")),
                delta_pct=None,
                notes="",
            ))
            continue

        if not in_current:
            pr = prior_map[key]
            change_type = (
                ChangeType.DROPPED_LOCATION if loc in dropped_locs
                else ChangeType.DROPPED_ANALYTE
            )
            changes.append(ChangeRecord(
                location_id=loc,
                analyte_name=analyte,
                change_type=change_type,
                prior_value=_parse_float(pr.get("ResultNumeric")),
                current_value=None,
                prior_exceeds=_parse_exceed(pr.get("ExceedsScreeningLevel")),
                current_exceeds=None,
                delta_pct=None,
                notes="",
            ))
            continue

        # --- Present in both — classify the change ---
        pr = prior_map[key]
        cr = current_map[key]

        prior_val = _parse_float(pr.get("ResultNumeric"))
        current_val = _parse_float(cr.get("ResultNumeric"))
        prior_exc = _parse_exceed(pr.get("ExceedsScreeningLevel"))
        current_exc = _parse_exceed(cr.get("ExceedsScreeningLevel"))

        # Compute delta_pct (used by exceedance branches too)
        delta_pct: Optional[float] = None
        if prior_val is not None and current_val is not None:
            if prior_val != 0:
                delta_pct = round((current_val - prior_val) / prior_val * 100, 2)
            else:
                qa.add(SEV_WARNING, "zero_prior_value",
                       f"{loc}/{analyte}: prior value is 0, delta_pct undefined")

        # Exceedance flips take priority over value-change classification
        if prior_exc is False and current_exc is True:
            change_type = ChangeType.NEW_EXCEEDANCE
        elif prior_exc is True and current_exc is False:
            change_type = ChangeType.CLEARED_EXCEEDANCE
        elif delta_pct is not None and abs(delta_pct) > delta_pct_threshold:
            change_type = ChangeType.VALUE_CHANGE
        else:
            change_type = ChangeType.NO_CHANGE

        changes.append(ChangeRecord(
            location_id=loc,
            analyte_name=analyte,
            change_type=change_type,
            prior_value=prior_val,
            current_value=current_val,
            prior_exceeds=prior_exc,
            current_exceeds=current_exc,
            delta_pct=delta_pct,
            notes="",
        ))

    new_loc_count = sum(1 for c in changes if c.change_type == ChangeType.NEW_LOCATION)
    dropped_loc_count = sum(1 for c in changes if c.change_type == ChangeType.DROPPED_LOCATION)
    new_exc_count = sum(1 for c in changes if c.change_type == ChangeType.NEW_EXCEEDANCE)
    cleared_exc_count = sum(1 for c in changes if c.change_type == ChangeType.CLEARED_EXCEEDANCE)

    qa.add(
        SEV_INFO, "changelog_complete",
        f"generate_event_changelog: {len(changes)} record(s) — "
        f"{new_loc_count} new location(s), {dropped_loc_count} dropped location(s), "
        f"{new_exc_count} new exceedance(s), {cleared_exc_count} cleared exceedance(s)",
    )
    return EventChangeResult(
        prior_event_id=prior_event_id,
        current_event_id=current_event_id,
        changes=changes,
        new_location_count=new_loc_count,
        dropped_location_count=dropped_loc_count,
        new_exceedance_count=new_exc_count,
        cleared_exceedance_count=cleared_exc_count,
        qa=qa,
    )


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

def write_changelog_csv(result: EventChangeResult, out_path: Path) -> None:
    """Write ChangeRecord list to a flat CSV."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fields = [f.name for f in dataclasses.fields(ChangeRecord)]
    with open(out_path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        for rec in result.changes:
            writer.writerow(dataclasses.asdict(rec))


def write_changelog_workbook(result: EventChangeResult, out_path: Path) -> None:
    """Write ChangeRecord list to Excel workbook — one sheet per change type.

    Sheets are created for every ChangeType constant, in declaration order.
    Empty sheets (no rows of that type) are still created so the workbook
    structure is predictable. Requires ``openpyxl`` (ADR-008).
    """
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError as exc:
        result.qa.add(SEV_ERROR, "openpyxl_missing",
                      f"openpyxl not installed: {exc}")
        return

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    fields = [f.name for f in dataclasses.fields(ChangeRecord)]
    change_type_order = [
        ChangeType.NEW_LOCATION,
        ChangeType.DROPPED_LOCATION,
        ChangeType.NEW_ANALYTE,
        ChangeType.DROPPED_ANALYTE,
        ChangeType.NEW_EXCEEDANCE,
        ChangeType.CLEARED_EXCEEDANCE,
        ChangeType.VALUE_CHANGE,
        ChangeType.NO_CHANGE,
    ]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    for ct in change_type_order:
        ws = wb.create_sheet(ct)
        ws.append(fields)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = ws["A2"]
        for rec in result.changes:
            if rec.change_type == ct:
                row_dict = dataclasses.asdict(rec)
                ws.append([row_dict.get(f) for f in fields])

    wb.save(out_path)
    result.qa.add(SEV_INFO, "workbook_written",
                  f"Changelog workbook written to {out_path} "
                  f"({len(result.changes)} record(s))")
