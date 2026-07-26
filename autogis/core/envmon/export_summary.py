"""Export Env_Samples + Env_AnalyticalResults to an Excel summary workbook.

Always: "All Results", "Detections", "Exceedances", "Summary by Analyte".
Added when provided: "Metadata" (site_id/event_id) and "Samples" (samples).

Headless: reads in-memory record lists, writes with openpyxl.  No arcpy.
"""
from __future__ import annotations

from collections import defaultdict
from dataclasses import asdict, fields as dc_fields
from pathlib import Path
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ..common.qa import QACollector
from .canonical_read import canonical_result_rows
from .gdb_schema import AnalyticalResultRecord, SampleRecord

_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _write_sheet(ws, rows: list, field_names: list) -> None:
    ws.append(field_names)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([row.get(f) for f in field_names])
    for i, _ in enumerate(field_names, 1):
        ws.column_dimensions[get_column_letter(i)].width = 14


def export_analytical_summary(
    samples: List[SampleRecord],
    results: List[AnalyticalResultRecord],
    output_path: Path,
    site_id: str,
    event_id: str = "",
    qa: Optional[QACollector] = None,
) -> Path:
    """Write an Excel summary workbook and return the written path.

    A "Metadata" sheet (SiteID/EventID) is added when either is non-empty,
    and a "Samples" sheet is added when *samples* is non-empty (issue #343:
    these were previously accepted but silently never written)."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    qa = qa or QACollector()

    result_fields = [f.name for f in dc_fields(AnalyticalResultRecord)]
    # Canonical-read policy: drop QC rows + resolve fraction pairs before the
    # analyte counts (and every sheet) so a summary never double-counts a
    # Total/Dissolved pair or tallies a QC blank as a detection (ADR-0075).
    all_rows = canonical_result_rows([asdict(r) for r in results], qa)
    det_rows = [r for r in all_rows if r.get("IsDetected")]
    exc_rows = [r for r in all_rows if r.get("ExceedsScreeningLevel")]

    counts: dict = defaultdict(lambda: defaultdict(int))
    for r in all_rows:
        a = r.get("AnalyteName", "")
        counts[a]["total"] += 1
        if r.get("IsDetected"):
            counts[a]["detected"] += 1
        if r.get("ExceedsScreeningLevel"):
            counts[a]["exceeds"] += 1
        if r.get("IsNonDetect"):
            counts[a]["nondetect"] += 1
    summary_fields = ["AnalyteName", "TotalCount", "DetectionCount",
                      "NonDetectCount", "ExceedanceCount"]
    summary_rows = [
        {"AnalyteName": a,
         "TotalCount": v["total"],
         "DetectionCount": v["detected"],
         "NonDetectCount": v["nondetect"],
         "ExceedanceCount": v["exceeds"]}
        for a, v in sorted(counts.items())]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if site_id or event_id:
        ws_meta = wb.create_sheet("Metadata")
        ws_meta.append(["Field", "Value"])
        for cell in ws_meta[1]:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
        if site_id:
            ws_meta.append(["SiteID", site_id])
        if event_id:
            ws_meta.append(["EventID", event_id])
        ws_meta.column_dimensions["A"].width = 14
        ws_meta.column_dimensions["B"].width = 24

    if samples:
        sample_fields = [f.name for f in dc_fields(SampleRecord)]
        ws_samples = wb.create_sheet("Samples")
        _write_sheet(ws_samples, [asdict(s) for s in samples], sample_fields)

    ws_all = wb.create_sheet("All Results")
    _write_sheet(ws_all, all_rows, result_fields)

    ws_det = wb.create_sheet("Detections")
    _write_sheet(ws_det, det_rows, result_fields)

    ws_exc = wb.create_sheet("Exceedances")
    _write_sheet(ws_exc, exc_rows, result_fields)

    ws_sum = wb.create_sheet("Summary by Analyte")
    _write_sheet(ws_sum, summary_rows, summary_fields)

    wb.save(output_path)
    return output_path
