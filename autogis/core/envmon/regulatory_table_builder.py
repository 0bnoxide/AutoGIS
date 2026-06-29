"""regulatory_table_builder.py — regulatory submission pivot table (openpyxl)."""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from ..common.qa import QACollector, QARecord, SEV_INFO, SEV_WARNING

ND_QUALIFIERS = frozenset({"ND", "U", "BDL"})
_FILL_EXCEED = PatternFill(fill_type="solid", fgColor="FF9999")
_FILL_HEADER = PatternFill(fill_type="solid", fgColor="D9E1F2")


@dataclass
class RegulatoryTableSpec:
    analyte_group: str
    analytes: list
    screening_levels: dict
    units: dict


@dataclass
class RegulatoryTableResult:
    workbook_path: Path
    group_count: int
    well_count: int
    exceedance_count: int
    qa: QACollector


def build_regulatory_table_specs(
    result_rows: list,
    group_map: Optional[dict] = None,
    screening_levels: Optional[dict] = None,
) -> list:
    gm = group_map or {}
    sl = screening_levels or {}
    by_group: dict[str, set] = defaultdict(set)
    units_map: dict[str, str] = {}
    for r in result_rows:
        analyte = r.get("AnalyteName", "")
        group = gm.get(analyte, "All Analytes")
        by_group[group].add(analyte)
        if r.get("ReportedUnits"):
            units_map[analyte] = r["ReportedUnits"]

    specs = []
    for group, analytes in sorted(by_group.items()):
        sorted_analytes = sorted(analytes)
        specs.append(RegulatoryTableSpec(
            analyte_group=group,
            analytes=sorted_analytes,
            screening_levels={a: sl[a] for a in sorted_analytes if a in sl},
            units={a: units_map.get(a, "ug/L") for a in sorted_analytes},
        ))
    return specs


def write_regulatory_workbook(
    result_rows: list,
    specs: list,
    out_path: Path,
    *,
    site_id: str = "",
    event_label: str = "",
    screening_levels: Optional[dict] = None,
    nd_text: str = "ND",
    exceed_marker: str = "**",
    qa: Optional[QACollector] = None,
) -> RegulatoryTableResult:
    if qa is None:
        qa = QACollector()
    sl = screening_levels or {}

    # Index: {(loc, analyte): row}
    data: dict[tuple, dict] = {}
    for r in result_rows:
        key = (r.get("LocationID", ""), r.get("AnalyteName", ""))
        data[key] = r

    all_wells = sorted({r.get("LocationID", "") for r in result_rows})
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    total_exceedances = 0

    for spec in specs:
        ws = wb.create_sheet(spec.analyte_group[:31])
        generated = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        # Row 1: site header
        ws.cell(1, 1, f"Site: {site_id}")
        ws.cell(1, 3, f"Event: {event_label}")
        ws.cell(1, 5, f"Generated: {generated}")

        # Row 2: analyte headers (offset from col 2)
        ws.cell(2, 1, "Location")
        for col_idx, analyte in enumerate(spec.analytes, start=2):
            cell = ws.cell(2, col_idx, analyte)
            cell.font = Font(bold=True)
            cell.fill = _FILL_HEADER

        # Row 3: MCL values
        ws.cell(3, 1, "MCL (ug/L)")
        ws.cell(3, 1).font = Font(italic=True)
        for col_idx, analyte in enumerate(spec.analytes, start=2):
            mcl = spec.screening_levels.get(analyte)
            ws.cell(3, col_idx, str(mcl) if mcl is not None else "—")

        # Data rows
        for row_idx, well in enumerate(all_wells, start=4):
            ws.cell(row_idx, 1, well)
            ws.cell(row_idx, 1).font = Font(bold=True)
            for col_idx, analyte in enumerate(spec.analytes, start=2):
                r = data.get((well, analyte))
                cell = ws.cell(row_idx, col_idx)
                if r is None:
                    cell.value = "—"
                else:
                    qual = r.get("ResultQualifier", "").upper().strip()
                    val_str = r.get("ResultValue", "")
                    if qual in ND_QUALIFIERS:
                        cell.value = nd_text
                    else:
                        try:
                            val = float(val_str)
                            mcl = sl.get(analyte) or spec.screening_levels.get(analyte)
                            if mcl and val > mcl:
                                cell.value = f"{val}{exceed_marker}"
                                cell.fill = _FILL_EXCEED
                                total_exceedances += 1
                            else:
                                cell.value = val
                        except (TypeError, ValueError):
                            cell.value = val_str

    # Footnotes sheet
    fn_ws = wb.create_sheet("Footnotes")
    fn_ws.cell(1, 1, "Symbol").font = Font(bold=True)
    fn_ws.cell(1, 2, "Meaning").font = Font(bold=True)
    fn_ws.cell(2, 1, exceed_marker)
    fn_ws.cell(2, 2, "Concentration exceeds the MCL/RSL")
    fn_ws.cell(3, 1, nd_text)
    fn_ws.cell(3, 2, "Not detected at or above the method detection limit")

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    qa.add(QARecord(SEV_INFO, "reg_tables_built",
                    f"{len(specs)} groups, {len(all_wells)} wells, "
                    f"{total_exceedances} exceedances → {out_path}"))

    return RegulatoryTableResult(
        workbook_path=out_path,
        group_count=len(specs),
        well_count=len(all_wells),
        exceedance_count=total_exceedances,
        qa=qa,
    )
