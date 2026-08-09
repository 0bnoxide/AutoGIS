"""Export normalized AnalyticalResultRecord lists to formatted Excel workbooks.

Headless: openpyxl only, no arcpy. Entry point: export_summary_tables().
"""
from __future__ import annotations

import datetime as _dt
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from ..common.logging import get_logger
from ..common.qa import QACollector, SEV_INFO, SEV_WARNING
from .gdb_schema import AnalyticalResultRecord
from .canonical_read import canonical_records

LOG = get_logger(__name__)

_RED_FILL = PatternFill(fill_type="solid", fgColor="FFCCCC")
_BOLD = Font(bold=True)
_MAX_COL_WIDTH = 40


def _date_str(d) -> str:
    if isinstance(d, (_dt.date, _dt.datetime)):
        return d.strftime("%Y-%m-%d")
    return str(d or "")


def _build_current_event(
    records: Sequence[AnalyticalResultRecord],
) -> Tuple[List[str], List[List], Set[Tuple[int, int]]]:
    """Rows = locations (latest sample date only), columns = analytes."""
    gw = [r for r in records if r.Matrix != "SOIL"]
    latest: Dict[str, str] = {}
    for r in gw:
        d = _date_str(r.SampleDate)
        if r.LocationID not in latest or d > latest[r.LocationID]:
            latest[r.LocationID] = d
    cells: Dict[Tuple[str, str], AnalyticalResultRecord] = {}
    for r in gw:
        if _date_str(r.SampleDate) == latest.get(r.LocationID):
            cells[(r.LocationID, r.AnalyteCanonicalName)] = r
    analytes = sorted({r.AnalyteCanonicalName for r in gw})
    locations = sorted(latest.keys())
    headers = ["LocationID", "SampleDate"] + analytes
    rows: List[List] = []
    exceedance_coords: Set[Tuple[int, int]] = set()
    for ri, loc in enumerate(locations):
        row: List = [loc, latest[loc]]
        for ci, an in enumerate(analytes):
            r = cells.get((loc, an))
            row.append(r.DisplayText if r else "--")
            if r and r.ExceedsScreeningLevel == 1:
                exceedance_coords.add((ri, ci + 2))
        rows.append(row)
    return headers, rows, exceedance_coords


def _build_gw_by_event(
    records: Sequence[AnalyticalResultRecord],
) -> Tuple[List[str], List[List], Set[Tuple[int, int]]]:
    """Stacked (location, analyte) rows × sorted event-date columns."""
    gw = [r for r in records if r.Matrix != "SOIL"]
    dates = sorted({_date_str(r.SampleDate) for r in gw})
    pairs = sorted({(r.LocationID, r.AnalyteCanonicalName) for r in gw})
    cells: Dict[Tuple[str, str, str], AnalyticalResultRecord] = {}
    for r in gw:
        cells[(r.LocationID, r.AnalyteCanonicalName, _date_str(r.SampleDate))] = r
    headers = ["Location", "Analyte"] + dates
    rows: List[List] = []
    exceedance_coords: Set[Tuple[int, int]] = set()
    for ri, (loc, an) in enumerate(pairs):
        row: List = [loc, an]
        for ci, d in enumerate(dates):
            r = cells.get((loc, an, d))
            row.append(r.DisplayText if r else "--")
            if r and r.ExceedsScreeningLevel == 1:
                exceedance_coords.add((ri, ci + 2))
        rows.append(row)
    return headers, rows, exceedance_coords


def _build_soil_by_depth(
    records: Sequence[AnalyticalResultRecord],
) -> Tuple[List[str], List[List], Set[Tuple[int, int]]]:
    """Rows = (location, depth interval), columns = analytes. SOIL matrix only."""
    soil = [r for r in records if r.Matrix == "SOIL"]
    analytes = sorted({r.AnalyteCanonicalName for r in soil})
    pairs = sorted({(r.LocationID, r.DepthIntervalText or "") for r in soil})
    cells: Dict[Tuple[str, str, str], AnalyticalResultRecord] = {}
    for r in soil:
        cells[(r.LocationID, r.DepthIntervalText or "", r.AnalyteCanonicalName)] = r
    headers = ["Location", "Depth"] + analytes
    rows: List[List] = []
    exceedance_coords: Set[Tuple[int, int]] = set()
    for ri, (loc, dep) in enumerate(pairs):
        row: List = [loc, dep]
        for ci, an in enumerate(analytes):
            r = cells.get((loc, dep, an))
            row.append(r.DisplayText if r else "--")
            if r and r.ExceedsScreeningLevel == 1:
                exceedance_coords.add((ri, ci + 2))
        rows.append(row)
    return headers, rows, exceedance_coords


def _apply_sheet_style(ws, exceedance_coords: Set[Tuple[int, int]]) -> None:
    """Bold header, freeze panes at A2, auto-column width, red exceedance cells.

    exceedance_coords are 0-based (row_idx, col_idx) into data rows.
    Worksheet conversion: ws_row = row_idx + 2  (data starts at row 2),
                          ws_col = col_idx + 1  (1-based columns).
    """
    for cell in ws[1]:
        cell.font = _BOLD
    ws.freeze_panes = "A2"
    for ri, ci in exceedance_coords:
        ws.cell(row=ri + 2, column=ci + 1).fill = _RED_FILL
    for col_cells in ws.columns:
        width = max(
            (len(str(c.value or "")) for c in col_cells),
            default=8,
        )
        ws.column_dimensions[col_cells[0].column_letter].width = min(
            width + 2, _MAX_COL_WIDTH
        )


def export_summary_tables(
    records: List[AnalyticalResultRecord],
    output_path: Path,
    *,
    site_id: str = "",
    include_current_event: bool = True,
    include_gw_by_event: bool = True,
    include_soil_by_depth: bool = True,
    qa: Optional[QACollector] = None,
) -> Path:
    """Write analytical summary sheets to one Excel workbook.

    Sheets (when enabled): "Current Event", "GW by Event", "Soil by Depth".
    Returns output_path.
    """
    qa = qa or QACollector()
    output_path = Path(output_path)
    if site_id:
        records = [r for r in records if r.SiteID == site_id]
    records = canonical_records(records, qa)

    builders = []
    if include_current_event:
        builders.append(("Current Event", _build_current_event))
    if include_gw_by_event:
        builders.append(("GW by Event", _build_gw_by_event))
    if include_soil_by_depth:
        if not any(r.Matrix == "SOIL" for r in records):
            qa.add(SEV_WARNING, "soil_records_absent",
                   "include_soil_by_depth=True but no SOIL records found.")
        builders.append(("Soil by Depth", _build_soil_by_depth))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)   # drop the default empty sheet
    total_rows = 0

    for sheet_name, builder in builders:
        headers, rows, exc_coords = builder(records)
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for row in rows:
            ws.append(row)
        if not rows:
            qa.add(SEV_WARNING, "no_records_for_sheet",
                   f"Sheet '{sheet_name}' has no data rows.")
        _apply_sheet_style(ws, exc_coords)
        total_rows += len(rows)

    wb.save(output_path)
    qa.add(SEV_INFO, "export_complete",
           f"Wrote {len(builders)} sheet(s) to {output_path}; "
           f"{total_rows} total data rows.")
    LOG.info("export_summary_tables: %s sheets, %s rows -> %s",
             len(builders), total_rows, output_path)
    return output_path
