"""report_appendix_builder.py — Tool 9.2 BuildMonitoringReportAppendix.

Build a multi-sheet Excel analytical-data appendix from long-format result
rows. One sheet per analyte group; rows = analytes, columns = (well, event
date). Detected cells are filled yellow, exceedances red, non-detects left
unfilled with text "ND". Two summary rows per sheet ("Detects" frequency and
"Max") sit below the data block.

Headless: openpyxl only, no arcpy. Input rows are raw ``csv.DictReader`` dicts
with keys LocationID, AnalyteName, ResultValue, ResultQualifier, ReportedUnits,
SampleDate (matching the sibling headless tools on this branch).
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from ..common.logging import get_logger
from ..common.qa import QACollector, SEV_INFO, SEV_WARNING

LOG = get_logger(__name__)

YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFFF99")
RED_FILL = PatternFill(fill_type="solid", fgColor="FF9999")
_BOLD = Font(bold=True)
_MAX_COL_WIDTH = 24
_DEFAULT_GROUP = "Analytical Results"
ND_QUALIFIERS = frozenset({"ND", "U", "BDL"})


@dataclass
class AppendixSheetSpec:
    sheet_name: str
    analyte_group: str
    analytes: List[str]
    screening_levels: Dict[str, float] = field(default_factory=dict)
    units: str = ""


@dataclass
class AppendixBuildResult:
    workbook_path: Path
    sheet_count: int
    well_count: int
    event_count: int
    qa: QACollector


def _parse_num(v) -> Optional[float]:
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _is_nondetect(row: dict) -> bool:
    qual = str(row.get("ResultQualifier", "")).upper().strip()
    if qual in ND_QUALIFIERS:
        return True
    return _parse_num(row.get("ResultValue", "")) is None


def build_appendix_sheet_specs(
    result_rows: List[dict],
    screening_levels: Optional[Dict[str, float]] = None,
    group_map: Optional[Dict[str, str]] = None,
) -> List[AppendixSheetSpec]:
    """Partition the analytes present in ``result_rows`` into one spec per group.

    ``group_map`` maps analyte -> group label. Analytes absent from the map (or
    when no map is given) fall into a single default group. Within a spec,
    analytes are sorted; groups are emitted in sorted order for determinism.
    """
    sl = dict(screening_levels or {})
    gmap = group_map or {}
    units_by_analyte: Dict[str, str] = {}
    groups: Dict[str, set] = {}
    for r in result_rows:
        analyte = r.get("AnalyteName", "")
        if not analyte:
            continue
        group = gmap.get(analyte, _DEFAULT_GROUP)
        groups.setdefault(group, set()).add(analyte)
        units_by_analyte.setdefault(analyte, r.get("ReportedUnits", ""))

    specs: List[AppendixSheetSpec] = []
    for group in sorted(groups):
        analytes = sorted(groups[group])
        units = next((units_by_analyte[a] for a in analytes
                      if units_by_analyte.get(a)), "")
        specs.append(AppendixSheetSpec(
            sheet_name=group[:31],  # Excel sheet-name limit
            analyte_group=group,
            analytes=analytes,
            screening_levels={a: sl[a] for a in analytes if a in sl},
            units=units,
        ))
    return specs


def _cell_text(row: Optional[dict], nd_qualifier: str) -> str:
    if row is None:
        return "--"
    if _is_nondetect(row):
        return nd_qualifier
    val = _parse_num(row.get("ResultValue", ""))
    return val if val is not None else nd_qualifier


def write_appendix_workbook(
    result_rows: List[dict],
    specs: List[AppendixSheetSpec],
    out_path: Path,
    *,
    site_id: str = "",
    event_dates: Optional[List[str]] = None,
    nd_qualifier: str = "ND",
    qa: Optional[QACollector] = None,
) -> AppendixBuildResult:
    """Build the multi-sheet appendix workbook. Returns AppendixBuildResult."""
    qa = qa or QACollector()
    out_path = Path(out_path)
    rows = result_rows
    if site_id:
        rows = [r for r in rows if r.get("SiteID", site_id) == site_id]

    wells = sorted({r.get("LocationID", "") for r in rows if r.get("LocationID")})
    all_dates = sorted({r.get("SampleDate", "") for r in rows if r.get("SampleDate")})
    dates = [d for d in (event_dates or all_dates) if d in all_dates] or all_dates
    # (well, date) column ordering: well-major, date within.
    columns = [(w, d) for w in wells for d in dates]

    # index: (analyte, well, date) -> row dict (last write wins)
    index: Dict[tuple, dict] = {}
    for r in rows:
        index[(r.get("AnalyteName", ""), r.get("LocationID", ""),
               r.get("SampleDate", ""))] = r

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    for spec in specs:
        ws = wb.create_sheet(spec.sheet_name)
        # Header row 1: well IDs spanning their date columns.
        h1 = ["", ""]
        for w in wells:
            h1.extend([w] + [""] * (len(dates) - 1))
        ws.append(h1)
        # Header row 2: Analyte / Units + a date per column.
        h2 = ["Analyte", "Units"] + [d for (_, d) in columns]
        ws.append(h2)
        for cell in ws[1] + ws[2]:
            cell.font = _BOLD

        for analyte in spec.analytes:
            sl = spec.screening_levels.get(analyte)
            data_row = [analyte, spec.units or ""]
            ws.append(data_row + [None] * len(columns))
            r_idx = ws.max_row
            for c_offset, (w, d) in enumerate(columns):
                src = index.get((analyte, w, d))
                cell = ws.cell(row=r_idx, column=3 + c_offset)
                cell.value = _cell_text(src, nd_qualifier)
                if src is None or _is_nondetect(src):
                    continue
                val = _parse_num(src.get("ResultValue", ""))
                if sl is not None and val is not None and val > sl:
                    cell.fill = RED_FILL
                else:
                    cell.fill = YELLOW_FILL

        # Summary rows (per column): detection frequency + max detected.
        detects = ["Detects", ""]
        maxes = ["Max", ""]
        for (w, d) in columns:
            present = [index.get((a, w, d)) for a in spec.analytes]
            present = [p for p in present if p is not None]
            n = len(present)
            det = [p for p in present if not _is_nondetect(p)]
            detects.append(f"{len(det)}/{n}" if n else "--")
            det_vals = [_parse_num(p.get("ResultValue", "")) for p in det]
            det_vals = [v for v in det_vals if v is not None]
            maxes.append(max(det_vals) if det_vals else nd_qualifier)
        ws.append(detects)
        ws.append(maxes)
        for cell in (ws[ws.max_row - 1] + ws[ws.max_row]):
            cell.font = _BOLD

        for col_cells in ws.columns:
            width = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[col_cells[0].column_letter].width = min(
                width + 2, _MAX_COL_WIDTH)

        if not spec.analytes:
            qa.add(SEV_WARNING, "empty_appendix_sheet",
                   f"Sheet '{spec.sheet_name}' has no analytes.")

    wb.save(out_path)
    qa.add(SEV_INFO, "appendix_built",
           f"Wrote {len(specs)} sheet(s), {len(wells)} well(s), "
           f"{len(dates)} event(s) to {out_path}.")
    LOG.info("report_appendix: %s sheets, %s wells, %s events -> %s",
             len(specs), len(wells), len(dates), out_path)
    return AppendixBuildResult(
        workbook_path=out_path, sheet_count=len(specs),
        well_count=len(wells), event_count=len(dates), qa=qa)
