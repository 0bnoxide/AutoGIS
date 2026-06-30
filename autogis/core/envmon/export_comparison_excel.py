"""Export ComparisonRecord CSV to a formatted Excel workbook (Tool 4.8).

Reads comparison records (output of Tool 4.7 ``compare-events``) and writes a
two-sheet workbook: ``AllResults`` (all records, TrendClass colour-coded) and
``Exceedances`` (rows where ``CurrentExceedance == "1"``).

Headless: openpyxl only (imported lazily), no arcpy.
"""
from __future__ import annotations

from pathlib import Path
from typing import List

from ..common.qa import QACollector, SEV_INFO, SEV_WARNING, SEV_ERROR

# Keys MUST match the TrendClass tokens emitted by compare_events._classify():
# INCREASED / DECREASED / STABLE / NEW_DETECTION / NO_LONGER_DETECTED /
# NONDETECT_BOTH / INDETERMINATE.
_TREND_HEX = {
    "INCREASED":          "FFCCCC",  # red tint
    "DECREASED":          "CCE5FF",  # blue tint
    "STABLE":             "CCFFCC",  # green tint
    "NEW_DETECTION":      "FFFF99",
    "NO_LONGER_DETECTED": "FFFF99",
    "NONDETECT_BOTH":     "FFFF99",
    "INDETERMINATE":      "FFFF99",
}

# compare_events._exc() emits CurrentExceedance as "Y"/"N"/""; accept a few
# lenient spellings so hand-edited CSVs still work.
_EXCEEDANCE_TRUE = frozenset({"Y", "YES", "1", "TRUE"})


def export_comparison_excel(
    records: List[dict],
    output_path: Path,
    *,
    overwrite: bool = False,
    qa: QACollector,
) -> Path:
    """Write comparison records to a two-sheet Excel workbook. Returns the path."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    output_path = Path(output_path)
    if output_path.exists() and not overwrite:
        qa.add(SEV_ERROR, "output_exists",
               f"Output already exists: {output_path}. Use overwrite=True.")
        return output_path

    if not records:
        qa.add(SEV_WARNING, "no_records", "No comparison records to export.")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    def _write_sheet(ws, rows, bold_header=True):
        if not rows:
            return
        headers = list(rows[0].keys())
        ws.append(headers)
        if bold_header:
            for cell in ws[1]:
                cell.font = Font(bold=True)
        # Use the string form, NOT ws["A2"] — accessing the cell object
        # materializes an empty row 2 and pushes appended data to row 3.
        ws.freeze_panes = "A2"
        trend_col = (headers.index("TrendClass") + 1
                     if "TrendClass" in headers else None)
        for row_dict in rows:
            ws.append([row_dict.get(h, "") for h in headers])
            row_idx = ws.max_row
            if trend_col:
                hex_col = _TREND_HEX.get(str(row_dict.get("TrendClass", "")))
                if hex_col:
                    ws.cell(row=row_idx, column=trend_col).fill = PatternFill(
                        fill_type="solid", fgColor=hex_col)

    ws_all = wb.active
    ws_all.title = "AllResults"
    _write_sheet(ws_all, records)

    ws_exc = wb.create_sheet("Exceedances")
    exceed_rows = [r for r in records
                   if str(r.get("CurrentExceedance", "")).strip().upper()
                   in _EXCEEDANCE_TRUE]
    _write_sheet(ws_exc, exceed_rows)

    wb.save(output_path)
    qa.add(SEV_INFO, "export_complete",
           f"Wrote {len(records)} row(s) to {output_path} "
           f"({len(exceed_rows)} exceedance(s))")
    return output_path
