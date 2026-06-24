"""Profile-driven Excel reading.

The workbook is opened twice: once with data_only=True (cached formula
results) and once with data_only=False (formula text). This is the only way
openpyxl can both use cached values and detect formula cells whose cached
value is missing — a mandatory QA condition (spec: Excel parsing
requirements). The source workbook is never modified.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterator, List, Optional, Protocol, runtime_checkable

import openpyxl
from openpyxl.utils import get_column_letter

from ..common.config import ParserProfile, SheetProfile
from ..common.qa import QACollector, SEV_ERROR, SEV_CRITICAL


@runtime_checkable
class WorkbookReaderProtocol(Protocol):
    """Interface every workbook reader must satisfy.

    ``ProfileWorkbookReader`` is the production adapter. Test code can
    provide ``InMemoryWorkbookReader`` (defined in ``tests/envmon/conftest.py``)
    to run normalizer logic without openpyxl or a real file.
    """
    path: Path
    date_system: str

    def sheet_names(self) -> List[str]: ...
    def require_sheet(self, sheet_profile: SheetProfile) -> bool: ...
    def cell(self, sheet_name: str, row: int, col: int) -> "CellValue": ...
    def header_text(self, sheet_profile: SheetProfile,
                    row: Optional[int], col: int) -> str: ...
    def iter_data_rows(self, sheet_profile: SheetProfile) -> Iterator[int]: ...


@dataclass
class CellValue:
    value: object            # cached/parsed value (never the formula text)
    raw_text: str            # str(value) before any coercion
    is_formula: bool
    formula_text: str
    cached_missing: bool     # formula present but no cached value
    row: int
    column: int
    ref: str                 # e.g. "F12"
    sheet: str


class ProfileWorkbookReader:
    """Read-only access to a workbook through a parser profile."""

    def __init__(self, workbook_path: Path, profile: ParserProfile,
                 qa: Optional[QACollector] = None):
        self.path = Path(workbook_path)
        self.profile = profile
        self.qa = qa or QACollector()
        if not self.path.exists():
            self.qa.add(SEV_CRITICAL, "workbook_not_found",
                        f"workbook could not be opened: {self.path}",
                        source_workbook=str(self.path))
            raise FileNotFoundError(self.path)
        try:
            self._wb_values = openpyxl.load_workbook(
                self.path, data_only=True, read_only=False)
            self._wb_formulas = openpyxl.load_workbook(
                self.path, data_only=False, read_only=False)
        except Exception as exc:
            self.qa.add(SEV_CRITICAL, "workbook_open_failed",
                        f"workbook could not be opened: {exc}",
                        source_workbook=str(self.path))
            raise
        # 1904 date-system detection
        self.date_system = "1904" if getattr(
            self._wb_values, "epoch", None) and "1904" in str(
            self._wb_values.epoch) else profile.date_system

    # ------------------------------------------------------------------
    def sheet_names(self) -> List[str]:
        return list(self._wb_values.sheetnames)

    def require_sheet(self, sheet_profile: SheetProfile) -> bool:
        if sheet_profile.sheet_name not in self._wb_values.sheetnames:
            self.qa.add(SEV_ERROR, "expected_sheet_missing",
                        f"expected sheet '{sheet_profile.sheet_name}' not found",
                        source_workbook=self.path.name,
                        recommended_action="update the parser profile sheet list")
            return False
        return True

    def cell(self, sheet_name: str, row: int, col: int) -> CellValue:
        ws_v = self._wb_values[sheet_name]
        ws_f = self._wb_formulas[sheet_name]
        v = ws_v.cell(row=row, column=col).value
        f = ws_f.cell(row=row, column=col).value
        is_formula = isinstance(f, str) and f.startswith("=")
        cached_missing = is_formula and v is None
        return CellValue(
            value=v, raw_text="" if v is None else str(v),
            is_formula=is_formula,
            formula_text=f if is_formula else "",
            cached_missing=cached_missing,
            row=row, column=col,
            ref=f"{get_column_letter(col)}{row}", sheet=sheet_name)

    def header_text(self, sheet_profile: SheetProfile, row: Optional[int],
                    col: int) -> str:
        """Header text for a column, walking left through merged regions."""
        if row is None:
            return ""
        ws = self._wb_values[sheet_profile.sheet_name]
        cell = ws.cell(row=row, column=col)
        if cell.value not in (None, ""):
            return str(cell.value).strip()
        # merged or visually merged header: take the anchor cell of any merged
        # range covering this position
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                anchor = ws.cell(row=rng.min_row, column=rng.min_col).value
                return "" if anchor is None else str(anchor).strip()
        return ""

    # ------------------------------------------------------------------
    def iter_data_rows(self, sheet_profile: SheetProfile) -> Iterator[int]:
        """Yield 1-based data row numbers per the end-detection strategy.

        Blank ID rows inside the data block are *skipped, not dropped
        silently*: the caller sees them via QA when appropriate; here they
        only count toward the consecutive-blank terminator.
        """
        if not self.require_sheet(sheet_profile):
            return
        ws = self._wb_values[sheet_profile.sheet_name]
        strategy = sheet_profile.data_end_detection_strategy
        max_row = ws.max_row
        if strategy == "max_row":
            limit_blanks = None
        elif strategy.startswith("consecutive_blank_ids"):
            limit_blanks = int(strategy.split(":")[1]) if ":" in strategy else 3
        else:
            limit_blanks = 3
        blanks = 0
        id_col = sheet_profile.id_column or 1
        for row in range(sheet_profile.data_start_row, max_row + 1):
            id_val = ws.cell(row=row, column=id_col).value
            if id_val in (None, ""):
                blanks += 1
                if limit_blanks is not None and blanks >= limit_blanks:
                    return
                continue
            blanks = 0
            yield row
