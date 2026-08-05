"""Groundwater normalization: combined analytical + water-level sheets.

Implements normalize_gw_table_2 (spec priority function). Water-level columns
(MPE/DTW/GWE) and analyte columns coexist on the same sheet; both streams are
extracted per row with full source-cell tracking.
"""
from __future__ import annotations

from typing import List, Tuple

from ..common.config import ParserProfile, SheetProfile
from .excel_profile_reader import ProfileWorkbookReader
from .gdb_schema import (AnalyticalResultRecord, SampleRecord,
                        WaterLevelRecord)
from ..common.qa import QACollector, SEV_ERROR, SEV_WARNING
from .result_parser import parse_excel_date, parse_result_value
from .table_normalizer import normalize_matrix_table

# DTW/GWE arithmetic mismatch tolerance, ft. Larger than rounding noise,
# smaller than any plausible transcription error.
_WL_TOLERANCE_FT = 0.02


def _is_non_data_water_level(rec: WaterLevelRecord) -> bool:
    """A footnote/annotation row masquerading as a water-level row: no valid
    date AND no water-level measurement (no numeric MPE/DTW/GWE and no
    recognized dry/NM/NS/NA status → MeasurementStatus 'UNKNOWN'). See
    ADR-0106."""
    return (rec.EventDate is None
            and rec.MeasurementStatus == "UNKNOWN"
            and rec.MonitoringPointElevation_ft is None
            and rec.DepthToWater_ft is None
            and rec.GroundwaterElevation_ft is None)


def _water_level_for_row(reader: ProfileWorkbookReader, sheet: SheetProfile,
                         row: int, site_id: str, batch_id: str,
                         qa: QACollector,
                         gwe_range: tuple | None) -> WaterLevelRecord:
    wb = reader.path.name
    loc = reader.cell(sheet.sheet_name, row, sheet.id_column).raw_text.strip()
    date_cell = reader.cell(sheet.sheet_name, row, sheet.date_column)
    event_date = parse_excel_date(date_cell.value, reader.date_system)

    def _num(col):
        if col is None:
            return None, None
        cell = reader.cell(sheet.sheet_name, row, col)
        if cell.cached_missing:
            qa.add(SEV_ERROR, "formula_cell_missing_cached_value",
                   f"water-level formula '{cell.formula_text}' has no cached "
                   f"value", site_id=site_id, location_id=loc,
                   import_batch_id=batch_id, source_workbook=wb,
                   source_sheet=sheet.sheet_name, source_row=row,
                   source_cell=cell.ref,
                   recommended_action="open/save workbook in Excel to "
                                      "recalculate, then re-import")
            return None, cell
        return parse_result_value(cell.value), cell

    mpe, mpe_cell = _num(sheet.mpe_column)
    dtw, dtw_cell = _num(sheet.dtw_column)
    gwe, gwe_cell = _num(sheet.gwe_column)

    gwe_raw = gwe_cell.raw_text if gwe_cell else ""
    status = "MEASURED"
    is_dry = is_measured = 0
    excl = ""
    gwe_val = None

    # Status can appear in either DTW or GWE cells on report tables.
    status_src = None
    for p in (gwe, dtw):
        if p is not None and p.status_code in ("DRY", "NOT_MEASURED",
                                               "NOT_SAMPLED", "NOT_ANALYZED"):
            status_src = p
            break
    if status_src is not None:
        status = {"DRY": "DRY", "NOT_MEASURED": "NM", "NOT_SAMPLED": "NS",
                  "NOT_ANALYZED": "NA"}[status_src.status_code]
        is_dry = int(status == "DRY")
        excl = f"status={status}"
        qa_cat = "dry_well" if is_dry else "well_not_measured"
        qa.add(SEV_WARNING if not is_dry else SEV_WARNING, qa_cat,
               f"{loc}: water level status {status}; excluded from contouring",
               site_id=site_id, location_id=loc, import_batch_id=batch_id,
               source_workbook=wb, source_sheet=sheet.sheet_name,
               source_row=row)
    elif gwe is not None and gwe.is_numeric:
        gwe_val = gwe.result_numeric
        is_measured = 1
        # arithmetic consistency check MPE - DTW = GWE
        if (mpe is not None and mpe.is_numeric and dtw is not None
                and dtw.is_numeric):
            calc = mpe.result_numeric - dtw.result_numeric
            if abs(calc - gwe_val) > _WL_TOLERANCE_FT:
                qa.add(SEV_WARNING, "water_level_formula_conflict",
                       f"{loc}: MPE-DTW={calc:.2f} but GWE={gwe_val:.2f}",
                       site_id=site_id, location_id=loc,
                       import_batch_id=batch_id, source_workbook=wb,
                       source_sheet=sheet.sheet_name, source_row=row,
                       source_cell=gwe_cell.ref if gwe_cell else "")
        if gwe_range and not (gwe_range[0] <= gwe_val <= gwe_range[1]):
            qa.add(SEV_ERROR, "groundwater_elevation_out_of_range",
                   f"{loc}: GWE {gwe_val} outside configured plausible range "
                   f"{gwe_range}", site_id=site_id, location_id=loc,
                   import_batch_id=batch_id, source_workbook=wb,
                   source_sheet=sheet.sheet_name, source_row=row)
            is_measured = 0
            status = "MEASURED"
            excl = "GWE outside configured range"
    else:
        status = "UNKNOWN"
        excl = "no numeric GWE and no recognized status"
        qa.add(SEV_WARNING, "result_value_unparsable",
               f"{loc}: groundwater elevation '{gwe_raw}' unparsable",
               site_id=site_id, location_id=loc, import_batch_id=batch_id,
               source_workbook=wb, source_sheet=sheet.sheet_name,
               source_row=row, source_cell=gwe_cell.ref if gwe_cell else "")

    use_for_contour = int(is_measured == 1 and gwe_val is not None
                          and not is_dry)
    if not use_for_contour and not excl:
        excl = f"status={status}"
    from openpyxl.utils import get_column_letter as _gcl
    return WaterLevelRecord(
        ImportBatchID=batch_id, SiteID=site_id, LocationID=loc,
        EventDate=event_date, SampleDateRaw=date_cell.raw_text,
        MonitoringPointElevation_ft=(mpe.result_numeric
                                     if mpe and mpe.is_numeric else None),
        DepthToWater_ft=(dtw.result_numeric
                         if dtw and dtw.is_numeric else None),
        GroundwaterElevation_ft=gwe_val,
        GroundwaterElevationRawText=gwe_raw,
        IsDry=is_dry, IsMeasured=is_measured, MeasurementStatus=status,
        UseForContour=use_for_contour, ExclusionReason=excl,
        SourceWorkbook=wb, SourceSheet=sheet.sheet_name, SourceRow=row,
        SourceColumn_MPE=_gcl(sheet.mpe_column) if sheet.mpe_column else "",
        SourceColumn_DTW=_gcl(sheet.dtw_column) if sheet.dtw_column else "",
        SourceColumn_GWE=_gcl(sheet.gwe_column) if sheet.gwe_column else "")


def normalize_gw_table_2(workbook_path, profile: ParserProfile, site_id: str,
                         batch_id: str, analyte_dictionary: dict,
                         screening_levels: dict, qa: QACollector,
                         reader: ProfileWorkbookReader | None = None,
                         plausible_gwe_range_ft=None
                         ) -> Tuple[List[WaterLevelRecord],
                                    List[SampleRecord],
                                    List[AnalyticalResultRecord]]:
    """Normalize all GW_ANALYTICAL_AND_WATER_LEVEL, GW_WATER_LEVEL_ONLY and
    GW_ANALYTICAL sheets in the profile. GW_ANALYTICAL sheets carry analyte
    columns but no MPE/DTW/GWE, so water-level extraction is skipped for them
    (running it would emit a spurious QA warning per row)."""
    reader = reader or ProfileWorkbookReader(workbook_path, profile, qa)
    water_levels: List[WaterLevelRecord] = []
    samples: List[SampleRecord] = []
    results: List[AnalyticalResultRecord] = []
    gwe_range = None
    # The site config owns this window, not the parser profile (#445): it
    # describes the site's AMSL elevation band, not a workbook's layout, and
    # `config_validation` requires+type-checks it on the site config while the
    # parser-profile template never carried it at all. Callers pass the site
    # config's value; a profile-only value is no longer consulted, so the two
    # surfaces cannot drift with the wrong one silently winning.
    rng = plausible_gwe_range_ft
    if rng and len(rng) == 2:
        gwe_range = (float(rng[0]), float(rng[1]))
    for sheet in (profile.sheets_of_type("GW_ANALYTICAL_AND_WATER_LEVEL")
                  + profile.sheets_of_type("GW_WATER_LEVEL_ONLY")
                  + profile.sheets_of_type("GW_ANALYTICAL")):
        if not reader.require_sheet(sheet):
            continue
        if sheet.data_type != "GW_ANALYTICAL":
            for row in reader.iter_data_rows(sheet):
                rec = _water_level_for_row(
                    reader, sheet, row, site_id, batch_id, qa, gwe_range)
                # A row with no valid date AND no water-level measurement (no
                # numeric MPE/DTW/GWE and no recognized status) is not data —
                # it is a footnote/annotation row whose label sits in the id
                # column (e.g. H272 'NOTES:'). Skip it with a visible QA
                # warning rather than import a bogus water level. See ADR-0106.
                if _is_non_data_water_level(rec):
                    qa.add(SEV_WARNING, "skipped_non_data_row",
                           f"row {row} ('{rec.LocationID or '(blank)'}') has no "
                           f"date and no water-level measurement; treated as a "
                           f"non-data (footnote/annotation) row and skipped",
                           site_id=site_id, location_id=rec.LocationID,
                           import_batch_id=batch_id,
                           source_workbook=reader.path.name,
                           source_sheet=sheet.sheet_name, source_row=row)
                    continue
                water_levels.append(rec)
        if sheet.analyte_columns:
            s, r = normalize_matrix_table(
                reader, sheet, matrix="GW", analytical_group="VPH_EPH_VOC",
                site_id=site_id, batch_id=batch_id,
                analyte_dictionary=analyte_dictionary,
                screening_levels=screening_levels, qa=qa)
            samples.extend(s)
            results.extend(r)
    return water_levels, samples, results
