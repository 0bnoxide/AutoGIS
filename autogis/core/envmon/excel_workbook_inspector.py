"""Workbook structure inspection (Tool 1 backend).

Produces a JSON/CSV structure report and can propose a *draft* parser
profile. The draft is a starting point that a human must review — heuristics
on report-formatted workbooks are never authoritative.
"""
from __future__ import annotations

import csv
import dataclasses
import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
from openpyxl.utils import get_column_letter

from ..common.config import col_index
from ..common.qa import QACollector, SEV_WARNING
from .result_parser import FORMULA_ERRORS, STATUS_CODES

_UNIT_PAT = re.compile(r"(ug/l|µg/l|mg/l|mg/kg|ug/kg|µg/kg|ppm|ppb|ft|feet|%)",
                       re.IGNORECASE)
_SCREEN_PAT = re.compile(r"(rbsl|screening|criteri|standard|mcl|cleanup|tier)",
                         re.IGNORECASE)
_DATE_HDR = re.compile(r"\bdate\b", re.IGNORECASE)
_ID_HDR = re.compile(r"(location|well|sample|boring)\s*(id|no\.?|number)?",
                     re.IGNORECASE)


@dataclass
class SheetInspectionInfo:
    """Typed result of inspecting one worksheet."""
    used_range: str
    max_row: Optional[int]
    max_col: Optional[int]
    merged_ranges: List[str] = field(default_factory=list)
    row_classes: Dict[int, str] = field(default_factory=dict)
    title_rows: List[int] = field(default_factory=list)
    header_rows: List[int] = field(default_factory=list)
    units_rows: List[int] = field(default_factory=list)
    screening_rows: List[int] = field(default_factory=list)
    blank_rows: List[int] = field(default_factory=list)
    data_start_row: Optional[int] = None
    date_columns: List[str] = field(default_factory=list)
    id_columns: List[str] = field(default_factory=list)
    formula_cells: int = 0
    formula_cells_missing_cached: List[str] = field(default_factory=list)
    formula_errors: List[str] = field(default_factory=list)
    nonnumeric_tokens: List[str] = field(default_factory=list)
    unsafe_field_names: List[str] = field(default_factory=list)


@dataclass
class WorkbookInspectionReport:
    """Typed result of inspecting a whole workbook."""
    workbook: str
    workbook_path: str
    inspected: str
    date_system: str
    sheet_names: List[str]
    sheets: Dict[str, SheetInspectionInfo] = field(default_factory=dict)


def _row_values(ws, row: int, max_col: int) -> List:
    return [ws.cell(row=row, column=c).value for c in range(1, max_col + 1)]


def _classify_row(values: List) -> str:
    non_blank = [v for v in values if v not in (None, "")]
    if not non_blank:
        return "blank"
    texts = [str(v) for v in non_blank]
    joined = " ".join(texts)
    if any(_UNIT_PAT.fullmatch(t.strip()) or _UNIT_PAT.search(t)
           and len(t) < 12 for t in texts) and len(non_blank) > 1:
        return "units"
    if _SCREEN_PAT.search(joined):
        return "screening_level"
    n_num = sum(isinstance(v, (int, float)) for v in non_blank)
    n_date = sum(isinstance(v, (datetime, date)) for v in non_blank)
    if n_date >= 1 and n_num >= 1:
        return "data"
    if n_num >= max(2, len(non_blank) // 2):
        return "data_or_screening"
    if len(non_blank) == 1 and len(texts[0]) > 20:
        return "title"
    if len(non_blank) >= 3 and n_num == 0:
        return "header"
    return "text"


def _guess_analyte_columns(info: SheetInspectionInfo) -> Optional[dict]:
    """Return {from, to} column range for likely analyte columns.

    Heuristic: analyte columns are columns not claimed as id or date columns,
    from the first such column through max_col. When none are detected, column 1
    (A) is reserved to match propose_parser_profile's id_column fallback.
    Returns None when the sheet has too few columns to make a guess.
    """
    max_col = info.max_col or 0
    if max_col < 2:
        return None
    special: set = set()
    for letter in info.id_columns + info.date_columns:
        try:
            idx = col_index(letter)
            if idx is not None:
                special.add(idx)
        except Exception:
            pass
    if not special:
        special.add(1)
    first_analyte = next(
        (c for c in range(1, max_col + 1) if c not in special), None
    )
    if first_analyte is None:
        return None
    return {"from": get_column_letter(first_analyte),
            "to": get_column_letter(max_col)}


def inspect_workbook_structure(workbook_path: Path,
                               qa: Optional[QACollector] = None,
                               scan_rows: int = 40) -> WorkbookInspectionReport:
    qa = qa or QACollector()
    path = Path(workbook_path)
    wb_v = openpyxl.load_workbook(path, data_only=True)
    wb_f = openpyxl.load_workbook(path, data_only=False)
    date_system = "1904" if "1904" in str(getattr(wb_v, "epoch", "")) else "1900"
    sheets: Dict[str, SheetInspectionInfo] = {}
    for name in wb_v.sheetnames:
        ws, wsf = wb_v[name], wb_f[name]
        if not hasattr(ws, "max_row"):
            # Chartsheet (and any other non-data sheet type) — a saved chart
            # tab has no cell grid to inspect.
            qa.add(SEV_WARNING, "non_data_sheet_skipped",
                   f"sheet '{name}' is a {type(ws).__name__}, not a data "
                   f"worksheet (e.g. a saved chart) — skipped.",
                   source_sheet=name, source_workbook=path.name)
            continue
        max_row, max_col = ws.max_row, ws.max_column
        used_range = (f"A1:{get_column_letter(max_col)}{max_row}"
                      if max_col else "A1")
        info = SheetInspectionInfo(
            used_range=used_range,
            max_row=max_row,
            max_col=max_col,
            merged_ranges=[str(r) for r in ws.merged_cells.ranges],
        )
        scan_to = min(scan_rows, max_row) if max_row else 0
        for r in range(1, scan_to + 1):
            cls = _classify_row(_row_values(ws, r, max_col or 0))
            info.row_classes[r] = cls
            if cls == "blank":
                info.blank_rows.append(r)
            elif cls == "title":
                info.title_rows.append(r)
            elif cls == "header":
                info.header_rows.append(r)
            elif cls == "units":
                info.units_rows.append(r)
            elif cls == "screening_level":
                info.screening_rows.append(r)
            elif cls == "data" and info.data_start_row is None:
                info.data_start_row = r
        # column scans on the data block
        start = info.data_start_row or 1
        for c in range(1, (max_col or 0) + 1):
            col_dates = sum(isinstance(ws.cell(row=r, column=c).value,
                                       (datetime, date))
                            for r in range(start, min(start + 15, (max_row or 0) + 1)))
            if col_dates >= 2:
                info.date_columns.append(get_column_letter(c))
            hdr_texts = " ".join(str(ws.cell(row=r, column=c).value or "")
                                 for r in range(1, start))
            if _DATE_HDR.search(hdr_texts) and get_column_letter(c) not in info.date_columns:
                info.date_columns.append(get_column_letter(c))
            if _ID_HDR.search(hdr_texts):
                info.id_columns.append(get_column_letter(c))
            if hdr_texts.strip():
                last_hdr = hdr_texts.strip().split("  ")[-1]
                if re.search(r"[^A-Za-z0-9_]", last_hdr.replace(" ", "_")):
                    info.unsafe_field_names.append(last_hdr[:60])
        # whole-sheet formula / token scan
        nonnumeric: set = set()
        for r in range(1, (max_row or 0) + 1):
            for c in range(1, (max_col or 0) + 1):
                f = wsf.cell(row=r, column=c).value
                v = ws.cell(row=r, column=c).value
                ref = f"{get_column_letter(c)}{r}"
                if isinstance(f, str) and f.startswith("="):
                    info.formula_cells += 1
                    if v is None:
                        info.formula_cells_missing_cached.append(ref)
                if isinstance(v, str):
                    sv = v.strip()
                    if sv.upper() in FORMULA_ERRORS:
                        info.formula_errors.append(ref)
                    elif sv.upper() in STATUS_CODES:
                        nonnumeric.add(sv)
        info.nonnumeric_tokens = sorted(nonnumeric)
        sheets[name] = info
        # QA
        if len(info.header_rows) > 2:
            qa.add(SEV_WARNING, "ambiguous_header_row",
                   f"sheet '{name}': multiple candidate header rows "
                   f"{info.header_rows}", source_sheet=name,
                   source_workbook=path.name,
                   recommended_action="confirm analyte_header_row in the parser profile")
        if info.formula_cells_missing_cached:
            qa.add(SEV_WARNING, "formula_cell_missing_cached_value",
                   f"sheet '{name}': {len(info.formula_cells_missing_cached)} "
                   f"formula cells have no cached value "
                   f"(e.g. {info.formula_cells_missing_cached[:5]})",
                   source_sheet=name, source_workbook=path.name,
                   recommended_action="open and save the workbook in Excel to "
                                      "recalculate, then re-import")
        if info.data_start_row is None:
            qa.add(SEV_WARNING, "nonstandard_layout",
                   f"sheet '{name}': no data row detected in first "
                   f"{scan_to} rows", source_sheet=name,
                   source_workbook=path.name)
    return WorkbookInspectionReport(
        workbook=path.name,
        workbook_path=str(path),
        inspected=datetime.now().isoformat(timespec="seconds"),
        date_system=date_system,
        sheet_names=list(wb_v.sheetnames),
        sheets=sheets,
    )


def propose_parser_profile(report: WorkbookInspectionReport,
                           profile_id: str = "DRAFT") -> dict:
    """Draft a parser profile from an inspection report. Marked DRAFT;
    a human must verify row/column anchors before import."""
    sheets = []
    for name, info in report.sheets.items():
        sheets.append({
            "sheet_name": name,
            "data_type": "UNKNOWN_REVIEW_REQUIRED",
            "title_rows": info.title_rows,
            "group_header_row": (info.header_rows[0]
                                 if len(info.header_rows) > 1 else None),
            "analyte_header_row": (info.header_rows[-1]
                                   if info.header_rows else None),
            "units_row": info.units_rows[0] if info.units_rows else None,
            "screening_level_row": (info.screening_rows[0]
                                    if info.screening_rows else None),
            "data_start_row": info.data_start_row or 2,
            "id_column": info.id_columns[0] if info.id_columns else "A",
            "date_column": info.date_columns[0] if info.date_columns else None,
            "analyte_columns": _guess_analyte_columns(info),
            "_TODO": "REVIEW EVERY ANCHOR — heuristic draft, not validated",
        })
    return {"profile_id": profile_id,
            "profile_description": "DRAFT auto-proposed from inspection — "
                                   "review before use",
            "workbook_type": "UNKNOWN",
            "date_system": report.date_system,
            "formula_handling": "use_cached_then_qa",
            "sheets": sheets}


def write_inspection_outputs(report: WorkbookInspectionReport, json_path: Path,
                             csv_path: Optional[Path] = None) -> None:
    json_path = Path(json_path)
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(
        json.dumps(dataclasses.asdict(report), indent=2, default=str),
        encoding="utf-8")
    if csv_path:
        csv_path = Path(csv_path)
        csv_path.parent.mkdir(parents=True, exist_ok=True)
        with csv_path.open("w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["sheet", "item", "value"])
            for name, info in report.sheets.items():
                info_dict = dataclasses.asdict(info)
                for key in ("used_range", "title_rows", "header_rows",
                            "units_rows", "screening_rows", "data_start_row",
                            "date_columns", "id_columns", "formula_cells",
                            "formula_cells_missing_cached", "formula_errors",
                            "nonnumeric_tokens", "merged_ranges",
                            "unsafe_field_names"):
                    w.writerow([name, key, json.dumps(info_dict[key], default=str)])
