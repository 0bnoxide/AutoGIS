"""qc_sample_summary.py — QC sample type classifier + per-type summary workbook."""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill

from ..common.qa import QACollector, QARecord, SEV_INFO, SEV_WARNING
from .sample_id import QC_SUFFIXES as _SUFFIX_MAP

QC_TYPES = (
    "method_blank", "field_blank", "trip_blank",
    "matrix_spike", "matrix_spike_duplicate",
    "lab_duplicate", "field_duplicate",
    "primary",
)

_QC_TYPE_SET = frozenset(QC_TYPES)
ND_QUALIFIERS = frozenset({"ND", "U", "BDL"})

_FILL_FAIL = PatternFill(fill_type="solid", fgColor="FF9999")
_FILL_WARN = PatternFill(fill_type="solid", fgColor="FFFF99")


@dataclass
class QCRecord:
    sample_id: str
    qc_type: str
    location_id: str
    analyte_name: str
    result_value: Optional[float]
    qualifier: str
    units: str
    sample_date: str
    rpd: Optional[float]
    pct_recovery: Optional[float]
    pass_fail: str


@dataclass
class QCSummaryResult:
    records: list
    blank_detections: int
    spike_failures: int
    duplicate_failures: int
    qa: QACollector


def _infer_qc_type(sample_id: str, declared_type: str) -> str:
    if declared_type and declared_type.lower() in _QC_TYPE_SET:
        return declared_type.lower()
    sid_lower = sample_id.lower()
    for suffix, qtype in sorted(_SUFFIX_MAP.items(), key=lambda x: -len(x[0])):
        if sid_lower.endswith(suffix):
            return qtype
    return "primary"


def _parse_float(v: str) -> Optional[float]:
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def compute_rpd(val1: float, val2: float) -> float:
    denom = (val1 + val2) / 2
    if denom == 0:
        return 0.0
    return abs(val1 - val2) / denom * 100


def classify_qc_rows(
    result_rows: list,
    *,
    qc_type_field: str = "SampleType",
    sample_id_field: str = "SampleID",
) -> list:
    records = []
    for r in result_rows:
        sid = r.get(sample_id_field, "")
        declared = r.get(qc_type_field, "")
        qc_type = _infer_qc_type(sid, declared)
        if qc_type == "primary":
            continue
        qual = r.get("ResultQualifier", "").upper().strip()
        val = None if qual in ND_QUALIFIERS else _parse_float(r.get("ResultValue", ""))
        records.append(QCRecord(
            sample_id=sid, qc_type=qc_type,
            location_id=r.get("LocationID", ""),
            analyte_name=r.get("AnalyteName", ""),
            result_value=val, qualifier=qual,
            units=r.get("ReportedUnits", ""),
            sample_date=r.get("SampleDate", ""),
            rpd=None, pct_recovery=None, pass_fail="na",
        ))
    return records


def write_qc_summary_workbook(
    result: QCSummaryResult,
    out_path: Path,
) -> Path:
    # ponytail: RPD/recovery pass-fail scoring is deferred — it needs duplicate
    # pairing and spike-amount data the current model does not carry. The
    # building block compute_rpd() is kept and tested for when that lands.
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    by_type: dict[str, list] = {}
    for r in result.records:
        by_type.setdefault(r.qc_type, []).append(r)

    _headers = ["SampleID", "AnalyteName", "ResultValue", "Qualifier",
                "Units", "SampleDate", "RPD/Recovery", "Pass/Fail"]

    for qc_type, records in sorted(by_type.items()):
        ws = wb.create_sheet(qc_type[:31])
        ws.append(_headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for rec in records:
            rpd_val = rec.rpd
            pf = rec.pass_fail
            ws.append([
                rec.sample_id, rec.analyte_name,
                rec.result_value if rec.result_value is not None else rec.qualifier,
                rec.qualifier, rec.units, rec.sample_date,
                f"{rpd_val:.1f}%" if rpd_val is not None else "—", pf,
            ])

    # Summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["QC Type", "Count", "Blank Detections",
                        "Spike Failures", "Duplicate Failures"])
    summary_ws.append(["All QC", len(result.records),
                        result.blank_detections, result.spike_failures,
                        result.duplicate_failures])
    for qc_type, records in sorted(by_type.items()):
        summary_ws.append([qc_type, len(records), "", "", ""])

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path
