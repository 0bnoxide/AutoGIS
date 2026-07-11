"""well_inspection_photo_report.py — Tool 7.4 GenerateWellInspectionPhotoReport.

Assembles a per-well inspection photo workbook (XLSX, openpyxl) from two
inputs:

- the attachment harvester's ``manifest.csv``/``.json`` (read via
  ``index_field_attachments.load_manifest``) — the photo inventory; and
- a user-supplied inspection metadata CSV whose headers align with the
  sibling ``well_inspection_report.py`` maintenance-log schema:
  ``WellID, InspectionDate, Inspector, Condition, Notes`` plus optional
  ``GPS_Lat, GPS_Lon, DepthToWaterFt``. GPS must come from this CSV — the
  harvest manifest's ``geometry`` column is reserved and always empty.

PILOT ASSUMPTION (isolated in ``match_photos_to_wells``): the harvest run's
``group_template`` rendered a well ID, so every ``saved_path`` follows
``{harvest_dir}/{well_id}/{filename}``. If a site's group_template is not
well-id-based, photo group keys won't match inspection WellIDs — the
``wells_without_photos`` / ``photos_without_record`` QA WARNINGs are the
intended tripwire for that mismatch, not a silent failure.

Pillow is required only to embed photos (openpyxl's image support needs it)
and is lazy-imported inside the write path, so this module imports with
neither arcpy, arcgis, nor Pillow installed. Install: pip install
"autogis[report]".

No arcpy. No arcgis.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from autogis.core.common.qa import QACollector, SEV_INFO, SEV_WARNING

_USABLE_DISPOSITIONS = {"downloaded", "skipped"}
_PILLOW_HINT = ('Pillow is required to embed photos (openpyxl image '
                'support). Install with: pip install "autogis[report]"')


@dataclass
class PhotoInspectionRecord:
    """One well's inspection metadata (internal record — snake_case per
    ADR-0038; the CSV headers are PascalCase for schema alignment with
    well_inspection_report.py's maintenance log)."""
    well_id: str
    inspection_date: str
    inspector: str = ""
    condition: str = ""
    notes: str = ""
    gps_lat: Optional[float] = None
    gps_lon: Optional[float] = None
    depth_to_water_ft: Optional[float] = None


@dataclass
class PhotoReportResult:
    workbook_path: Path
    well_count: int
    photos_embedded: int
    photos_missing: int


def _float_or_none(value, *, field: str, row_no: int,
                   qa: Optional[QACollector]) -> Optional[float]:
    s = (value or "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        if qa is not None:
            qa.add(SEV_WARNING, "bad_number",
                   f"inspections row {row_no}: {field}={s!r} is not a "
                   f"number — treated as blank.")
        return None


def load_inspection_records(
    path: Path, *, qa: Optional[QACollector] = None,
) -> list[PhotoInspectionRecord]:
    """Load the inspection metadata CSV (headers documented in the module
    docstring). Rows with a blank WellID are skipped with a QA WARNING."""
    records: list[PhotoInspectionRecord] = []
    with Path(path).open(newline="", encoding="utf-8-sig") as fh:
        for row_no, row in enumerate(csv.DictReader(fh), start=2):
            well_id = (row.get("WellID") or "").strip()
            if not well_id:
                if qa is not None:
                    qa.add(SEV_WARNING, "missing_well_id",
                           f"inspections row {row_no}: blank WellID — "
                           f"row skipped.")
                continue
            records.append(PhotoInspectionRecord(
                well_id=well_id,
                inspection_date=(row.get("InspectionDate") or "").strip(),
                inspector=(row.get("Inspector") or "").strip(),
                condition=(row.get("Condition") or "").strip(),
                notes=(row.get("Notes") or "").strip(),
                gps_lat=_float_or_none(row.get("GPS_Lat"), field="GPS_Lat",
                                       row_no=row_no, qa=qa),
                gps_lon=_float_or_none(row.get("GPS_Lon"), field="GPS_Lon",
                                       row_no=row_no, qa=qa),
                depth_to_water_ft=_float_or_none(
                    row.get("DepthToWaterFt"), field="DepthToWaterFt",
                    row_no=row_no, qa=qa),
            ))
    return records


def match_photos_to_wells(
    manifest_rows: list[dict], harvest_dir: Path, *,
    qa: Optional[QACollector] = None,
) -> dict[str, list[dict]]:
    """Group usable harvester-manifest rows by well ID.

    PILOT ASSUMPTION lives here: the well ID is the first path component of
    ``saved_path`` relative to *harvest_dir* (i.e. the harvest
    ``group_template`` rendered a well ID). Rows whose effective disposition
    (``disposition`` falling back to ``status``) is not downloaded/skipped,
    or whose saved_path is blank, are excluded; rows not under *harvest_dir*
    are dropped with one aggregated QA WARNING.
    """
    harvest_dir = Path(harvest_dir)
    photo_map: dict[str, list[dict]] = {}
    outside = 0
    for row in manifest_rows:
        eff = ((row.get("disposition") or row.get("status") or "").strip())
        saved = (row.get("saved_path") or "").strip()
        if eff not in _USABLE_DISPOSITIONS or not saved:
            continue
        try:
            well_id = Path(saved).relative_to(harvest_dir).parts[0]
        except (ValueError, IndexError):
            outside += 1
            continue
        photo_map.setdefault(well_id, []).append(row)
    if outside and qa is not None:
        qa.add(SEV_WARNING, "manifest_rows_outside_harvest_dir",
               f"{outside} manifest row(s) have a saved_path outside "
               f"{harvest_dir} — dropped. Wrong --harvest-dir, or the "
               f"manifest came from another machine?")
    for rows in photo_map.values():
        rows.sort(key=lambda r: r.get("saved_path") or "")
    return photo_map


def prepare_image_bytes(path: Path, box: tuple[int, int]) -> Optional[bytes]:
    """EXIF-corrected, RGB, thumbnail-to-*box* JPEG bytes; None if the file
    is missing on disk. Pillow is lazy-imported here (the only place)."""
    p = Path(path)
    if not p.is_file():
        return None
    try:
        from PIL import Image, ImageOps
    except ImportError as exc:
        raise ImportError(_PILLOW_HINT) from exc
    img = Image.open(p)
    img = ImageOps.exif_transpose(img)  # field photos: honor orientation tag
    if img.mode != "RGB":
        img = img.convert("RGB")  # JPEG has no alpha
    img.thumbnail(box)  # never upscales
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=80)
    return buf.getvalue()


def write_photo_report(
    records: list[PhotoInspectionRecord],
    photo_map: dict[str, list[dict]],
    out_path: Path,
    *,
    site_id: str = "",
    photo_width_px: int = 300,
    photo_height_px: int = 225,
    qa: Optional[QACollector] = None,
) -> PhotoReportResult:
    """Write the per-well photo workbook. One row per photo (well metadata on
    the well's first row); wells with no photos get a placeholder row.

    Covers the union of wells in *records* and *photo_map*; mismatches in
    either direction raise aggregated QA WARNINGs (the tripwire for the
    group_template pilot assumption — see module docstring).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    qa = qa if qa is not None else QACollector()
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    record_by_id: dict[str, PhotoInspectionRecord] = {}
    duplicates: set[str] = set()
    for rec in records:
        prev = record_by_id.get(rec.well_id)
        if prev is not None:
            duplicates.add(rec.well_id)
            if rec.inspection_date <= prev.inspection_date:  # ISO sorts
                continue
        record_by_id[rec.well_id] = rec
    if duplicates:
        qa.add(SEV_WARNING, "duplicate_well_id",
               f"{len(duplicates)} well(s) have multiple inspection rows — "
               f"kept the latest InspectionDate: "
               f"{', '.join(sorted(duplicates))}")

    all_ids = sorted(set(record_by_id) | set(photo_map))

    wb = Workbook()
    ws = wb.active
    ws.title = f"Inspection {site_id}" if site_id else "Inspection"
    headers = ["Well ID", "Photo", "GPS (lat, lon)", "Inspection Info",
               "Notes"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    widths = {1: 20, 2: max(12, int(photo_width_px / 7)), 3: 24, 4: 30, 5: 45}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    wrap = Alignment(wrap_text=True, vertical="top")

    photos_embedded = 0
    missing_names: list[str] = []
    wells_without_photos: list[str] = []
    photos_without_record: list[str] = []
    row_no = 1

    for well_id in all_ids:
        rec = record_by_id.get(well_id)
        photos = photo_map.get(well_id, [])
        if rec is None:
            photos_without_record.append(well_id)
        if rec is not None and not photos:
            wells_without_photos.append(well_id)

        well_label = (f"{well_id} ({rec.condition})"
                      if rec is not None and rec.condition else well_id)
        if rec is None:
            gps = ""
            info = "No inspection record"
            notes = ""
        else:
            gps = (f"{rec.gps_lat:.6f}, {rec.gps_lon:.6f}"
                   if rec.gps_lat is not None and rec.gps_lon is not None
                   else "")
            info_parts = [rec.inspection_date]
            if rec.inspector:
                info_parts.append(rec.inspector)
            if rec.depth_to_water_ft is not None:
                info_parts.append(f"DTW: {rec.depth_to_water_ft:.2f} ft")
            info = "\n".join(p for p in info_parts if p)
            notes = rec.notes

        first = True
        for entry in photos or [None]:
            row_no += 1
            if first:
                ws.cell(row=row_no, column=1, value=well_label)
                ws.cell(row=row_no, column=3, value=gps).alignment = wrap
                ws.cell(row=row_no, column=4, value=info).alignment = wrap
                ws.cell(row=row_no, column=5, value=notes).alignment = wrap
                first = False
            if entry is None:
                ws.cell(row=row_no, column=2, value="No photos")
                continue
            data = prepare_image_bytes(
                entry.get("saved_path", ""),
                (photo_width_px, photo_height_px))
            if data is None:
                missing_names.append(entry.get("original_name")
                                     or entry.get("saved_path", ""))
                ws.cell(row=row_no, column=2,
                        value=f"[missing on disk: "
                              f"{entry.get('original_name', '?')}]")
                continue
            from openpyxl.drawing.image import Image as XLImage
            ws.add_image(XLImage(io.BytesIO(data)), f"B{row_no}")
            # px -> points is x0.75; pad a little so the image fits the row.
            ws.row_dimensions[row_no].height = photo_height_px * 0.75 + 6
            photos_embedded += 1

    wb.save(out_path)

    if wells_without_photos:
        qa.add(SEV_WARNING, "wells_without_photos",
               f"{len(wells_without_photos)} well(s) have an inspection "
               f"record but no matched photos: "
               f"{', '.join(wells_without_photos)}. If ALL wells lack "
               f"photos, check that the harvest group_template renders a "
               f"well ID (pilot assumption).")
    if photos_without_record:
        qa.add(SEV_WARNING, "photos_without_record",
               f"{len(photos_without_record)} photo group(s) have no "
               f"inspection record: {', '.join(photos_without_record)}. "
               f"Unexpected group keys usually mean the harvest "
               f"group_template is not well-id-based (pilot assumption).")
    if missing_names:
        qa.add(SEV_WARNING, "photo_files_missing",
               f"{len(missing_names)} manifest photo(s) not found on disk: "
               f"{', '.join(missing_names)}")
    qa.add(SEV_INFO, "photo_report_summary",
           f"{len(all_ids)} well(s), {photos_embedded} photo(s) embedded, "
           f"{len(missing_names)} missing -> {out_path}")

    return PhotoReportResult(
        workbook_path=out_path,
        well_count=len(all_ids),
        photos_embedded=photos_embedded,
        photos_missing=len(missing_names),
    )


# Back-compat: this was module-private before it was reused by the HTML report.
_prepared_image_bytes = prepare_image_bytes
