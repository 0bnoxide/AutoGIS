"""sampling_event_writer.py — serialize a SamplingEventPlan to XLSX.

Three-sheet workbook:
  Expected_Samples  — one row per (well × analyte_group), includes FD rows
  Crew_Assignment   — one row per well
  COC_Draft         — same rows as Expected_Samples, formatted for field crew

Headless: openpyxl only, no arcpy.
"""
from __future__ import annotations

from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill

from .create_sampling_event import SamplingEventPlan, ExpectedSampleRow, CrewAssignmentRow

_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")

_ES_HEADERS = [
    "SampleID", "LocationID", "EventDate", "Matrix", "AnalyteGroup",
    "SampleType", "ContainerType", "Preservative", "HoldTime_hr",
    "BottleCount", "COCNumber", "AssignedTo",
]

_CA_HEADERS = ["LocationID", "AssignedTo", "BottleCount"]

_COC_HEADERS = [
    "COCNumber", "SampleID", "LocationID", "EventDate", "Matrix",
    "AnalyteGroup", "SampleType", "ContainerType", "Preservative",
    "HoldTime_hr", "BottleCount", "LabName", "ExtraBottles",
    "SamplerSignature", "DateTimeSampled",
]


def _write_header(ws, headers: List[str]) -> None:
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(1, col, h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL


def _row_es(r: ExpectedSampleRow) -> list:
    return [
        r.sample_id, r.location_id, r.event_date, r.matrix,
        r.analyte_group, r.sample_type, r.container_type, r.preservative,
        r.hold_time_hr, r.bottle_count, r.coc_number, r.assigned_to,
    ]


def _row_ca(r: CrewAssignmentRow) -> list:
    return [r.location_id, r.assigned_to, r.bottle_count]


def _row_coc(r: ExpectedSampleRow, lab_name: str) -> list:
    return [
        r.coc_number, r.sample_id, r.location_id, r.event_date,
        r.matrix, r.analyte_group, r.sample_type, r.container_type,
        r.preservative, r.hold_time_hr, r.bottle_count,
        lab_name,
        "",   # ExtraBottles — blank for field crew
        "",   # SamplerSignature — blank for field crew
        "",   # DateTimeSampled — blank for field crew
    ]


def write_sampling_event_workbook(
    plan: SamplingEventPlan,
    out_path: Path,
) -> Path:
    """Write the plan to a three-sheet XLSX workbook and return out_path."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # ── Expected_Samples ──
    ws_es = wb.create_sheet("Expected_Samples")
    _write_header(ws_es, _ES_HEADERS)
    for row_idx, sample in enumerate(plan.expected_samples, start=2):
        for col, val in enumerate(_row_es(sample), start=1):
            ws_es.cell(row_idx, col, val)

    # ── Crew_Assignment ──
    ws_ca = wb.create_sheet("Crew_Assignment")
    _write_header(ws_ca, _CA_HEADERS)
    for row_idx, assignment in enumerate(plan.crew_assignments, start=2):
        for col, val in enumerate(_row_ca(assignment), start=1):
            ws_ca.cell(row_idx, col, val)

    # ── COC_Draft ──
    ws_coc = wb.create_sheet("COC_Draft")
    _write_header(ws_coc, _COC_HEADERS)
    for row_idx, sample in enumerate(plan.expected_samples, start=2):
        for col, val in enumerate(_row_coc(sample, plan.lab_name), start=1):
            ws_coc.cell(row_idx, col, val)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path
