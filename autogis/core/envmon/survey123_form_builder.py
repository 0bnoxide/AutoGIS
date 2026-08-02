"""Build a Survey123 XLSForm workbook from site, event, and analyte configs.

Headless — no arcpy imports (ADR-002). openpyxl only (ADR-008).

XLSForm spec: three sheets named 'survey', 'choices', 'settings'.
The resulting .xlsx is opened in ArcGIS Survey123 Connect by the user.
"""
from __future__ import annotations

import re
from typing import Dict, List

import openpyxl

from .sample_id import xform_sample_id_calc

_QA_FLAGS = [
    ("resampled", "Resampled"),
    ("field_dup_a", "Field Duplicate A"),
    ("field_dup_b", "Field Duplicate B"),
    ("equipment_blank", "Equipment Blank"),
    ("turbid", "Turbid / High Turbidity"),
    ("other", "Other"),
]

_MATRIX_LABELS = {
    "GW": "Groundwater",
    "SOIL": "Soil",
    "SW": "Surface Water",
    "SEDIMENT": "Sediment",
}

_SURVEY_HEADERS = ["type", "name", "label", "hint", "required", "calculation",
                   "appearance", "default"]
_CHOICES_HEADERS = ["list_name", "name", "label"]
_SETTINGS_HEADERS = ["form_title", "form_id", "version", "instance_name"]


def _slug(text: str) -> str:
    """Lower-case alphanumeric slug for choice names."""
    s = re.sub(r"[^0-9A-Za-z]+", "_", text).strip("_").lower() or "v"
    if s[0].isdigit():
        s = "v_" + s
    return s


def _field_name(analyte: str, used: set) -> str:
    """XLSForm-safe field name for an analyte, unique within *used*."""
    s = re.sub(r"[^0-9A-Za-z_]", "_", analyte).strip("_") or "F"
    if s[0].isdigit():
        s = "F_" + s
    s = s[:60]
    base, i = s, 1
    while s.upper() in used:
        i += 1
        s = f"{base[:57]}_{i}"
    used.add(s.upper())
    return s


def _write_row(ws, row_idx: int, values: list) -> None:
    for col, val in enumerate(values, start=1):
        ws.cell(row_idx, col, val)


def build_xlsform(
    site_config: dict,
    event_config: dict,
    analyte_dict: dict,
) -> openpyxl.Workbook:
    """Return an XLSForm workbook for the given site/event/analyte configs.

    site_config keys used: site_id, site_name
    event_config keys: analyte_groups, crew_list, coc_prefix, matrices,
                       location_ids
    analyte_dict structure: {analytes: {name: {abbreviation, display_order,
                                               default_units_by_matrix}}}
    """
    analytes_meta: Dict[str, dict] = analyte_dict.get("analytes", {})
    matrices: List[str] = event_config.get("matrices", ["GW"])
    primary_matrix = matrices[0] if matrices else "GW"
    location_ids: List[str] = event_config.get("location_ids", [])
    crew_list: List[str] = event_config.get("crew_list", [])
    analyte_groups: Dict[str, List[str]] = event_config.get("analyte_groups", {})
    site_id: str = site_config.get("site_id", "SITE")
    site_name: str = site_config.get("site_name", site_id)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ------------------------------------------------------------------ survey
    survey = wb.create_sheet("survey")
    _write_row(survey, 1, _SURVEY_HEADERS)
    r = 2

    def row(*vals):
        nonlocal r
        _write_row(survey, r, list(vals))
        r += 1

    sample_id_calc = xform_sample_id_calc()

    row("select_one well_list", "WellID", "Well ID", "", "yes")
    row("date", "SamplingDate", "Sampling Date", "", "yes")
    row("select_one matrix_list", "Matrix", "Sample Matrix", "", "yes")
    row("select_one crew_list", "SampledBy", "Sampled By", "", "yes")
    row("text", "COCNumber", "COC Number")

    used_names: set = set()
    for group_name, analyte_names in analyte_groups.items():
        group_slug = _slug(group_name)
        row("begin_group", f"grp_{group_slug}", group_name)
        for analyte in analyte_names:
            meta = analytes_meta.get(analyte, {})
            abbrev = meta.get("abbreviation", analyte)
            units = (meta.get("default_units_by_matrix") or {}).get(
                primary_matrix, "")
            label = f"{abbrev} ({units})" if units else abbrev
            fname = _field_name(analyte, used_names)
            row("decimal", fname, label)
        row("end_group")

    row("decimal", "DepthToWater_ft", "Depth to Water (ft)", "GW only")
    row("select_multiple qa_flags", "QAFlags", "QA Flags")
    # After QAFlags, which the duplicate leg reads: XLSForm resolves
    # calculates by dependency rather than row order, but a backward
    # reference needs no such guarantee and costs nothing — the question is
    # a calculate, so it renders nothing wherever it sits. The leg reuses the
    # QAFlags question; its A/B choices give two same-day duplicates distinct
    # identities without adding a second field-duplicate affordance.
    row("calculate", "SampleID", "", "", "", sample_id_calc)
    row("text", "Notes", "Notes")

    # ----------------------------------------------------------------- choices
    choices = wb.create_sheet("choices")
    _write_row(choices, 1, _CHOICES_HEADERS)
    cr = 2

    def crow(*vals):
        nonlocal cr
        _write_row(choices, cr, list(vals))
        cr += 1

    for loc in location_ids:
        crow("well_list", loc, loc)

    for matrix in matrices:
        crow("matrix_list", matrix, _MATRIX_LABELS.get(matrix, matrix))

    for member in crew_list:
        crow("crew_list", _slug(member), member)

    for flag_name, flag_label in _QA_FLAGS:
        crow("qa_flags", flag_name, flag_label)

    # ---------------------------------------------------------------- settings
    settings = wb.create_sheet("settings")
    _write_row(settings, 1, _SETTINGS_HEADERS)
    form_id = _slug(f"{site_id}_sampling")
    _write_row(settings, 2, [f"{site_name} Field Sampling", form_id, "1", ""])

    return wb
