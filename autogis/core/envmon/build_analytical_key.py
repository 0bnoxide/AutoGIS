"""build_analytical_key.py — analytical key/legend table for map layouts (Tool 5.5).

All functions here are arcpy-free (headless): the key is emitted as CSV / XLSX /
Markdown. (A GDB-table writer for a planned Env_AnalyticalKey feature class was
specced in 2026-06 but never wired — no table, no CLI option, no caller — and
was removed as dead code in ADR-0098.)

Default analyte selection = ``include_in_default_figures == True``, sorted by
``display_order`` then canonical name; override via ``analyte_filter``. All
values in the current DRAFT config are null — that is expected and renders as
``NE`` (Not Established) with a draft_flag. A screening value of exactly 0.0 is
a real, established threshold, not "missing".
"""
from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

_BOLD = Font(bold=True)
_DRAFT_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")  # light amber
_MAX_COL_WIDTH = 40


@dataclass
class AnalyticalKeyRow:
    canonical: str
    abbreviation: str
    display_order: int
    units: str
    screening_value: Optional[float]   # None = not established
    screening_units: str
    screening_source: str
    level_established: bool             # True only when screening_value is not None
    draft_flag: bool                   # True when source has "_TODO" or value is None


def _is_draft(entry: Optional[dict]) -> bool:
    """A screening entry is a DRAFT stub when absent, null-valued, or _TODO-sourced."""
    if entry is None:
        return True
    if entry.get("value") is None:        # is-None: 0.0 is a real value, not draft
        return True
    if "_TODO" in str(entry.get("source", "")):
        return True
    return False


def build_analytical_key(
    analytes: dict,
    screening_levels: dict,
    matrix: str,
    analyte_filter: Optional[list] = None,
) -> list:
    """Assemble a sorted AnalyticalKeyRow list from analyte dict + screening levels.

    ``matrix`` is required — units and screening lookups are both keyed by it.
    """
    matrix_screening = (screening_levels.get(matrix) or {})

    rows: list = []
    for canonical, entry in analytes.items():
        if canonical.startswith("_") or not isinstance(entry, dict):
            continue
        if analyte_filter is not None:
            if canonical not in analyte_filter:
                continue
        elif not entry.get("include_in_default_figures", False):
            continue

        units = (entry.get("default_units_by_matrix") or {}).get(matrix, "")

        sl_entry = matrix_screening.get(canonical)
        if sl_entry and isinstance(sl_entry, dict):
            screening_value = sl_entry.get("value")
            screening_units = sl_entry.get("units", "")
            screening_source = sl_entry.get("source", "")
        else:
            screening_value = None
            screening_units = ""
            screening_source = ""

        rows.append(AnalyticalKeyRow(
            canonical=canonical,
            abbreviation=str(entry.get("abbreviation", canonical)),
            display_order=int(entry.get("display_order", 9999)),
            units=units,
            screening_value=screening_value,
            screening_units=screening_units,
            screening_source=screening_source,
            level_established=screening_value is not None,  # 0.0 IS established
            draft_flag=_is_draft(sl_entry),
        ))

    rows.sort(key=lambda r: (r.display_order, r.canonical))
    return rows


_CSV_FIELDNAMES = [
    "canonical", "abbreviation", "display_order", "units",
    "screening_value", "screening_units", "screening_source",
    "level_established", "draft_flag",
]

_XLSX_HEADERS = [
    "Analyte", "Abbr", "Order", "Units",
    "Screening Level", "SL Units", "SL Source",
    "Level Established", "Draft/Stub",
]


def format_key_markdown(rows: list, matrix: str, site_id: str = "") -> str:
    """Render rows as a markdown table with header and a draft footnote."""
    title = f"Analytical Key — Matrix: {matrix}"
    if site_id:
        title += f"  |  Site: {site_id}"

    lines = [f"# {title}", ""]
    lines.append(f"| {'Analyte':<28} | {'Abbr':<6} | {'Units':<10} | "
                 f"{'Screening Level':>16} | {'SL Units':<10} | Source |")
    lines.append(f"| {'-'*28} | {'-'*6} | {'-'*10} | {'-'*16} | {'-'*10} | "
                 f"{'-'*30} |")
    has_draft = False
    for r in rows:
        sl_display = f"{r.screening_value}" if r.level_established else "NE"
        if r.draft_flag:
            sl_display += " *"
            has_draft = True
        lines.append(f"| {r.canonical:<28} | {r.abbreviation:<6} | "
                     f"{r.units:<10} | {sl_display:>16} | "
                     f"{r.screening_units:<10} | {r.screening_source} |")

    lines.append("")
    lines.append("**Legend:** Values in **bold** on map figures exceed the "
                 "screening level. J = estimated value, U = non-detect at "
                 "reporting limit.")
    lines.append("")
    if has_draft:
        lines.append("> **NE** = screening level Not Established (null or "
                     "absent). **\\*** = unverified / pre-production: the level "
                     "is null, absent, or its source contains _TODO. Do not use "
                     "NE or starred rows for regulatory reporting until replaced "
                     "with verified citations.")
    return "\n".join(lines)


def write_key_csv(rows: list, path: Path) -> None:
    """Write key rows to CSV."""
    with Path(path).open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=_CSV_FIELDNAMES)
        writer.writeheader()
        for r in rows:
            writer.writerow({
                "canonical": r.canonical,
                "abbreviation": r.abbreviation,
                "display_order": r.display_order,
                "units": r.units,
                # is-not-None: a 0.0 level writes "0.0", not blank.
                "screening_value": (r.screening_value
                                    if r.screening_value is not None else ""),
                "screening_units": r.screening_units,
                "screening_source": r.screening_source,
                "level_established": r.level_established,
                "draft_flag": r.draft_flag,
            })


def write_key_xlsx(rows: list, path: Path, matrix: str = "") -> None:
    """Write key rows to an Excel workbook (headless, openpyxl only)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "AnalyticalKey"

    for ci, header in enumerate(_XLSX_HEADERS, start=1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.font = _BOLD

    for ri, r in enumerate(rows, start=2):
        sl_display = r.screening_value if r.level_established else None
        values = [r.canonical, r.abbreviation, r.display_order, r.units,
                  sl_display, r.screening_units, r.screening_source,
                  r.level_established, r.draft_flag]
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if r.draft_flag:
                cell.fill = _DRAFT_FILL

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None),
                      default=0)
        ws.column_dimensions[col[0].column_letter].width = min(
            max_len + 2, _MAX_COL_WIDTH)

    if matrix:
        ws.cell(row=len(rows) + 3, column=1, value=f"Matrix: {matrix}")

    wb.save(Path(path))
