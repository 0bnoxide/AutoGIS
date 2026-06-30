"""Generate per-location/analyte trend charts in Excel (Tool 4.6).

Reads a history CSV (columns: LocationID, AnalyteName, SampleDate,
ResultValue, ReportedUnits, ScreeningLevel) and writes an openpyxl
workbook with one LineChart per location/analyte pair, grouped by
analyte into named sheets.

Headless: openpyxl only, no arcpy.
"""
from __future__ import annotations

import csv
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

# Each series block occupies this many rows so stacked charts never overlap.
_BLOCK_ROWS = 20

# Data columns within each block:
#   Col A (1): SampleDate
#   Col B (2): ResultValue
#   Col C (3): ScreeningLevel  (present only when screening_level is not None)
# Charts are anchored at Col E (5) of the block start row.
_COL_DATE = 1
_COL_VALUE = 2
_COL_SL = 3
_COL_CHART_ANCHOR = 5


@dataclass
class TrendSeries:
    location_id: str
    analyte_name: str
    dates: list[str] = field(default_factory=list)
    values: list[float] = field(default_factory=list)
    screening_level: Optional[float] = None
    units: str = ""


def load_history_csv(path: Path) -> list[TrendSeries]:
    """Read history CSV; group by (LocationID, AnalyteName); skip ND rows; sort by date.

    Rows are skipped when ResultValue is empty, 'ND', 'NONE', 'N/A', or
    non-numeric. ScreeningLevel is taken from the first row that carries one and
    back-filled if an earlier row of the same key lacked it.
    """
    path = Path(path)
    groups: dict[tuple[str, str], TrendSeries] = {}

    with path.open(newline="", encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            loc = row.get("LocationID", "").strip()
            analyte = row.get("AnalyteName", "").strip()
            raw_val = row.get("ResultValue", "").strip()
            sample_date = row.get("SampleDate", "").strip()
            units = row.get("ReportedUnits", "").strip()
            screening_raw = row.get("ScreeningLevel", "").strip()

            if not raw_val or raw_val.upper() in ("ND", "NONE", "N/A"):
                continue
            try:
                val = float(raw_val)
            except ValueError:
                continue

            key = (loc, analyte)
            if key not in groups:
                sl: Optional[float] = None
                try:
                    sl = float(screening_raw) if screening_raw else None
                except ValueError:
                    sl = None
                groups[key] = TrendSeries(
                    location_id=loc, analyte_name=analyte,
                    units=units, screening_level=sl,
                )
            else:
                if groups[key].screening_level is None and screening_raw:
                    try:
                        groups[key].screening_level = float(screening_raw)
                    except ValueError:
                        pass

            groups[key].dates.append(sample_date)
            groups[key].values.append(val)

    result: list[TrendSeries] = []
    for series in groups.values():
        if series.dates:
            paired = sorted(zip(series.dates, series.values))
            dates, values = zip(*paired)
            series.dates = list(dates)
            series.values = list(values)
        result.append(series)
    return result


def _safe_sheet_name(name: str, suffix: str = "") -> str:
    """Produce a valid Excel sheet name (max 31 chars, no forbidden chars)."""
    for ch in r"/\?*[]:'":
        name = name.replace(ch, "_")
    return (name + suffix)[:31]


def write_trend_charts(
    series_list: list[TrendSeries],
    out_path: Path,
    *,
    max_per_sheet: int = 20,
) -> int:
    """Write one Excel workbook with one LineChart per location/analyte pair.

    Sheets are named after analytes; more than *max_per_sheet* locations for one
    analyte spill into additional sheets suffixed "(2)", "(3)", etc. Each series
    occupies a fixed block of _BLOCK_ROWS rows so stacked charts do not overlap.

    Returns the total number of charts written.
    """
    out_path = Path(out_path)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    by_analyte: dict[str, list[TrendSeries]] = defaultdict(list)
    for series in series_list:
        by_analyte[series.analyte_name].append(series)

    chart_count = 0

    for analyte, analyte_series in sorted(by_analyte.items()):
        pages = [
            analyte_series[i:i + max_per_sheet]
            for i in range(0, max(1, len(analyte_series)), max_per_sheet)
        ]
        for page_idx, page_series in enumerate(pages):
            suffix = f" ({page_idx + 1})" if len(pages) > 1 else ""
            ws = wb.create_sheet(title=_safe_sheet_name(analyte, suffix))

            data_start_row = 1
            for series in page_series:
                n = len(series.dates)

                ws.cell(data_start_row, _COL_DATE, "Date")
                ws.cell(data_start_row, _COL_VALUE, series.location_id)
                has_sl = series.screening_level is not None
                if has_sl:
                    ws.cell(data_start_row, _COL_SL, "Screening Level")

                for i, (d, v) in enumerate(zip(series.dates, series.values)):
                    r = data_start_row + 1 + i
                    ws.cell(r, _COL_DATE, d)
                    ws.cell(r, _COL_VALUE, v)
                    if has_sl:
                        ws.cell(r, _COL_SL, series.screening_level)

                chart = LineChart()
                chart.title = f"{series.location_id} — {analyte}"
                chart.y_axis.title = series.units or "Concentration"
                chart.x_axis.title = "Sample Date"
                chart.style = 10
                chart.grouping = "standard"
                chart.width = 18
                chart.height = 10

                values_ref = Reference(
                    ws, min_col=_COL_VALUE, max_col=_COL_VALUE,
                    min_row=data_start_row, max_row=data_start_row + n,
                )
                chart.add_data(values_ref, titles_from_data=True)

                cats_ref = Reference(
                    ws, min_col=_COL_DATE,
                    min_row=data_start_row + 1, max_row=data_start_row + n,
                )
                chart.set_categories(cats_ref)

                if has_sl:
                    sl_ref = Reference(
                        ws, min_col=_COL_SL, max_col=_COL_SL,
                        min_row=data_start_row, max_row=data_start_row + n,
                    )
                    chart.add_data(sl_ref, titles_from_data=True)

                anchor = f"{get_column_letter(_COL_CHART_ANCHOR)}{data_start_row}"
                ws.add_chart(chart, anchor)
                chart_count += 1

                data_start_row += _BLOCK_ROWS

    if not wb.worksheets:
        # No series (empty input): openpyxl refuses to save a zero-sheet
        # workbook, so emit a placeholder sheet instead of crashing.
        ws = wb.create_sheet(title="No Data")
        ws.cell(1, 1, "No trend data for the given inputs.")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return chart_count
