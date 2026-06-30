"""Seeded synthetic environmental-workbook generator (Tool 10.6).

Produces deterministic fake-but-realistic .xlsx workbooks so the parsers and
importers can be hardened against ugly real-world data without exposing client
project data. A scenario selects which "messiness" features to inject; a fixed
seed makes output reproducible for tests. openpyxl-only, headless, no arcpy.
"""
from __future__ import annotations

import random
from dataclasses import dataclass, field
from datetime import date, timedelta
from pathlib import Path
from typing import Set

import openpyxl

MESSINESS = (
    "merged_headers", "formulas", "formula_error", "nondetects", "qualifiers",
    "missing_dates", "unknown_well", "rpd_sheet", "soil_depths", "metals", "ibi",
)

#: A well ID deliberately absent from the generated well network.
UNKNOWN_WELL_ID = "MW-UNKNOWN"

_BASE_ANALYTES = ["Benzene", "Toluene"]
_METAL_ANALYTES = ["Lead", "Arsenic"]


@dataclass
class WorkbookScenario:
    site_id: str
    n_wells: int
    n_events: int
    features: Set[str] = field(default_factory=set)
    seed: int = 0


def _headers(soil: bool) -> list:
    cols = ["LocationID", "SampleDate", "Analyte"]
    if soil:
        cols += ["DepthTop", "DepthBottom"]
    cols += ["Result", "Units", "Qualifier"]
    return cols


def generate_workbook(scenario: WorkbookScenario, out_path: Path) -> Path:
    """Write a deterministic synthetic workbook with the requested quirks."""
    feats = set(scenario.features)
    unknown = {f for f in feats if f not in MESSINESS}
    if unknown:
        raise ValueError(f"unknown messiness features: {sorted(unknown)}")
    rng = random.Random(scenario.seed)
    soil = "soil_depths" in feats
    units = "mg/kg" if soil else "ug/L"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    header = _headers(soil)
    row_i = 1
    if "merged_headers" in feats:
        ws.cell(row=1, column=1,
                value=f"Site {scenario.site_id} — Analytical Results")
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=len(header))
        row_i = 2
    for c, name in enumerate(header, start=1):
        ws.cell(row=row_i, column=c, value=name)
    data_start = row_i + 1

    wells = [f"MW-{i + 1}" for i in range(scenario.n_wells)]
    if "unknown_well" in feats:
        wells.append(UNKNOWN_WELL_ID)
    analytes = list(_BASE_ANALYTES)
    if "metals" in feats:
        analytes += _METAL_ANALYTES
    dates = [date(2026, 1, 1) + timedelta(days=90 * e)
             for e in range(scenario.n_events)]

    r = data_start
    first_result_cell = None
    for well in wells:
        for d in dates:
            for analyte in analytes:
                rec = {c: "" for c in header}
                rec["LocationID"] = well
                # missing_dates: occasionally blank the date.
                rec["SampleDate"] = (
                    "" if ("missing_dates" in feats and rng.random() < 0.2)
                    else d.isoformat())
                rec["Analyte"] = analyte
                if soil:
                    top = rng.choice([0, 2, 8, 14])
                    rec["DepthTop"] = top
                    rec["DepthBottom"] = top + 2
                rec["Units"] = units
                # nondetect vs detected.
                if "nondetects" in feats and rng.random() < 0.3:
                    rl = round(rng.uniform(0.1, 1.0), 2)
                    rec["Result"] = f"<{rl}"
                    rec["Qualifier"] = "U"
                else:
                    rec["Result"] = round(rng.uniform(0.1, 100.0), 2)
                    if "qualifiers" in feats and rng.random() < 0.25:
                        rec["Qualifier"] = rng.choice(["J", "B"])
                for c, name in enumerate(header, start=1):
                    cell = ws.cell(row=r, column=c, value=rec[name])
                    if name == "Result" and first_result_cell is None:
                        first_result_cell = cell
                r += 1

    # A real formula (string; openpyxl does not evaluate).
    if "formulas" in feats:
        ws.cell(row=r, column=1, value="=SUM(1,2)")
        r += 1
    # A cached error value, as a saved workbook with a broken formula reads.
    if "formula_error" in feats and first_result_cell is not None:
        first_result_cell.value = "#VALUE!"

    if "rpd_sheet" in feats:
        rpd = wb.create_sheet("RPD")
        for c, name in enumerate(
                ["ParentSampleID", "DuplicateSampleID", "Analyte",
                 "ParentResult", "DuplicateResult", "RPD"], start=1):
            rpd.cell(row=1, column=c, value=name)
        for i, analyte in enumerate(analytes, start=2):
            parent = round(rng.uniform(1, 50), 2)
            dup = round(parent * rng.uniform(0.8, 1.2), 2)
            rpd.append([f"{wells[0]}-S", f"{wells[0]}-D", analyte, parent, dup,
                        round(abs(parent - dup) / ((parent + dup) / 2) * 100, 1)])

    if "ibi" in feats:
        ibi = wb.create_sheet("IBI")
        for c, name in enumerate(["StationID", "Metric", "Score"], start=1):
            ibi.cell(row=1, column=c, value=name)
        for i, well in enumerate(wells[:scenario.n_wells], start=2):
            ibi.append([well, "TaxaRichness", rng.randint(1, 10)])

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out
