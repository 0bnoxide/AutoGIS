"""CLI surface for envmon export-report-format-summary-tables.

Wires export_summary_tables (report-appendix cross-tabs) into the CLI,
kept distinct from export-summary (flat QA sheets).
"""
import csv
from dataclasses import asdict, fields as dc_fields
from datetime import date

import openpyxl
from click.testing import CliRunner

from autogis.adapters.cli import autogis
from autogis.core.envmon.gdb_schema import AnalyticalResultRecord


def _r(LocationID, AnalyteCanonicalName, *, SampleDate=date(2026, 4, 1),
       DisplayText="--", ExceedsScreeningLevel=0, Matrix="GW",
       DepthIntervalText="", **kwargs):
    """Build a full AnalyticalResultRecord (all 39 fields required)."""
    defaults = {
        "ImportBatchID": "TEST_BATCH", "SiteID": "TEST_SITE", "Matrix": Matrix,
        "LocationID": LocationID,
        "SampleID": f"S_{LocationID}_{AnalyteCanonicalName}", "ParentSampleID": "",
        "SampleDate": SampleDate, "DepthTop_ft": None, "DepthBottom_ft": None,
        "DepthIntervalText": DepthIntervalText, "AnalyticalGroup": "VOC",
        "MethodGroup": "EPA8260", "AnalyteName": AnalyteCanonicalName,
        "AnalyteCanonicalName": AnalyteCanonicalName,
        "AnalyteAbbreviation": AnalyteCanonicalName[:3],
        "ResultRawText": DisplayText, "ResultNumeric": None,
        "ReportingLimit": None, "DetectionLimit": None, "Units": "ug/L",
        "Qualifier": "", "IsNonDetect": 0, "IsDetected": 1, "IsEstimated": 0,
        "IsDiluted": 0, "IsNotAnalyzed": 0, "IsNotSampled": 0,
        "IsNotMeasured": 0, "ScreeningLevel": 5.0, "ScreeningLevelSource": "RBSL",
        "ExceedsScreeningLevel": ExceedsScreeningLevel, "DisplayText": DisplayText,
        "DisplayColorClass": "EXCEED" if ExceedsScreeningLevel else "OK",
        "SourceWorkbook": "test.xlsx", "SourceSheet": "Sheet1", "SourceRow": 1,
        "SourceColumn": "A", "SourceCell": "A1",
    }
    defaults.update(kwargs)
    return AnalyticalResultRecord(**defaults)


def _write_results_csv(path, records):
    field_names = [f.name for f in dc_fields(AnalyticalResultRecord)]
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=field_names)
        w.writeheader()
        for r in records:
            w.writerow(asdict(r))


GW = [
    _r("MW-1", "Benzene", DisplayText="5.5", ExceedsScreeningLevel=1),
    _r("MW-2", "Benzene", DisplayText="<1.0"),
]
SOIL = [
    _r("SB-1", "Benzene", DisplayText="10", Matrix="SOIL",
       DepthIntervalText="0-2 ft"),
]


def test_happy_path_writes_three_sheets(tmp_path):
    csv_path = tmp_path / "results.csv"
    out = tmp_path / "report.xlsx"
    _write_results_csv(csv_path, GW + SOIL)
    result = CliRunner().invoke(autogis, [
        "envmon", "export-report-format-summary-tables",
        "--results-csv", str(csv_path), "--output", str(out)])
    assert result.exit_code == 0, result.output
    assert out.exists()
    wb = openpyxl.load_workbook(out)
    assert set(wb.sheetnames) == {"Current Event", "GW by Event", "Soil by Depth"}


def test_no_soil_flag_drops_sheet(tmp_path):
    csv_path = tmp_path / "results.csv"
    out = tmp_path / "report.xlsx"
    _write_results_csv(csv_path, GW + SOIL)
    result = CliRunner().invoke(autogis, [
        "envmon", "export-report-format-summary-tables",
        "--results-csv", str(csv_path), "--output", str(out),
        "--no-soil-by-depth"])
    assert result.exit_code == 0, result.output
    wb = openpyxl.load_workbook(out)
    assert "Soil by Depth" not in wb.sheetnames
    assert set(wb.sheetnames) == {"Current Event", "GW by Event"}


def test_fail_on_warning_escalates_for_missing_soil(tmp_path):
    csv_path = tmp_path / "results.csv"
    out = tmp_path / "report.xlsx"
    _write_results_csv(csv_path, GW)   # GW only -> soil_records_absent warning
    result = CliRunner().invoke(autogis, [
        "envmon", "export-report-format-summary-tables",
        "--results-csv", str(csv_path), "--output", str(out),
        "--fail-on", "warning"])
    assert result.exit_code != 0
    assert "soil_records_absent" in result.output


def test_command_registered_help():
    result = CliRunner().invoke(autogis, [
        "envmon", "export-report-format-summary-tables", "--help"])
    assert result.exit_code == 0
    assert "--results-csv" in result.output
    assert "--output" in result.output
    assert "--fail-on" in result.output
