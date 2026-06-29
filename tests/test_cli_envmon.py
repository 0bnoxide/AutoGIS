"""CLI surface for the envmon sub-group.

(a) A LOCAL tool (``import-gdb``) must fail cleanly under a no-arcpy condition
    with a message that mentions arcpy — a click error, not a traceback.
(b) A headless tool (``inspect``) must exit 0 against a real workbook.
(c) Regression: import-edd, import-rtk-survey, route-survey123 must produce a
    clean ClickException (not a raw KeyError traceback) when arcpy is absent.
"""
import builtins

from click.testing import CliRunner

from autogis.adapters.cli import autogis


def _no_arcpy_monkeypatch(monkeypatch):
    """Helper: make 'import arcpy' raise ModuleNotFoundError."""
    real = builtins.__import__

    def fake(name, *a, **k):
        if name == "arcpy":
            raise ModuleNotFoundError("No module named 'arcpy'")
        return real(name, *a, **k)

    monkeypatch.setattr(builtins, "__import__", fake)


def test_local_tool_clean_error_without_arcpy(monkeypatch, tmp_path):
    real = builtins.__import__

    def fake(name, *a, **k):
        if name == "arcpy":
            raise ModuleNotFoundError("No module named 'arcpy'")
        return real(name, *a, **k)

    monkeypatch.setattr(builtins, "__import__", fake)
    # Two existing path args so click's exists=True checks pass and we reach
    # the guard.
    site = tmp_path / "site.yaml"
    site.write_text("site_id: X\n", encoding="utf-8")
    wb = tmp_path / "wb.xlsx"
    wb.write_text("x", encoding="utf-8")

    result = CliRunner().invoke(
        autogis, ["envmon", "import-gdb", str(site), str(wb)])
    assert result.exit_code != 0
    assert "arcpy" in result.output.lower()
    assert result.exception is None or isinstance(result.exception, SystemExit)


def test_inspect_runs_headless(tmp_path):
    import openpyxl

    path = tmp_path / "synthetic.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Well ID"
    ws["B1"] = "Date"
    ws["C1"] = "Benzene"
    ws["A2"] = "MW-1"
    ws["B2"] = "2026-01-01"
    ws["C2"] = 1.5
    wb.save(path)

    result = CliRunner().invoke(autogis, ["envmon", "inspect", str(path)])
    assert result.exit_code == 0, result.output
    assert "Data" not in result.output or "sheet" in result.output.lower()


def test_import_edd_cmd_registered():
    """import-edd command exists and shows help without error."""
    runner = CliRunner()
    result = runner.invoke(autogis, ["envmon", "import-edd", "--help"])
    assert result.exit_code == 0
    assert "--edd" in result.output
    assert "--profile-path" in result.output
    assert "--site" in result.output
    assert "--gdb" in result.output


def test_import_edd_clean_error_without_arcpy(monkeypatch, tmp_path):
    """Regression #62: import-edd must raise a ClickException, not KeyError."""
    _no_arcpy_monkeypatch(monkeypatch)
    edd = tmp_path / "edd.csv"
    edd.write_text("col\n", encoding="utf-8")
    profile = tmp_path / "profile.yaml"
    profile.write_text("name: test\n", encoding="utf-8")

    result = CliRunner().invoke(
        autogis,
        ["envmon", "import-edd",
         "--edd", str(edd),
         "--profile-path", str(profile),
         "--site", "H281",
         "--gdb", str(tmp_path / "out.gdb")],
    )
    assert result.exit_code != 0
    assert "arcpy" in result.output.lower()
    # Must be a clean click error -- no raw KeyError traceback
    assert result.exception is None or isinstance(result.exception, SystemExit)


def test_import_rtk_survey_clean_error_without_arcpy(monkeypatch, tmp_path):
    """Regression #62: import-rtk-survey must raise a ClickException, not KeyError."""
    _no_arcpy_monkeypatch(monkeypatch)
    csv = tmp_path / "survey.csv"
    csv.write_text("col\n", encoding="utf-8")

    result = CliRunner().invoke(
        autogis,
        ["envmon", "import-rtk-survey",
         str(csv),
         "--site", "H281",
         "--gdb", str(tmp_path / "out.gdb")],
    )
    assert result.exit_code != 0
    assert "arcpy" in result.output.lower()
    assert result.exception is None or isinstance(result.exception, SystemExit)


def test_route_survey123_clean_error_without_arcpy(monkeypatch, tmp_path):
    """Regression #62: route-survey123 must raise a ClickException, not KeyError."""
    _no_arcpy_monkeypatch(monkeypatch)
    inp = tmp_path / "input.csv"
    inp.write_text("col\n", encoding="utf-8")

    result = CliRunner().invoke(
        autogis,
        ["envmon", "route-survey123",
         str(inp),
         "--site", "H281",
         "--gdb", str(tmp_path / "out.gdb")],
    )
    assert result.exit_code != 0
    assert "arcpy" in result.output.lower()
    assert result.exception is None or isinstance(result.exception, SystemExit)
