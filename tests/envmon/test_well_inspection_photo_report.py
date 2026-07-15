"""Tests for Tool 7.4 GenerateWellInspectionPhotoReport (headless).

Photo-embedding tests require Pillow and are gated per-test with
``pytest.importorskip("PIL")``; everything else is stdlib+openpyxl only.
"""
import builtins
import csv
import importlib
import sys
from pathlib import Path

import pytest
from click.testing import CliRunner

from autogis.adapters.cli import autogis
from autogis.core.common.qa import QACollector, SEV_WARNING
from autogis.core.envmon.well_inspection_photo_report import (
    PhotoInspectionRecord, load_inspection_records, match_photos_to_wells,
    write_photo_report)


# ── helpers ──────────────────────────────────────────────────────────────────

_INSPECTION_FIELDS = ["WellID", "InspectionDate", "Inspector", "Condition",
                      "Notes", "GPS_Lat", "GPS_Lon", "DepthToWaterFt"]


def _write_inspections_csv(path: Path, rows=None) -> Path:
    rows = rows if rows is not None else [{
        "WellID": "MW-01", "InspectionDate": "2026-06-15",
        "Inspector": "J. Smith", "Condition": "Good",
        "Notes": "No visible damage", "GPS_Lat": "48.196800",
        "GPS_Lon": "-106.636500", "DepthToWaterFt": "12.34",
    }, {
        "WellID": "MW-02", "InspectionDate": "2026-06-15",
        "Inspector": "J. Smith", "Condition": "Fair",
        "Notes": "Minor corrosion", "GPS_Lat": "", "GPS_Lon": "",
        "DepthToWaterFt": "",
    }]
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=_INSPECTION_FIELDS)
        w.writeheader()
        w.writerows(rows)
    return path


def _manifest_row(saved_path, *, status="downloaded", disposition=None,
                  name="photo.jpg"):
    """One harvester-manifest row (AttachmentResult CSV columns)."""
    return {
        "objectid": "1", "attachment_id": "101", "original_name": name,
        "saved_path": str(saved_path) if saved_path else "",
        "size": "1000", "status": status,
        "error": "", "disposition": disposition if disposition is not None else status,
        "checksum": "", "algorithm": "", "geometry": "",
        "source_table": "", "relationship_id": "",
    }


def _make_jpeg(path: Path, size=(8, 6)) -> Path:
    from PIL import Image
    path.parent.mkdir(parents=True, exist_ok=True)
    Image.new("RGB", size, color=(200, 40, 40)).save(path, format="JPEG")
    return path


def _cell_texts(xlsx_path: Path) -> list[str]:
    from openpyxl import load_workbook
    ws = load_workbook(xlsx_path).active
    return [str(c.value) for row in ws.iter_rows() for c in row
            if c.value is not None]


# ── load_inspection_records ──────────────────────────────────────────────────

def test_load_inspection_records_parses_rows(tmp_path):
    p = _write_inspections_csv(tmp_path / "inspections.csv")
    records = load_inspection_records(p)
    assert len(records) == 2
    r = records[0]
    assert r.well_id == "MW-01"
    assert r.condition == "Good"
    assert r.inspector == "J. Smith"
    assert r.gps_lat == pytest.approx(48.1968)
    assert r.gps_lon == pytest.approx(-106.6365)
    assert r.depth_to_water_ft == pytest.approx(12.34)


def test_load_inspection_records_blank_optionals_are_none(tmp_path):
    p = _write_inspections_csv(tmp_path / "inspections.csv")
    r = load_inspection_records(p)[1]
    assert r.gps_lat is None and r.gps_lon is None
    assert r.depth_to_water_ft is None


def test_load_inspection_records_bad_float_warns(tmp_path):
    p = _write_inspections_csv(tmp_path / "inspections.csv", rows=[{
        "WellID": "MW-01", "InspectionDate": "2026-06-15",
        "Inspector": "", "Condition": "Good", "Notes": "",
        "GPS_Lat": "not-a-number", "GPS_Lon": "", "DepthToWaterFt": "",
    }])
    qa = QACollector()
    records = load_inspection_records(p, qa=qa)
    assert records[0].gps_lat is None
    assert any(r.severity == SEV_WARNING for r in qa.records)


def test_load_inspection_records_blank_well_id_skipped(tmp_path):
    p = _write_inspections_csv(tmp_path / "inspections.csv", rows=[{
        "WellID": "", "InspectionDate": "2026-06-15", "Inspector": "",
        "Condition": "Good", "Notes": "", "GPS_Lat": "", "GPS_Lon": "",
        "DepthToWaterFt": "",
    }])
    qa = QACollector()
    assert load_inspection_records(p, qa=qa) == []
    assert any(r.category == "missing_well_id" for r in qa.records)


# ── match_photos_to_wells ────────────────────────────────────────────────────

def test_match_photos_groups_by_well_dir(tmp_path):
    rows = [
        _manifest_row(tmp_path / "MW-01" / "a.jpg", name="a.jpg"),
        _manifest_row(tmp_path / "MW-01" / "b.jpg", status="skipped",
                      name="b.jpg"),
        _manifest_row(tmp_path / "MW-02" / "c.jpg", name="c.jpg"),
    ]
    photo_map = match_photos_to_wells(rows, tmp_path)
    assert sorted(photo_map) == ["MW-01", "MW-02"]
    assert len(photo_map["MW-01"]) == 2


def test_match_photos_excludes_failed_and_blank_paths(tmp_path):
    rows = [
        _manifest_row("", status="failed", name="lost.jpg"),
        _manifest_row(tmp_path / "MW-01" / "a.jpg", name="a.jpg"),
    ]
    photo_map = match_photos_to_wells(rows, tmp_path)
    names = [r["original_name"] for rs in photo_map.values() for r in rs]
    assert names == ["a.jpg"]


def test_match_photos_blank_disposition_falls_back_to_status(tmp_path):
    rows = [_manifest_row(tmp_path / "MW-02" / "d.jpg", disposition="",
                          name="d.jpg")]
    photo_map = match_photos_to_wells(rows, tmp_path)
    assert list(photo_map) == ["MW-02"]


def test_match_photos_outside_harvest_dir_dropped_with_warning(tmp_path):
    rows = [_manifest_row(tmp_path / "elsewhere" / "stray.jpg",
                          name="stray.jpg")]
    qa = QACollector()
    photo_map = match_photos_to_wells(rows, tmp_path / "harvest", qa=qa)
    assert photo_map == {}
    assert any(r.category == "manifest_rows_outside_harvest_dir"
               for r in qa.records)


# ── write_photo_report ───────────────────────────────────────────────────────

def test_write_photo_report_embeds_photo(tmp_path):
    pytest.importorskip("PIL")
    from openpyxl import load_workbook
    photo = _make_jpeg(tmp_path / "harvest" / "MW-01" / "a.jpg")
    records = [PhotoInspectionRecord("MW-01", "2026-06-15", "J. Smith",
                                     "Good", "OK")]
    photo_map = {"MW-01": [_manifest_row(photo, name="a.jpg")]}
    out = tmp_path / "report.xlsx"
    qa = QACollector()
    result = write_photo_report(records, photo_map, out, site_id="H281",
                                qa=qa)
    assert out.exists()
    assert result.photos_embedded == 1
    assert result.photos_missing == 0
    assert result.well_count == 1
    assert len(load_workbook(out).active._images) == 1


def test_write_photo_report_condition_and_gps_in_cells(tmp_path):
    records = [PhotoInspectionRecord("MW-01", "2026-06-15", "J. Smith",
                                     "Good", "OK", gps_lat=48.1968,
                                     gps_lon=-106.6365)]
    out = tmp_path / "report.xlsx"
    write_photo_report(records, {}, out, site_id="H281", qa=QACollector())
    texts = _cell_texts(out)
    assert any("MW-01" in t and "Good" in t for t in texts)
    assert any("48.196800" in t and "-106.636500" in t for t in texts)


def test_write_photo_report_missing_file_placeholder(tmp_path):
    """Manifest row whose file is gone on disk -> placeholder + WARNING.

    Runs without Pillow: the on-disk check happens before the lazy import.
    """
    records = [PhotoInspectionRecord("MW-01", "2026-06-15")]
    gone = tmp_path / "harvest" / "MW-01" / "gone.jpg"
    photo_map = {"MW-01": [_manifest_row(gone, name="gone.jpg")]}
    out = tmp_path / "report.xlsx"
    qa = QACollector()
    result = write_photo_report(records, photo_map, out, site_id="H281",
                                qa=qa)
    assert result.photos_missing == 1
    assert result.photos_embedded == 0
    assert any("gone.jpg" in t for t in _cell_texts(out))
    assert any(r.category == "photo_files_missing" for r in qa.records)


def test_write_photo_report_record_without_photos_warns(tmp_path):
    records = [PhotoInspectionRecord("MW-09", "2026-06-15")]
    out = tmp_path / "report.xlsx"
    qa = QACollector()
    result = write_photo_report(records, {}, out, site_id="H281", qa=qa)
    assert result.well_count == 1
    assert any("No photos" in t for t in _cell_texts(out))
    assert any(r.category == "wells_without_photos" for r in qa.records)


def test_write_photo_report_photos_without_record_warns(tmp_path):
    pytest.importorskip("PIL")
    photo = _make_jpeg(tmp_path / "harvest" / "MW-77" / "a.jpg")
    photo_map = {"MW-77": [_manifest_row(photo, name="a.jpg")]}
    out = tmp_path / "report.xlsx"
    qa = QACollector()
    result = write_photo_report([], photo_map, out, site_id="H281", qa=qa)
    assert result.well_count == 1
    assert any("No inspection record" in t for t in _cell_texts(out))
    assert any(r.category == "photos_without_record" for r in qa.records)


def test_write_photo_report_wells_sorted(tmp_path):
    from openpyxl import load_workbook
    records = [PhotoInspectionRecord("MW-02", "2026-06-15"),
               PhotoInspectionRecord("MW-01", "2026-06-15")]
    out = tmp_path / "report.xlsx"
    write_photo_report(records, {}, out, site_id="H281", qa=QACollector())
    ws = load_workbook(out).active
    col_a = [c.value for (c,) in ws.iter_rows(min_col=1, max_col=1)
             if c.value]
    assert col_a.index(next(v for v in col_a if "MW-01" in v)) < \
        col_a.index(next(v for v in col_a if "MW-02" in v))


def test_write_photo_report_duplicate_well_keeps_latest(tmp_path):
    records = [
        PhotoInspectionRecord("MW-01", "2026-01-10", condition="Poor"),
        PhotoInspectionRecord("MW-01", "2026-06-15", condition="Good"),
    ]
    out = tmp_path / "report.xlsx"
    qa = QACollector()
    result = write_photo_report(records, {}, out, site_id="H281", qa=qa)
    assert result.well_count == 1
    texts = _cell_texts(out)
    assert any("Good" in t for t in texts)
    assert not any("Poor" in t for t in texts)
    assert any(r.category == "duplicate_well_id" for r in qa.records)


def test_write_photo_report_helpful_error_without_pillow(tmp_path, monkeypatch):
    """Photo on disk but Pillow unavailable -> ImportError with install hint."""
    pytest.importorskip("PIL")  # need Pillow to create the fixture jpeg
    photo = _make_jpeg(tmp_path / "harvest" / "MW-01" / "a.jpg")
    real = builtins.__import__

    def fake(name, *a, **k):
        if name.split(".")[0] == "PIL":
            raise ModuleNotFoundError("No module named 'PIL'")
        return real(name, *a, **k)

    monkeypatch.setattr(builtins, "__import__", fake)
    records = [PhotoInspectionRecord("MW-01", "2026-06-15")]
    photo_map = {"MW-01": [_manifest_row(photo, name="a.jpg")]}
    with pytest.raises(ImportError, match=r"autogis\[report\]"):
        write_photo_report(records, photo_map, tmp_path / "r.xlsx",
                           site_id="H281", qa=QACollector())


# ── import invariant ─────────────────────────────────────────────────────────

def test_module_imports_without_arcpy_arcgis_or_pillow(monkeypatch):
    real = builtins.__import__

    def fake(name, *a, **k):
        if name.split(".")[0] in ("arcpy", "arcgis", "PIL"):
            raise ModuleNotFoundError(f"No module named '{name}'")
        return real(name, *a, **k)

    monkeypatch.setattr(builtins, "__import__", fake)
    monkeypatch.delitem(
        sys.modules, "autogis.core.envmon.well_inspection_photo_report",
        raising=False)
    mod = importlib.import_module(
        "autogis.core.envmon.well_inspection_photo_report")
    assert hasattr(mod, "write_photo_report")


# ── CLI ──────────────────────────────────────────────────────────────────────

def test_generate_inspection_report_in_help():
    r = CliRunner().invoke(autogis, ["envmon", "--help"])
    assert "generate-inspection-report" in r.output


def test_generate_inspection_report_registered_as_cloud():
    from autogis.runtime.capabilities import Runtime, TOOLS
    assert TOOLS["generate-inspection-report"] is Runtime.CLOUD


def _empty_manifest(path: Path) -> Path:
    fieldnames = list(_manifest_row("").keys())
    with path.open("w", newline="", encoding="utf-8") as fh:
        csv.DictWriter(fh, fieldnames=fieldnames).writeheader()
    return path


def test_generate_inspection_report_end_to_end_no_photos(tmp_path):
    insp = _write_inspections_csv(tmp_path / "inspections.csv")
    manifest = _empty_manifest(tmp_path / "manifest.csv")
    harvest = tmp_path / "harvest"
    harvest.mkdir()
    out = tmp_path / "report.xlsx"
    r = CliRunner().invoke(autogis, [
        "envmon", "generate-inspection-report",
        "--inspections", str(insp), "--manifest", str(manifest),
        "--harvest-dir", str(harvest), "--site", "H281",
        "--out", str(out)])
    assert r.exit_code == 0, r.output
    assert out.exists()
    assert "2 well(s)" in r.output
    assert "Status: PASS" in r.output


def test_generate_inspection_report_end_to_end_with_photo(tmp_path):
    pytest.importorskip("PIL")
    insp = _write_inspections_csv(tmp_path / "inspections.csv")
    harvest = tmp_path / "harvest"
    photo = _make_jpeg(harvest / "MW-01" / "a.jpg")
    manifest = tmp_path / "manifest.csv"
    fieldnames = list(_manifest_row("").keys())
    with manifest.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fieldnames)
        w.writeheader()
        w.writerow(_manifest_row(photo, name="a.jpg"))
    out = tmp_path / "report.xlsx"
    r = CliRunner().invoke(autogis, [
        "envmon", "generate-inspection-report",
        "--inspections", str(insp), "--manifest", str(manifest),
        "--harvest-dir", str(harvest), "--site", "H281",
        "--out", str(out)])
    assert r.exit_code == 0, r.output
    assert "1 photo(s) embedded" in r.output


def test_prepare_image_bytes_is_public():
    from autogis.core.envmon import well_inspection_photo_report as m
    assert hasattr(m, "prepare_image_bytes")
    # missing file returns None without importing Pillow
    assert m.prepare_image_bytes("does_not_exist.jpg", (10, 10)) is None
