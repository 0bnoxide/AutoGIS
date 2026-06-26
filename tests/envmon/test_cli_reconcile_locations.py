import openpyxl
import yaml
from click.testing import CliRunner

from autogis.adapters.cli import autogis


def _profile(tmp_path):
    data = {"profile_id": "P", "sheets": [
        {"sheet_name": "S", "data_type": "METALS", "data_start_row": 2,
         "id_column": "A"}]}
    p = tmp_path / "profile.yaml"
    p.write_text(yaml.safe_dump(data), encoding="utf-8")
    return str(p)


def _workbook(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "Well ID"
    ws["A2"] = "MW-1"
    ws["A3"] = "MW-7A"   # typo vs MW-07A
    p = tmp_path / "wb.xlsx"
    wb.save(p)
    return str(p)


def _site(tmp_path):
    p = tmp_path / "site.yaml"
    p.write_text(yaml.safe_dump({"site_id": "H281",
                                 "monitoring_wells_fc": "MonitoringWells"}),
                 encoding="utf-8")
    return str(p)


def test_reconcile_locations_cli_reports_typo_and_unmatched(tmp_path):
    wells = tmp_path / "wells.csv"
    wells.write_text("LocationID\nMW-1\nMW-07A\n", encoding="utf-8")
    r = CliRunner().invoke(autogis, [
        "envmon", "reconcile-locations", _site(tmp_path), _workbook(tmp_path),
        "--profile", _profile(tmp_path), "--wells-csv", str(wells)])
    # MW-7A -> WARNING typo only; no ERROR -> default --fail-on error => PASS (exit 0)
    assert r.exit_code == 0, r.output
    assert "location_id_typo" in r.output


def test_reconcile_locations_cli_gdb_path_redirects(tmp_path):
    r = CliRunner().invoke(autogis, [
        "envmon", "reconcile-locations", _site(tmp_path), _workbook(tmp_path),
        "--profile", _profile(tmp_path), "--gdb"])
    assert r.exit_code != 0
    assert "ArcGIS Pro" in r.output or "toolbox" in r.output
