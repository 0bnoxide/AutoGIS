"""Tests for build_analytical_key core logic (arcpy-free)."""
import csv

from autogis.core.envmon.build_analytical_key import (
    AnalyticalKeyRow,
    build_analytical_key,
    format_key_markdown,
    write_key_csv,
    write_key_xlsx,
)

# Inline fixtures — do NOT source from real YAML config. The real
# screening_levels.yaml has value: null + _TODO sources throughout; that is the
# NORMAL case and must render cleanly as "NE".
_ANALYTES = {
    "Benzene": {
        "canonical_name": "Benzene", "abbreviation": "B",
        "analytical_group": "VPH_VOC", "display_order": 10,
        "default_units_by_matrix": {"GW": "ug/L", "SOIL": "mg/kg"},
        "include_in_default_figures": True, "significant_figures": 2,
    },
    "Toluene": {
        "canonical_name": "Toluene", "abbreviation": "T",
        "analytical_group": "VPH_VOC", "display_order": 20,
        "default_units_by_matrix": {"GW": "ug/L", "SOIL": "mg/kg"},
        "include_in_default_figures": True, "significant_figures": 2,
    },
    "Methane": {
        "canonical_name": "Methane", "abbreviation": "CH4",
        "analytical_group": "GAS", "display_order": 200,
        "default_units_by_matrix": {"GW": "mg/L"},
        "include_in_default_figures": False, "significant_figures": 2,
    },
}

_SCREENING_NULL = {
    "GW": {
        "Benzene": {"value": None, "units": "ug/L", "source": "_TODO MDEQ RBSL"},
        "Toluene": {"value": None, "units": "ug/L", "source": "_TODO MDEQ RBSL"},
    }
}

_SCREENING_POPULATED = {
    "GW": {
        "Benzene": {"value": 5.0, "units": "ug/L", "source": "MDEQ MCL Table A"},
        "Toluene": {"value": 1000.0, "units": "ug/L", "source": "MDEQ MCL Table A"},
    }
}


def test_default_filter_excludes_non_figure_analytes():
    rows = build_analytical_key(_ANALYTES, _SCREENING_NULL, matrix="GW")
    assert "Methane" not in [r.canonical for r in rows]


def test_default_filter_includes_figure_analytes():
    rows = build_analytical_key(_ANALYTES, _SCREENING_NULL, matrix="GW")
    canonicals = [r.canonical for r in rows]
    assert "Benzene" in canonicals and "Toluene" in canonicals


def test_rows_sorted_by_display_order():
    rows = build_analytical_key(_ANALYTES, _SCREENING_NULL, matrix="GW")
    assert [r.display_order for r in rows] == sorted(r.display_order for r in rows)
    assert rows[0].canonical == "Benzene"


def test_analyte_filter_restricts_output():
    rows = build_analytical_key(_ANALYTES, _SCREENING_NULL, matrix="GW",
                                analyte_filter=["Benzene"])
    assert len(rows) == 1 and rows[0].canonical == "Benzene"


def test_units_are_matrix_specific():
    gw = build_analytical_key(_ANALYTES, _SCREENING_NULL, matrix="GW")
    soil = build_analytical_key(_ANALYTES, _SCREENING_NULL, matrix="SOIL")
    assert next(r for r in gw if r.canonical == "Benzene").units == "ug/L"
    assert next(r for r in soil if r.canonical == "Benzene").units == "mg/kg"


def test_units_fallback_when_matrix_absent():
    rows = build_analytical_key(_ANALYTES, {}, matrix="SOIL",
                                analyte_filter=["Methane"])
    assert rows[0].units == ""


def test_null_screening_produces_not_established():
    rows = build_analytical_key(_ANALYTES, _SCREENING_NULL, matrix="GW")
    benzene = next(r for r in rows if r.canonical == "Benzene")
    assert benzene.level_established is False
    assert benzene.screening_value is None
    assert benzene.draft_flag is True
    assert "_TODO" in benzene.screening_source


def test_analyte_with_no_screening_entry_also_not_established():
    rows = build_analytical_key(_ANALYTES, _SCREENING_NULL, matrix="GW",
                                analyte_filter=["Methane"])
    assert rows[0].level_established is False
    assert rows[0].draft_flag is True


def test_populated_screening_sets_level_established():
    rows = build_analytical_key(_ANALYTES, _SCREENING_POPULATED, matrix="GW")
    benzene = next(r for r in rows if r.canonical == "Benzene")
    assert benzene.level_established is True
    assert benzene.screening_value == 5.0
    assert benzene.draft_flag is False
    assert benzene.screening_units == "ug/L"


def test_zero_screening_value_is_established_not_missing():
    """A screening level of exactly 0.0 is a real, established threshold (#82)."""
    screening = {"GW": {"Benzene": {"value": 0.0, "units": "ug/L",
                                    "source": "MDEQ MCL"}}}
    rows = build_analytical_key(_ANALYTES, screening, matrix="GW",
                                analyte_filter=["Benzene"])
    assert rows[0].screening_value == 0.0
    assert rows[0].level_established is True   # 0.0 is NOT "Not Established"
    assert rows[0].draft_flag is False


def test_format_markdown_ne_and_draft_banner_for_null():
    md = format_key_markdown(
        build_analytical_key(_ANALYTES, _SCREENING_NULL, matrix="GW"),
        matrix="GW")
    assert "B" in md and "NE" in md
    assert ("pre-production" in md.lower() or "not established" in md.lower())


def test_format_markdown_shows_value_when_populated():
    md = format_key_markdown(
        build_analytical_key(_ANALYTES, _SCREENING_POPULATED, matrix="GW"),
        matrix="GW")
    assert "5.0" in md or "5" in md


def test_write_key_csv_headers_and_null_value(tmp_path):
    out = tmp_path / "key.csv"
    write_key_csv(build_analytical_key(_ANALYTES, _SCREENING_NULL, matrix="GW"),
                  out)
    with out.open(newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        data = list(reader)
        fieldnames = reader.fieldnames or []
    for col in ("canonical", "abbreviation", "units", "screening_value",
                "level_established", "draft_flag"):
        assert col in fieldnames
    benzene = next(r for r in data if r["canonical"] == "Benzene")
    assert benzene["screening_value"] in ("", "None")
    assert benzene["level_established"] in ("False", "false", "0")


def test_write_key_xlsx_creates_file_with_header(tmp_path):
    from openpyxl import load_workbook
    out = tmp_path / "key.xlsx"
    rows = build_analytical_key(_ANALYTES, _SCREENING_NULL, matrix="GW")
    write_key_xlsx(rows, out, matrix="GW")
    assert out.exists() and out.stat().st_size > 0
    ws = load_workbook(out).active
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert "Analyte" in header
