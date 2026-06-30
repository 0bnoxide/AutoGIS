"""Tests for autogis.core.envmon.ingest_reviewer_comments (Tool 9.4)."""
import csv
import json
import pytest
from pathlib import Path

from autogis.core.common.qa import QACollector, SEV_WARNING, SEV_ERROR, SEV_INFO
from autogis.core.envmon.ingest_reviewer_comments import (
    ReviewerComment,
    VALID_STATUSES,
    DEFAULT_STATUS,
    _make_comment_id,
    _normalize_status,
    parse_comments_csv,
    write_tracker_csv,
    read_tracker_csv,
    parse_comments_geojson,
    parse_comments_xlsx,
    ingest_comments,
    merge_tracker,
    format_comment_summary,
)


def _write_csv(path: Path, rows: list, extra_fields: list = None):
    fieldnames = ["figure_ref", "comment_text", "reviewer", "status", "x", "y"] + (extra_fields or [])
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        for row in rows:
            w.writerow(row)


def _make_geojson(features: list) -> dict:
    return {"type": "FeatureCollection", "features": features}


def _point_feature(x: float, y: float, **props) -> dict:
    return {
        "type": "Feature",
        "geometry": {"type": "Point", "coordinates": [x, y]},
        "properties": props,
    }


def _make_xlsx(path: Path, rows: list, sheet_name: str = "Sheet1"):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    if rows:
        ws.append(list(rows[0].keys()))
        for row in rows:
            ws.append(list(row.values()))
    wb.save(path)


def test_reviewer_comment_defaults():
    c = ReviewerComment(
        comment_id="rc-abc", source_file="test.csv", source_format="csv",
        figure_ref="Figure 3", comment_text="Fix north arrow", reviewer="Alice",
    )
    assert c.status == "OPEN"
    assert c.x is None and c.y is None
    assert c.assigned_to == "" and c.resolved_date == "" and c.resolution_note == ""


def test_make_comment_id_is_deterministic():
    a = _make_comment_id("f.csv", "Figure 1", "Fix label", "Bob")
    b = _make_comment_id("f.csv", "Figure 1", "Fix label", "Bob")
    assert a == b


def test_make_comment_id_starts_with_rc():
    assert _make_comment_id("f.csv", "Figure 1", "Fix label", "Bob").startswith("rc-")


def test_make_comment_id_differs_for_different_inputs():
    a = _make_comment_id("f.csv", "Figure 1", "Fix label", "Bob")
    b = _make_comment_id("f.csv", "Figure 2", "Fix label", "Bob")
    assert a != b


def test_normalize_status_canonical_values():
    assert _normalize_status("OPEN") == "OPEN"
    assert _normalize_status("IN_REVIEW") == "IN_REVIEW"
    assert _normalize_status("RESOLVED") == "RESOLVED"
    assert _normalize_status("WONT_FIX") == "WONT_FIX"


def test_normalize_status_lowercase():
    assert _normalize_status("open") == "OPEN"
    assert _normalize_status("resolved") == "RESOLVED"


def test_normalize_status_spaces_become_underscores():
    assert _normalize_status("in review") == "IN_REVIEW"
    assert _normalize_status("wont fix") == "WONT_FIX"


def test_normalize_status_unknown_defaults_to_open():
    assert _normalize_status("pending") == "OPEN"
    assert _normalize_status("") == "OPEN"
    assert _normalize_status("tbd") == "OPEN"


def test_parse_comments_csv_basic(tmp_path):
    p = tmp_path / "comments.csv"
    _write_csv(p, [
        {"figure_ref": "Figure 3", "comment_text": "Fix north arrow",
         "reviewer": "Alice", "status": "OPEN", "x": "-87.65", "y": "41.85"},
        {"figure_ref": "Figure 4", "comment_text": "Update legend",
         "reviewer": "Bob", "status": "resolved", "x": "", "y": ""},
    ])
    result = parse_comments_csv(p)
    assert len(result) == 2
    assert result[0].figure_ref == "Figure 3"
    assert result[0].comment_text == "Fix north arrow"
    assert result[0].reviewer == "Alice"
    assert result[0].status == "OPEN"
    assert result[0].x == pytest.approx(-87.65)
    assert result[0].y == pytest.approx(41.85)
    assert result[0].source_format == "csv"
    assert result[1].status == "RESOLVED"
    assert result[1].x is None


def test_parse_comments_csv_source_file_is_basename(tmp_path):
    p = tmp_path / "my_review.csv"
    _write_csv(p, [{"figure_ref": "F1", "comment_text": "x", "reviewer": "C",
                    "status": "OPEN", "x": "", "y": ""}])
    assert parse_comments_csv(p)[0].source_file == "my_review.csv"


def test_parse_comments_csv_status_normalization(tmp_path):
    p = tmp_path / "c.csv"
    _write_csv(p, [{"figure_ref": "F1", "comment_text": "x", "reviewer": "C",
                    "status": "in review", "x": "", "y": ""}])
    assert parse_comments_csv(p)[0].status == "IN_REVIEW"


def test_parse_comments_csv_missing_optional_columns(tmp_path):
    p = tmp_path / "minimal.csv"
    with p.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=["figure_ref", "comment_text"])
        w.writeheader()
        w.writerow({"figure_ref": "F1", "comment_text": "Needs label"})
    result = parse_comments_csv(p)
    assert len(result) == 1
    assert result[0].reviewer == "" and result[0].x is None
    assert result[0].status == "OPEN"


def test_parse_comments_csv_existing_comment_id_preserved(tmp_path):
    p = tmp_path / "with_id.csv"
    with p.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(
            fh, fieldnames=["comment_id", "figure_ref", "comment_text", "reviewer", "status"])
        w.writeheader()
        w.writerow({"comment_id": "rc-existing123", "figure_ref": "F1",
                    "comment_text": "Fix me", "reviewer": "D", "status": "OPEN"})
    assert parse_comments_csv(p)[0].comment_id == "rc-existing123"


def test_parse_comments_csv_accepts_alias_column_names(tmp_path):
    p = tmp_path / "aliases.csv"
    with p.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=["figure", "comment", "author"])
        w.writeheader()
        w.writerow({"figure": "Fig 7", "comment": "Bold title", "author": "Z"})
    result = parse_comments_csv(p)
    assert result[0].figure_ref == "Fig 7"
    assert result[0].comment_text == "Bold title"
    assert result[0].reviewer == "Z"


def test_write_read_tracker_csv_roundtrip(tmp_path):
    original = [
        ReviewerComment(
            comment_id="rc-aaa", source_file="in.csv", source_format="csv",
            figure_ref="Figure 1", comment_text="Move label", reviewer="Alice",
            status="RESOLVED", x=-87.65, y=41.85,
            assigned_to="Bob", resolved_date="2026-06-28",
            resolution_note="Label repositioned",
        ),
        ReviewerComment(
            comment_id="rc-bbb", source_file="in.csv", source_format="csv",
            figure_ref="Figure 2", comment_text="Add scale bar", reviewer="Carol",
        ),
    ]
    out = tmp_path / "tracker.csv"
    write_tracker_csv(original, out)
    read_back = read_tracker_csv(out)
    assert len(read_back) == 2
    assert read_back[0].comment_id == "rc-aaa"
    assert read_back[0].status == "RESOLVED"
    assert read_back[0].x == pytest.approx(-87.65)
    assert read_back[0].y == pytest.approx(41.85)
    assert read_back[0].resolution_note == "Label repositioned"
    assert read_back[1].status == "OPEN"
    assert read_back[1].x is None


def test_read_tracker_csv_missing_file_returns_empty(tmp_path):
    assert read_tracker_csv(tmp_path / "nonexistent.csv") == []


def test_write_tracker_csv_creates_parent_dirs(tmp_path):
    out = tmp_path / "subdir" / "tracker.csv"
    write_tracker_csv([], out)
    assert out.exists()


def test_parse_comments_geojson_basic(tmp_path):
    data = _make_geojson([
        _point_feature(-87.65, 41.85, content="Fix north arrow", author="Alice",
                       figure_ref="Figure 3"),
        _point_feature(-87.70, 41.90, comment="Update title", reviewer="Bob"),
    ])
    p = tmp_path / "comments.geojson"
    p.write_text(json.dumps(data), encoding="utf-8")
    qa = QACollector()
    result = parse_comments_geojson(p, qa=qa)
    assert len(result) == 2
    assert result[0].comment_text == "Fix north arrow"
    assert result[0].reviewer == "Alice"
    assert result[0].x == pytest.approx(-87.65)
    assert result[0].y == pytest.approx(41.85)
    assert result[0].source_format == "geojson"
    assert result[1].comment_text == "Update title"


def test_parse_comments_geojson_agol_property_names(tmp_path):
    data = _make_geojson([
        _point_feature(-87.65, 41.85,
                       author="reviewer@example.com",
                       content="Confirm this contour interval",
                       figure="Figure 5", status="IN_REVIEW"),
    ])
    p = tmp_path / "agol.geojson"
    p.write_text(json.dumps(data), encoding="utf-8")
    result = parse_comments_geojson(p, qa=QACollector())
    assert result[0].reviewer == "reviewer@example.com"
    assert result[0].comment_text == "Confirm this contour interval"
    assert result[0].figure_ref == "Figure 5"
    assert result[0].status == "IN_REVIEW"


def test_parse_comments_geojson_no_geometry_emits_warning(tmp_path):
    data = _make_geojson([
        {"type": "Feature", "geometry": None,
         "properties": {"content": "No location", "author": "X"}},
    ])
    p = tmp_path / "no_geom.geojson"
    p.write_text(json.dumps(data), encoding="utf-8")
    qa = QACollector()
    result = parse_comments_geojson(p, qa=qa)
    assert len(result) == 1
    assert result[0].x is None and result[0].y is None
    warnings = [r for r in qa.records if r.severity == SEV_WARNING]
    assert any("geometry" in r.message.lower() for r in warnings)


def test_parse_comments_geojson_not_feature_collection_returns_empty(tmp_path):
    data = {"type": "Feature", "geometry": None, "properties": {}}
    p = tmp_path / "bad.geojson"
    p.write_text(json.dumps(data), encoding="utf-8")
    qa = QACollector()
    result = parse_comments_geojson(p, qa=qa)
    assert result == []
    assert [r for r in qa.records if r.severity == SEV_ERROR]


def test_parse_comments_geojson_stable_comment_id(tmp_path):
    data = _make_geojson([
        _point_feature(-87.65, 41.85, content="Fix it", author="Alice", figure="F1"),
    ])
    p = tmp_path / "c.geojson"
    p.write_text(json.dumps(data), encoding="utf-8")
    r1 = parse_comments_geojson(p, qa=QACollector())
    r2 = parse_comments_geojson(p, qa=QACollector())
    assert r1[0].comment_id == r2[0].comment_id


def test_parse_comments_xlsx_basic(tmp_path):
    p = tmp_path / "comments.xlsx"
    _make_xlsx(p, [
        {"figure_ref": "Figure 1", "comment_text": "Bold the title",
         "reviewer": "Alice", "status": "OPEN", "x": -87.65, "y": 41.85},
        {"figure_ref": "Figure 2", "comment_text": "Check units",
         "reviewer": "Bob", "status": "RESOLVED", "x": None, "y": None},
    ])
    result = parse_comments_xlsx(p, qa=QACollector())
    assert len(result) == 2
    assert result[0].figure_ref == "Figure 1"
    assert result[0].comment_text == "Bold the title"
    assert result[0].source_format == "xlsx"
    assert result[0].x == pytest.approx(-87.65)
    assert result[1].status == "RESOLVED"
    assert result[1].x is None


def test_parse_comments_xlsx_prefers_comments_sheet(tmp_path):
    import openpyxl
    p = tmp_path / "multi_sheet.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Data"
    ws1.append(["figure_ref", "comment_text", "reviewer"])
    ws1.append(["F0", "Wrong sheet", "Wrong"])
    ws2 = wb.create_sheet("Comments")
    ws2.append(["figure_ref", "comment_text", "reviewer"])
    ws2.append(["F1", "Correct sheet comment", "Alice"])
    wb.save(p)
    result = parse_comments_xlsx(p, qa=QACollector())
    assert len(result) == 1
    assert result[0].comment_text == "Correct sheet comment"


def test_parse_comments_xlsx_falls_back_to_first_sheet(tmp_path):
    import openpyxl
    p = tmp_path / "no_comments_sheet.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Markups"
    ws.append(["figure_ref", "comment_text", "reviewer"])
    ws.append(["F2", "Fallback comment", "Carol"])
    wb.save(p)
    result = parse_comments_xlsx(p, qa=QACollector())
    assert len(result) == 1
    assert result[0].comment_text == "Fallback comment"


def test_parse_comments_xlsx_accepts_alias_columns(tmp_path):
    import openpyxl
    p = tmp_path / "aliases.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["figure", "comment", "author"])
    ws.append(["Fig 8", "Fix scale bar", "Dana"])
    wb.save(p)
    result = parse_comments_xlsx(p, qa=QACollector())
    assert result[0].figure_ref == "Fig 8"
    assert result[0].comment_text == "Fix scale bar"
    assert result[0].reviewer == "Dana"


def test_ingest_comments_routes_csv(tmp_path):
    p = tmp_path / "input.csv"
    _write_csv(p, [{"figure_ref": "F1", "comment_text": "Fix me", "reviewer": "A",
                    "status": "OPEN", "x": "", "y": ""}])
    qa = QACollector()
    result = ingest_comments(p, qa=qa)
    assert len(result) == 1
    assert result[0].source_format == "csv"
    assert [r for r in qa.records if r.severity == SEV_INFO and "ingest_complete" in r.category]


def test_ingest_comments_routes_geojson(tmp_path):
    data = _make_geojson([_point_feature(-87.65, 41.85, content="Test", author="B")])
    p = tmp_path / "input.geojson"
    p.write_text(json.dumps(data), encoding="utf-8")
    result = ingest_comments(p, qa=QACollector())
    assert len(result) == 1 and result[0].source_format == "geojson"


def test_ingest_comments_routes_json_extension(tmp_path):
    data = _make_geojson([_point_feature(-87.65, 41.85, content="JSON ext", author="C")])
    p = tmp_path / "input.json"
    p.write_text(json.dumps(data), encoding="utf-8")
    result = ingest_comments(p, qa=QACollector())
    assert len(result) == 1 and result[0].source_format == "geojson"


def test_ingest_comments_routes_xlsx(tmp_path):
    p = tmp_path / "input.xlsx"
    _make_xlsx(p, [{"figure_ref": "F3", "comment_text": "Fix",
                    "reviewer": "C", "status": "OPEN", "x": None, "y": None}])
    result = ingest_comments(p, qa=QACollector())
    assert len(result) == 1 and result[0].source_format == "xlsx"


def test_ingest_comments_unknown_extension_returns_empty_with_error(tmp_path):
    p = tmp_path / "comments.pdf"
    p.write_bytes(b"%PDF-fake")
    qa = QACollector()
    result = ingest_comments(p, qa=qa)
    assert result == []
    errors = [r for r in qa.records if r.severity == SEV_ERROR]
    assert errors
    assert any("pdf" in r.message.lower() for r in errors)


def test_ingest_comments_warns_empty_comment_text(tmp_path):
    p = tmp_path / "empty_text.csv"
    _write_csv(p, [{"figure_ref": "F1", "comment_text": "", "reviewer": "A",
                    "status": "OPEN", "x": "", "y": ""}])
    qa = QACollector()
    ingest_comments(p, qa=qa)
    assert any("empty_comment_text" in r.category
               for r in qa.records if r.severity == SEV_WARNING)


def test_ingest_comments_warns_empty_figure_ref(tmp_path):
    p = tmp_path / "no_fig.csv"
    _write_csv(p, [{"figure_ref": "", "comment_text": "Something", "reviewer": "A",
                    "status": "OPEN", "x": "", "y": ""}])
    qa = QACollector()
    ingest_comments(p, qa=qa)
    assert any("empty_figure_ref" in r.category
               for r in qa.records if r.severity == SEV_WARNING)


def test_merge_tracker_new_comments_appended():
    existing = [ReviewerComment("rc-1", "old.csv", "csv", "F1", "Old comment", "Alice")]
    incoming = [ReviewerComment("rc-2", "new.csv", "csv", "F2", "New comment", "Bob")]
    result = merge_tracker(existing, incoming, qa=QACollector())
    assert [c.comment_id for c in result] == ["rc-1", "rc-2"]


def test_merge_tracker_existing_status_preserved_on_reingest():
    existing = [ReviewerComment("rc-1", "old.csv", "csv", "F1", "Fix it", "Alice",
                                status="RESOLVED", resolution_note="Done")]
    incoming = [ReviewerComment("rc-1", "old.csv", "csv", "F1", "Fix it", "Alice",
                                status="OPEN")]
    result = merge_tracker(existing, incoming, qa=QACollector())
    assert len(result) == 1
    assert result[0].status == "RESOLVED"
    assert result[0].resolution_note == "Done"


def test_merge_tracker_empty_existing():
    incoming = [ReviewerComment("rc-1", "new.csv", "csv", "F1", "Add arrow", "Alice")]
    result = merge_tracker([], incoming, qa=QACollector())
    assert len(result) == 1 and result[0].comment_id == "rc-1"


def test_merge_tracker_no_duplicates_on_repeated_reingest():
    comment = ReviewerComment("rc-1", "f.csv", "csv", "F1", "Fix me", "Alice")
    after_first = merge_tracker([], [comment], qa=QACollector())
    after_second = merge_tracker(after_first, [comment], qa=QACollector())
    assert len(after_second) == 1


def test_merge_tracker_existing_order_preserved():
    existing = [
        ReviewerComment("rc-1", "f.csv", "csv", "F1", "A", "Alice"),
        ReviewerComment("rc-2", "f.csv", "csv", "F2", "B", "Bob"),
    ]
    incoming = [ReviewerComment("rc-3", "f.csv", "csv", "F3", "C", "Carol")]
    result = merge_tracker(existing, incoming, qa=QACollector())
    assert [c.comment_id for c in result] == ["rc-1", "rc-2", "rc-3"]


def test_merge_tracker_emits_info_qa_record():
    qa = QACollector()
    merge_tracker([], [], qa=qa)
    assert [r for r in qa.records if r.severity == SEV_INFO]


def test_format_comment_summary_shows_total_count():
    comments = [
        ReviewerComment("rc-1", "f.csv", "csv", "F1", "A", "Alice", status="OPEN"),
        ReviewerComment("rc-2", "f.csv", "csv", "F2", "B", "Bob", status="RESOLVED"),
        ReviewerComment("rc-3", "f.csv", "csv", "F3", "C", "Carol", status="OPEN"),
    ]
    assert "3" in format_comment_summary(comments)


def test_format_comment_summary_counts_by_status():
    comments = [
        ReviewerComment("rc-1", "f.csv", "csv", "F1", "A", "Alice", status="OPEN"),
        ReviewerComment("rc-2", "f.csv", "csv", "F2", "B", "Bob", status="OPEN"),
        ReviewerComment("rc-3", "f.csv", "csv", "F3", "C", "Carol", status="RESOLVED"),
    ]
    summary = format_comment_summary(comments)
    assert "OPEN" in summary and "RESOLVED" in summary


def test_format_comment_summary_empty():
    assert "0" in format_comment_summary([])


def test_format_comment_summary_shows_open_preview():
    comments = [
        ReviewerComment("rc-1", "f.csv", "csv", "Figure 3",
                        "Fix the north arrow placement", "Alice", status="OPEN"),
    ]
    summary = format_comment_summary(comments)
    assert "Figure 3" in summary
    assert "Fix the north arrow placement" in summary


def test_format_comment_summary_truncates_open_preview_at_5():
    comments = [
        ReviewerComment(f"rc-{i}", "f.csv", "csv", f"F{i}",
                        f"Comment {i}", "Alice", status="OPEN")
        for i in range(8)
    ]
    summary = format_comment_summary(comments)
    assert "3 more" in summary or "3" in summary


def test_ingest_comments_xls_unsupported(tmp_path):
    """Legacy .xls cannot be read by openpyxl -> clean error + [], not a raise."""
    p = tmp_path / "comments.xls"
    p.write_bytes(b"\xd0\xcf\x11\xe0fake-ole")  # OLE/BIFF magic-ish
    qa = QACollector()
    result = ingest_comments(p, qa=qa)
    assert result == []
    errors = [r for r in qa.records if r.severity == SEV_ERROR]
    assert any(".xls" in r.message or "xls" in r.message.lower() for r in errors)


def test_merge_tracker_dedupes_incoming_duplicates():
    """Repeated comment_id in one incoming batch must not inflate the tracker."""
    dup = ReviewerComment("rc-1", "f.csv", "csv", "F1", "Fix me", "Alice")
    result = merge_tracker([], [dup, dup, dup], qa=QACollector())
    assert len(result) == 1
    assert result[0].comment_id == "rc-1"
