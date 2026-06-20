"""Shared fixtures: a synthetic workbook that mimics the H281 layout the
spec describes (titles, group header row, analyte header row, units row,
RBSL screening row with a footnote, formula GWE without a cached value,
Dry/NM statuses, qualifiers, an exceedance) plus the matching test profile.

The synthetic workbook is NOT the real H281 file — anchors here are chosen
by the tests, so these tests validate the engine, not the production
parser profile (which stays DRAFT until run against the real workbook).
"""

from datetime import datetime
from pathlib import Path

import pytest

import autogis

CONFIG = Path(autogis.__file__).resolve().parent / "config"
DATA = Path(__file__).resolve().parent / "data"

EVENT = datetime(2026, 4, 15)


def _build_workbook(path: Path) -> Path:
    import openpyxl
    wb = openpyxl.Workbook()

    gw = wb.active
    gw.title = "GW Table 2"
    gw["A1"] = "TABLE 2"
    gw["A2"] = "GROUNDWATER ANALYTICAL RESULTS AND WATER LEVELS"
    gw["A3"] = "H281 Glasgow"
    gw["G5"] = "VOLATILE ORGANIC COMPOUNDS"
    gw.merge_cells("G5:L5")
    headers = ["Well ID", "Screen Interval", "Date", "MPE", "DTW", "GWE",
               "Benzene", "Toluene", "Ethylbenzene", "Total Xylenes",
               "MTBE", "Total Purgeable Hydrocarbons"]
    for i, h in enumerate(headers, start=1):
        gw.cell(row=7, column=i, value=h)
    for c in range(7, 13):
        gw.cell(row=8, column=c, value="ug/L")
    gw["A9"] = "MDEQ RBSL"
    gw["G9"] = "5"
    gw["L9"] = "1,000 3"      # screening level with a footnote marker

    rows = [
        # MW-1: GWE is a formula with NO cached value -> QA error expected
        ["MW-1", "15-25", EVENT, 2105.50, 16.20, "=D10-E10",
         "12 J", "<1.0", "<0.50 U", "150", "NA", "250"],
        ["MW-2", "12-22", EVENT, 2103.20, 13.90, 2089.30,
         "<1.0", "<1.0", "<1.0", "<2.0", "<1.0", "<50"],
        ["MW-3", "10-20", EVENT, 2101.00, "Dry", "Dry",
         "NS", "NS", "NS", "NS", "NS", "NS"],
        ["MW-4", "11-21", EVENT, 2102.40, "NM", "NM",
         "NM", "NM", "NM", "NM", "NM", "NM"],
        # MW-5: benzene 6.4 exceeds the RBSL of 5
        ["MW-5", "14-24", EVENT, 2104.80, 14.70, 2090.10,
         "6.4", "<1.0", "<1.0", "<2.0", "<1.0", "120"],
    ]
    for r, vals in enumerate(rows, start=10):
        for c, v in enumerate(vals, start=1):
            gw.cell(row=r, column=c, value=v)

    ibi = wb.create_sheet("Table 3")
    ibi["A1"] = "TABLE 3"
    ibi["A2"] = "INORGANIC INDICATOR PARAMETERS"
    for i, h in enumerate(["Well ID", "Date", "Sulfate",
                           "Nitrate+Nitrite as N", "Iron", "Manganese",
                           "Methane"], start=1):
        ibi.cell(row=3, column=i, value=h)
    for c in range(3, 8):
        ibi.cell(row=4, column=c, value="mg/L")
    for r, vals in enumerate([
            ["MW-1", EVENT, "120", "0.5 J", "1.2", "<0.05", "NA"],
            ["MW-2", EVENT, "<10", "<0.1", "0.8", "0.3", "0.002"]],
            start=5):
        for c, v in enumerate(vals, start=1):
            ibi.cell(row=r, column=c, value=v)

    met = wb.create_sheet("Metals Table")
    met["A1"] = "DISSOLVED METALS"
    for i, h in enumerate(["Well ID", "Date", "Arsenic", "Lead", "Zinc"],
                          start=1):
        met.cell(row=3, column=i, value=h)
    for c in range(3, 6):
        met.cell(row=4, column=c, value="ug/L")
    for r, vals in enumerate([
            ["MW-1", EVENT, "<0.010", "0.005 J", "12"],
            ["MW-2", EVENT, "2.1", "<1.0", "<5"]], start=5):
        for c, v in enumerate(vals, start=1):
            met.cell(row=r, column=c, value=v)

    rpd = wb.create_sheet("RPD")
    rpd["A1"] = "FIELD DUPLICATE RPD"
    rpd["B2"] = EVENT
    for i, h in enumerate(["Analyte", "Parent", "Duplicate", "RPD", "RL"],
                          start=1):
        rpd.cell(row=3, column=i, value=h)
    for r, vals in enumerate([
            ["Benzene", "6.4", "6.0", "6.5", "1.0"],     # consistent
            ["Toluene", "<1.0", "<1.0", "NC", "1.0"],    # nondetect pair
            ["MTBE", "10", "8", "30", "1.0"]],           # workbook RPD wrong
            start=4):
        for c, v in enumerate(vals, start=1):
            rpd.cell(row=r, column=c, value=v)

    wb.save(path)
    return path


@pytest.fixture(scope="session")
def workbook(tmp_path_factory) -> Path:
    return _build_workbook(
        tmp_path_factory.mktemp("wb") / "H281_synthetic.xlsx")


@pytest.fixture(scope="session")
def profile():
    from autogis.core.common.config import ParserProfile
    return ParserProfile.load(DATA / "H281_test_profile.yaml")


@pytest.fixture(scope="session")
def adict():
    from autogis.core.common.config import load_analyte_dictionary
    return load_analyte_dictionary(
        CONFIG / "analytes" / "analyte_dictionary.yaml")


@pytest.fixture(scope="session")
def slevels():
    from autogis.core.common.config import load_screening_levels
    return load_screening_levels(
        CONFIG / "screening_levels" / "screening_levels.yaml")


@pytest.fixture()
def qa():
    from autogis.core.common.qa import QACollector
    return QACollector()
