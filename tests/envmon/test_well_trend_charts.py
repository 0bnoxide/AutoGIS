"""Tests for well_trend_charts (Tool 4.6)."""
from __future__ import annotations

import csv
from pathlib import Path

import openpyxl
import pytest

from click.testing import CliRunner
from autogis.adapters.cli import autogis as cli_root
from autogis.core.envmon.well_trend_charts import (
    TrendSeries,
    load_history_csv,
    write_trend_charts,
)


def _write_history_csv(path: Path, rows: list[dict]) -> Path:
    fieldnames = [
        "LocationID", "AnalyteName", "SampleDate",
        "ResultValue", "ReportedUnits", "ScreeningLevel",
    ]
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in fieldnames})
    return path


# --- load_history_csv -------------------------------------------------------

def test_load_basic(tmp_path):
    csv_path = _write_history_csv(tmp_path / "h.csv", [
        {"LocationID": "MW-1", "AnalyteName": "Benzene",
         "SampleDate": "2026-01-01", "ResultValue": "5.0",
         "ReportedUnits": "ug/L", "ScreeningLevel": "1.0"},
        {"LocationID": "MW-1", "AnalyteName": "Benzene",
         "SampleDate": "2026-04-01", "ResultValue": "8.0",
         "ReportedUnits": "ug/L", "ScreeningLevel": "1.0"},
    ])
    result = load_history_csv(csv_path)
    assert len(result) == 1
    s = result[0]
    assert s.location_id == "MW-1"
    assert s.analyte_name == "Benzene"
    assert s.values == [5.0, 8.0]
    assert s.dates == ["2026-01-01", "2026-04-01"]
    assert s.units == "ug/L"
    assert s.screening_level == pytest.approx(1.0)


def test_load_nd_rows_excluded(tmp_path):
    csv_path = _write_history_csv(tmp_path / "h.csv", [
        {"LocationID": "MW-1", "AnalyteName": "Benzene",
         "SampleDate": "2026-01-01", "ResultValue": "ND"},
        {"LocationID": "MW-1", "AnalyteName": "Benzene",
         "SampleDate": "2026-04-01", "ResultValue": ""},
        {"LocationID": "MW-1", "AnalyteName": "Benzene",
         "SampleDate": "2026-07-01", "ResultValue": "3.5"},
    ])
    result = load_history_csv(csv_path)
    assert len(result) == 1
    assert result[0].values == [3.5]


def test_load_multiple_analytes(tmp_path):
    csv_path = _write_history_csv(tmp_path / "h.csv", [
        {"LocationID": "MW-1", "AnalyteName": "Benzene",
         "SampleDate": "2026-01-01", "ResultValue": "5.0"},
        {"LocationID": "MW-1", "AnalyteName": "Toluene",
         "SampleDate": "2026-01-01", "ResultValue": "2.0"},
    ])
    result = load_history_csv(csv_path)
    assert len(result) == 2
    assert {s.analyte_name for s in result} == {"Benzene", "Toluene"}


def test_load_multiple_locations(tmp_path):
    csv_path = _write_history_csv(tmp_path / "h.csv", [
        {"LocationID": "MW-1", "AnalyteName": "Benzene",
         "SampleDate": "2026-01-01", "ResultValue": "5.0"},
        {"LocationID": "MW-2", "AnalyteName": "Benzene",
         "SampleDate": "2026-01-01", "ResultValue": "3.0"},
    ])
    result = load_history_csv(csv_path)
    assert {s.location_id for s in result} == {"MW-1", "MW-2"}


def test_load_sorted_by_date(tmp_path):
    csv_path = _write_history_csv(tmp_path / "h.csv", [
        {"LocationID": "MW-1", "AnalyteName": "Benzene",
         "SampleDate": "2026-07-01", "ResultValue": "9.0"},
        {"LocationID": "MW-1", "AnalyteName": "Benzene",
         "SampleDate": "2026-01-01", "ResultValue": "5.0"},
    ])
    result = load_history_csv(csv_path)
    assert result[0].dates == ["2026-01-01", "2026-07-01"]


def test_load_no_screening_level(tmp_path):
    csv_path = _write_history_csv(tmp_path / "h.csv", [
        {"LocationID": "MW-1", "AnalyteName": "Benzene",
         "SampleDate": "2026-01-01", "ResultValue": "5.0",
         "ScreeningLevel": ""},
    ])
    result = load_history_csv(csv_path)
    assert result[0].screening_level is None


# --- write_trend_charts -----------------------------------------------------

def _make_series(loc="MW-1", analyte="Benzene", n=3, sl=None) -> TrendSeries:
    dates = [f"2026-0{i+1}-01" for i in range(n)]
    values = [float(i + 1) for i in range(n)]
    return TrendSeries(
        location_id=loc, analyte_name=analyte,
        dates=dates, values=values, screening_level=sl, units="ug/L",
    )


def test_write_creates_xlsx(tmp_path):
    out = tmp_path / "charts.xlsx"
    count = write_trend_charts([_make_series()], out)
    assert out.exists()
    assert count == 1


def test_write_sheet_named_after_analyte(tmp_path):
    out = tmp_path / "charts.xlsx"
    write_trend_charts([_make_series(analyte="Benzene")], out)
    assert "Benzene" in openpyxl.load_workbook(str(out)).sheetnames


def test_write_chart_count_multiple_locations(tmp_path):
    out = tmp_path / "charts.xlsx"
    count = write_trend_charts([
        _make_series(loc="MW-1"), _make_series(loc="MW-2"),
    ], out)
    assert count == 2


def test_write_multiple_analytes_multiple_sheets(tmp_path):
    out = tmp_path / "charts.xlsx"
    write_trend_charts([
        _make_series(analyte="Benzene"), _make_series(analyte="Toluene"),
    ], out)
    wb = openpyxl.load_workbook(str(out))
    assert "Benzene" in wb.sheetnames and "Toluene" in wb.sheetnames


def test_write_single_point_series_no_crash(tmp_path):
    out = tmp_path / "charts.xlsx"
    s = TrendSeries(location_id="MW-1", analyte_name="Benzene",
                    dates=["2026-01-01"], values=[5.0],
                    screening_level=None, units="ug/L")
    count = write_trend_charts([s], out)
    assert out.exists() and count == 1


def test_write_pagination(tmp_path):
    out = tmp_path / "charts.xlsx"
    series = [_make_series(loc=f"MW-{i}") for i in range(5)]
    write_trend_charts(series, out, max_per_sheet=3)
    names = openpyxl.load_workbook(str(out)).sheetnames
    assert any("Benzene" in n for n in names)
    assert len(names) == 2


def test_write_chart_present_in_sheet(tmp_path):
    out = tmp_path / "charts.xlsx"
    write_trend_charts([_make_series()], out)
    ws = openpyxl.load_workbook(str(out)).active
    assert len(ws._charts) >= 1


def test_write_empty_series_list(tmp_path):
    out = tmp_path / "charts.xlsx"
    count = write_trend_charts([], out)
    assert count == 0
    assert out.exists()


def _long_series(loc, n=24):
    """A series with n points and valid monthly ISO dates (n can exceed 9)."""
    dates, values = [], []
    for i in range(n):
        y, m = 2021 + i // 12, i % 12 + 1
        dates.append(f"{y}-{m:02d}-01")
        values.append(float(i + 1))
    return TrendSeries(location_id=loc, analyte_name="Benzene",
                       dates=dates, values=values, screening_level=None,
                       units="ug/L")


def test_long_series_does_not_overwrite_next_block(tmp_path):
    """A series with >= _BLOCK_ROWS (20) points must not have its data clobbered
    by the next stacked series' header (regression: fixed-20-row blocks)."""
    out = tmp_path / "charts.xlsx"
    s1 = _long_series("MW-1", n=24)   # 24 points -> spills past a 20-row block
    s2 = _make_series(loc="MW-2", analyte="Benzene", n=3)
    write_trend_charts([s1, s2], out)
    ws = openpyxl.load_workbook(str(out))["Benzene"]
    # MW-1 header at row 1, its 24 data rows at rows 2..25 must be intact.
    assert ws.cell(row=1, column=2).value == "MW-1"
    last_val = ws.cell(row=25, column=2).value   # 24th data point
    assert last_val == 24.0
    # MW-2's header must land strictly below MW-1's data block, not on top of it.
    mw2_header_rows = [r for r in range(2, ws.max_row + 1)
                       if ws.cell(row=r, column=2).value == "MW-2"]
    assert mw2_header_rows and min(mw2_header_rows) >= 26


# --- CLI --------------------------------------------------------------------

def test_generate_trend_charts_in_help():
    result = CliRunner().invoke(cli_root, ["envmon", "--help"])
    assert "generate-trend-charts" in result.output


def test_cli_end_to_end(tmp_path):
    csv_path = tmp_path / "history.csv"
    _write_history_csv(csv_path, [
        {"LocationID": "MW-1", "AnalyteName": "Benzene",
         "SampleDate": "2026-01-01", "ResultValue": "5.0",
         "ReportedUnits": "ug/L", "ScreeningLevel": "1.0"},
        {"LocationID": "MW-1", "AnalyteName": "Benzene",
         "SampleDate": "2026-04-01", "ResultValue": "8.0",
         "ReportedUnits": "ug/L", "ScreeningLevel": "1.0"},
    ])
    out_path = tmp_path / "trends.xlsx"
    result = CliRunner().invoke(cli_root, [
        "envmon", "generate-trend-charts",
        "--history-csv", str(csv_path), "--out", str(out_path),
    ])
    assert result.exit_code == 0, result.output
    assert out_path.exists()
    assert "Benzene" in openpyxl.load_workbook(str(out_path)).sheetnames


def test_cli_analyte_filter(tmp_path):
    csv_path = tmp_path / "history.csv"
    _write_history_csv(csv_path, [
        {"LocationID": "MW-1", "AnalyteName": "Benzene",
         "SampleDate": "2026-01-01", "ResultValue": "5.0"},
        {"LocationID": "MW-1", "AnalyteName": "Toluene",
         "SampleDate": "2026-01-01", "ResultValue": "2.0"},
    ])
    out_path = tmp_path / "trends.xlsx"
    result = CliRunner().invoke(cli_root, [
        "envmon", "generate-trend-charts",
        "--history-csv", str(csv_path), "--out", str(out_path),
        "--analytes", "Benzene",
    ])
    assert result.exit_code == 0, result.output
    wb = openpyxl.load_workbook(str(out_path))
    assert "Benzene" in wb.sheetnames and "Toluene" not in wb.sheetnames
