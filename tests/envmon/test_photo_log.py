from pathlib import Path

import pytest

from autogis.core.envmon.photo_metadata import PhotoRecord
from autogis.core.envmon.photo_log import write_log


def _rec(path, **kw):
    base = dict(objectid=2, attachment_id=7, source_table="Obs",
                group="G/SeepSpring", saved_path=str(path),
                exif_lat=45.874, exif_lon=-103.487, heading_deg=231.5,
                heading_ref="T", taken_at="2026-05-05T08:17:36",
                camera="samsung SM-X308U")
    base.update(kw)
    return PhotoRecord(**base)


def test_log_xlsx(tmp_path, make_photo_jpeg):
    pytest.importorskip("openpyxl")
    p = make_photo_jpeg(directory=tmp_path / "G")
    out = tmp_path / "log.xlsx"
    n = write_log([_rec(p)], out, fmt="xlsx")
    assert n == 1
    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "Photo #"
    assert ws.cell(row=2, column=1).value == 1
    assert "SW" in ws.cell(row=2, column=5).value      # 231.5° -> SW
    assert ws.cell(row=2, column=7).value in (None, "")  # blank Description


def test_log_html(tmp_path, make_photo_jpeg):
    p = make_photo_jpeg(directory=tmp_path / "G")
    out = tmp_path / "log.html"
    n = write_log([_rec(p)], out, fmt="html", title="RILEY PASS photos")
    assert n == 1
    html = out.read_text(encoding="utf-8")
    assert "RILEY PASS photos" in html
    assert "data:image/jpeg;base64," in html
    assert "232° SW" in html and "45.874" in html


def test_log_docx(tmp_path, make_photo_jpeg):
    pytest.importorskip("docx")
    p = make_photo_jpeg(directory=tmp_path / "G")
    out = tmp_path / "log.docx"
    n = write_log([_rec(p)], out, fmt="docx")
    assert n == 1
    import docx
    d = docx.Document(str(out))
    text = "\n".join(par.text for par in d.paragraphs)
    assert "Photo 1" in text and "Description:" in text


def test_log_unknown_format(tmp_path):
    with pytest.raises(ValueError):
        write_log([], tmp_path / "x.pdf", fmt="pdf")


def test_log_photo_without_gps_or_file(tmp_path):
    # No GPS, file missing on disk: still a log row, blank coords, no crash.
    out = tmp_path / "log.html"
    n = write_log([_rec(tmp_path / "gone.jpg", exif_lat=None, exif_lon=None,
                        heading_deg=None, taken_at=None)], out, fmt="html")
    assert n == 1
