from pathlib import Path
import pytest
import openpyxl
from autogis.core.envmon.regulatory_table_builder import (
    RegulatoryTableSpec, build_regulatory_table_specs,
    write_regulatory_workbook,
)

_ROWS = [
    {"LocationID": "MW-01", "AnalyteName": "Benzene",
     "ResultValue": "12.0", "ResultQualifier": "", "ReportedUnits": "ug/L",
     "SampleDate": "2026-01-15"},
    {"LocationID": "MW-01", "AnalyteName": "Toluene",
     "ResultValue": "ND", "ResultQualifier": "ND", "ReportedUnits": "ug/L",
     "SampleDate": "2026-01-15"},
    {"LocationID": "MW-02", "AnalyteName": "Benzene",
     "ResultValue": "2.0", "ResultQualifier": "", "ReportedUnits": "ug/L",
     "SampleDate": "2026-01-15"},
]
_SL = {"Benzene": 5.0, "Toluene": 1000.0}
_GROUPS = {"Benzene": "VOC", "Toluene": "VOC"}


def test_build_specs_groups_analytes():
    specs = build_regulatory_table_specs(_ROWS, group_map=_GROUPS, screening_levels=_SL)
    assert len(specs) == 1
    assert specs[0].analyte_group == "VOC"
    assert "Benzene" in specs[0].analytes


def test_write_workbook_produces_xlsx(tmp_path):
    specs = build_regulatory_table_specs(_ROWS, group_map=_GROUPS, screening_levels=_SL)
    out = tmp_path / "reg_tables.xlsx"
    result = write_regulatory_workbook(_ROWS, specs, out, screening_levels=_SL)
    assert out.exists()


def test_exceedance_marker_in_cell(tmp_path):
    specs = build_regulatory_table_specs(_ROWS, group_map=_GROUPS, screening_levels=_SL)
    out = tmp_path / "reg_tables.xlsx"
    result = write_regulatory_workbook(_ROWS, specs, out, screening_levels=_SL)
    wb = openpyxl.load_workbook(str(out))
    ws = wb["VOC"]
    found_marker = any(
        "**" in str(cell.value)
        for row in ws.iter_rows()
        for cell in row
        if cell.value
    )
    assert found_marker


def test_nd_text_in_cell(tmp_path):
    specs = build_regulatory_table_specs(_ROWS, group_map=_GROUPS, screening_levels=_SL)
    out = tmp_path / "reg_tables.xlsx"
    write_regulatory_workbook(_ROWS, specs, out, screening_levels=_SL)
    wb = openpyxl.load_workbook(str(out))
    ws = wb["VOC"]
    found_nd = any(
        str(cell.value) == "ND"
        for row in ws.iter_rows()
        for cell in row
        if cell.value
    )
    assert found_nd


def test_mcl_row_present(tmp_path):
    specs = build_regulatory_table_specs(_ROWS, group_map=_GROUPS, screening_levels=_SL)
    out = tmp_path / "reg_tables.xlsx"
    write_regulatory_workbook(_ROWS, specs, out, screening_levels=_SL)
    wb = openpyxl.load_workbook(str(out))
    ws = wb["VOC"]
    values = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value]
    assert "5.0" in values


def test_exceedance_count(tmp_path):
    specs = build_regulatory_table_specs(_ROWS, group_map=_GROUPS, screening_levels=_SL)
    out = tmp_path / "reg_tables.xlsx"
    result = write_regulatory_workbook(_ROWS, specs, out, screening_levels=_SL)
    assert result.exceedance_count == 1  # MW-01 Benzene 12.0 > 5.0


def test_footnotes_sheet_present(tmp_path):
    specs = build_regulatory_table_specs(_ROWS, group_map=_GROUPS, screening_levels=_SL)
    out = tmp_path / "reg_tables.xlsx"
    write_regulatory_workbook(_ROWS, specs, out, screening_levels=_SL)
    wb = openpyxl.load_workbook(str(out))
    assert any("Footnote" in s or "footnote" in s.lower() for s in wb.sheetnames)


def test_missing_screening_level_no_crash(tmp_path):
    specs = build_regulatory_table_specs(_ROWS, group_map=_GROUPS)
    out = tmp_path / "reg_tables.xlsx"
    result = write_regulatory_workbook(_ROWS, specs, out)
    assert out.exists()
    assert result.exceedance_count == 0
