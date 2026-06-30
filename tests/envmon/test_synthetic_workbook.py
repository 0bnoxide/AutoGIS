"""Unit tests for synthetic_workbook (Tool 10.6)."""
import openpyxl

from autogis.core.envmon.result_parser import parse_result_value
from autogis.core.envmon.synthetic_workbook import (
    UNKNOWN_WELL_ID, WorkbookScenario, generate_workbook)


def _scn(features=(), seed=0, wells=4, events=2):
    return WorkbookScenario(site_id="TEST01", n_wells=wells, n_events=events,
                            features=set(features), seed=seed)


def _all_cells(path):
    wb = openpyxl.load_workbook(path)
    grid = {}
    for ws in wb.worksheets:
        grid[ws.title] = [[c.value for c in row] for row in ws.iter_rows()]
    return grid


def test_same_seed_is_deterministic(tmp_path):
    a = generate_workbook(_scn(features=("nondetects", "metals"), seed=7),
                          tmp_path / "a.xlsx")
    b = generate_workbook(_scn(features=("nondetects", "metals"), seed=7),
                          tmp_path / "b.xlsx")
    assert _all_cells(a) == _all_cells(b)


def test_different_seed_differs(tmp_path):
    a = generate_workbook(_scn(seed=1), tmp_path / "a.xlsx")
    b = generate_workbook(_scn(seed=2), tmp_path / "b.xlsx")
    assert _all_cells(a) != _all_cells(b)


def test_merged_headers_produces_merged_range(tmp_path):
    out = generate_workbook(_scn(features=("merged_headers",)),
                            tmp_path / "m.xlsx")
    wb = openpyxl.load_workbook(out)
    assert len(wb["Results"].merged_cells.ranges) >= 1


def test_formula_error_injects_value_error(tmp_path):
    out = generate_workbook(_scn(features=("formula_error",)),
                            tmp_path / "fe.xlsx")
    vals = {c.value for row in openpyxl.load_workbook(out)["Results"].iter_rows()
            for c in row}
    assert "#VALUE!" in vals


def test_nondetects_have_lt_prefix_and_u_qualifier(tmp_path):
    out = generate_workbook(_scn(features=("nondetects",)), tmp_path / "nd.xlsx")
    grid = _all_cells(out)["Results"]
    header = grid[0] if "merged" not in str(grid[0]) else grid[1]
    res_col = header.index("Result")
    qual_col = header.index("Qualifier")
    body = grid[1:]
    assert any(str(r[res_col]).startswith("<") for r in body if r[res_col])
    assert any(str(r[qual_col]).strip().upper() == "U"
               for r in body if r[qual_col])


def test_unknown_well_present(tmp_path):
    out = generate_workbook(_scn(features=("unknown_well",)), tmp_path / "uw.xlsx")
    vals = {c.value for row in openpyxl.load_workbook(out)["Results"].iter_rows()
            for c in row}
    assert UNKNOWN_WELL_ID in vals


def test_rpd_sheet_added(tmp_path):
    out = generate_workbook(_scn(features=("rpd_sheet",)), tmp_path / "rpd.xlsx")
    assert "RPD" in openpyxl.load_workbook(out).sheetnames


def test_clean_baseline_result_values_parse_cleanly(tmp_path):
    """Baseline (no messiness) result values parse numerically with no warning."""
    out = generate_workbook(_scn(features=()), tmp_path / "base.xlsx")
    grid = _all_cells(out)["Results"]
    header = grid[0]
    res_col = header.index("Result")
    parsed = [parse_result_value(r[res_col]) for r in grid[1:] if r[res_col] is not None]
    assert parsed  # there are result rows
    assert all(p.is_numeric and p.parse_warning == "" for p in parsed)
