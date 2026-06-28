"""Tests for export_summary_tables module."""
from __future__ import annotations

import datetime as dt
from datetime import date
import pytest
import openpyxl

from autogis.core.envmon.export_summary_tables import _build_current_event, _build_gw_by_event, _build_soil_by_depth, export_summary_tables
from autogis.core.envmon.gdb_schema import AnalyticalResultRecord
from autogis.core.common.qa import QACollector


# ============================================================================
# Shared Fixtures
# ============================================================================

def _r(
    LocationID: str,
    AnalyteCanonicalName: str,
    SampleDate: date = None,
    DisplayText: str = "--",
    ExceedsScreeningLevel: int = 0,
    Matrix: str = "GW",
    **kwargs
) -> AnalyticalResultRecord:
    """Helper to build minimal AnalyticalResultRecord for testing."""
    defaults = {
        "ImportBatchID": "TEST_BATCH",
        "SiteID": "TEST_SITE",
        "Matrix": Matrix,
        "LocationID": LocationID,
        "SampleID": f"SAMPLE_{LocationID}_{AnalyteCanonicalName}",
        "ParentSampleID": "",
        "SampleDate": SampleDate or date(2026, 4, 1),
        "DepthTop_ft": None,
        "DepthBottom_ft": None,
        "DepthIntervalText": "",
        "AnalyticalGroup": "VOC",
        "MethodGroup": "EPA8260",
        "AnalyteName": AnalyteCanonicalName,
        "AnalyteCanonicalName": AnalyteCanonicalName,
        "AnalyteAbbreviation": AnalyteCanonicalName[:3],
        "ResultRawText": DisplayText,
        "ResultNumeric": None,
        "ReportingLimit": None,
        "DetectionLimit": None,
        "Units": "ug/L",
        "Qualifier": "",
        "IsNonDetect": 0,
        "IsDetected": 1,
        "IsEstimated": 0,
        "IsDiluted": 0,
        "IsNotAnalyzed": 0,
        "IsNotSampled": 0,
        "IsNotMeasured": 0,
        "ScreeningLevel": 5.0,
        "ScreeningLevelSource": "RBSL",
        "ExceedsScreeningLevel": ExceedsScreeningLevel,
        "DisplayText": DisplayText,
        "DisplayColorClass": "EXCEED" if ExceedsScreeningLevel else "OK",
        "SourceWorkbook": "test.xlsx",
        "SourceSheet": "Sheet1",
        "SourceRow": 1,
        "SourceColumn": "A",
        "SourceCell": "A1",
    }
    defaults.update(kwargs)
    return AnalyticalResultRecord(**defaults)


# GW records: 3 locations (MW-1, MW-2, MW-3), 3 analytes (Benzene, MTBE, Toluene)
# from two dates (2025-10-01 and 2026-04-01), with MW-1/Benzene exceeding screening level
GW_RECORDS = [
    # MW-1 records from 2025-10-01
    _r("MW-1", "Benzene", SampleDate=date(2025, 10, 1), DisplayText="3.2"),
    _r("MW-1", "MTBE", SampleDate=date(2025, 10, 1), DisplayText="<1.0"),
    _r("MW-1", "Toluene", SampleDate=date(2025, 10, 1), DisplayText="<0.5"),
    # MW-2 records from 2025-10-01
    _r("MW-2", "Benzene", SampleDate=date(2025, 10, 1), DisplayText="<1.0"),
    _r("MW-2", "MTBE", SampleDate=date(2025, 10, 1), DisplayText="<1.0"),
    _r("MW-2", "Toluene", SampleDate=date(2025, 10, 1), DisplayText="<1.0"),
    # MW-3 records from 2025-10-01
    _r("MW-3", "Benzene", SampleDate=date(2025, 10, 1), DisplayText="<1.0"),
    _r("MW-3", "MTBE", SampleDate=date(2025, 10, 1), DisplayText="<1.0"),
    _r("MW-3", "Toluene", SampleDate=date(2025, 10, 1), DisplayText="<1.0"),
    # MW-1 records from 2026-04-01
    _r("MW-1", "Benzene", SampleDate=date(2026, 4, 1), DisplayText="5.5", ExceedsScreeningLevel=1),
    _r("MW-1", "MTBE", SampleDate=date(2026, 4, 1), DisplayText="<1.0"),
    _r("MW-1", "Toluene", SampleDate=date(2026, 4, 1), DisplayText="<0.5"),
    # MW-2 records from 2026-04-01
    _r("MW-2", "Benzene", SampleDate=date(2026, 4, 1), DisplayText="<1.0"),
    _r("MW-2", "MTBE", SampleDate=date(2026, 4, 1), DisplayText="<1.0"),
    _r("MW-2", "Toluene", SampleDate=date(2026, 4, 1), DisplayText="<1.0"),
    # MW-3 records from 2026-04-01
    _r("MW-3", "Benzene", SampleDate=date(2026, 4, 1), DisplayText="<1.0"),
    _r("MW-3", "MTBE", SampleDate=date(2026, 4, 1), DisplayText="<1.0"),
    _r("MW-3", "Toluene", SampleDate=date(2026, 4, 1), DisplayText="<1.0"),
]

# Soil records: 2 locs × 2 depths × 2 analytes
SOIL_RECORDS = [
    # SB-1, 0-2 ft
    _r("SB-1", "Benzene", DisplayText="10", Matrix="SOIL", DepthIntervalText="0-2 ft"),
    _r("SB-1", "Naphthalene", DisplayText="5", Matrix="SOIL", DepthIntervalText="0-2 ft"),
    # SB-1, 2-4 ft
    _r("SB-1", "Benzene", DisplayText="8", Matrix="SOIL", DepthIntervalText="2-4 ft"),
    _r("SB-1", "Naphthalene", DisplayText="3", Matrix="SOIL", DepthIntervalText="2-4 ft"),
    # SB-2, 0-2 ft
    _r("SB-2", "Benzene", DisplayText="6", Matrix="SOIL", DepthIntervalText="0-2 ft"),
    _r("SB-2", "Naphthalene", DisplayText="2", Matrix="SOIL", DepthIntervalText="0-2 ft"),
    # SB-2, 2-4 ft
    _r("SB-2", "Benzene", DisplayText="4", Matrix="SOIL", DepthIntervalText="2-4 ft"),
    _r("SB-2", "Naphthalene", DisplayText="1", Matrix="SOIL", DepthIntervalText="2-4 ft"),
]

# All records combined
ALL_RECORDS = GW_RECORDS + SOIL_RECORDS


# ============================================================================
# Task 1 Tests
# ============================================================================

def test_build_current_event_shape():
    headers, rows, _ = _build_current_event(GW_RECORDS)
    assert headers == ["LocationID", "SampleDate", "Benzene", "MTBE", "Toluene"]
    assert len(rows) == 3           # one row per location
    for row in rows:
        assert row[1] == "2026-04-01"  # only latest date


def test_exceedance_coords_collected():
    _, _, exc_coords = _build_current_event(GW_RECORDS)
    assert len(exc_coords) >= 1
    # MW-1 is first sorted location (row_idx=0).
    # Benzene is first analyte; fixed cols are LocationID+SampleDate, so col_idx=2.
    assert (0, 2) in exc_coords


# ============================================================================
# Task 2 Tests
# ============================================================================

def test_build_gw_by_event_stacked():
    headers, rows, _ = _build_gw_by_event(GW_RECORDS)
    assert headers == ["Location", "Analyte", "2025-10-01", "2026-04-01"]
    assert len(rows) == 9                       # 3 locs × 3 analytes
    assert all(len(row) == 4 for row in rows)

    # "--" appears when a (loc, analyte, date) combo is absent
    partial = [
        r for r in GW_RECORDS
        if not (r.LocationID == "MW-1"
                and r.AnalyteCanonicalName == "Benzene"
                and r.SampleDate == dt.date(2025, 10, 1))
    ]
    _, rows2, _ = _build_gw_by_event(partial)
    mw1_benz = next(r for r in rows2 if r[0] == "MW-1" and r[1] == "Benzene")
    assert mw1_benz[2] == "--"    # 2025-10-01 column removed
    assert mw1_benz[3] == "5.5"  # 2026-04-01 column still present


# ============================================================================
# Task 3 Tests
# ============================================================================

def test_build_soil_by_depth_shape():
    headers, rows, _ = _build_soil_by_depth(SOIL_RECORDS)
    assert headers == ["Location", "Depth", "Benzene", "Naphthalene"]
    assert len(rows) == 4          # 2 locs × 2 depths
    assert all(len(row) == 4 for row in rows)
    # Row keys are (location, depth) pairs
    loc_depth_pairs = [(r[0], r[1]) for r in rows]
    assert ("SB-1", "0-2 ft") in loc_depth_pairs
    assert ("SB-2", "2-4 ft") in loc_depth_pairs


# ============================================================================
# Task 4 Tests (Integration)
# ============================================================================

def test_export_writes_file(tmp_path):
    out = tmp_path / "summary.xlsx"
    export_summary_tables(ALL_RECORDS, out)
    assert out.exists()
    wb = openpyxl.load_workbook(out)
    assert set(wb.sheetnames) == {"Current Event", "GW by Event", "Soil by Depth"}


def test_export_sheet_row_count(tmp_path):
    out = tmp_path / "summary.xlsx"
    export_summary_tables(ALL_RECORDS, out)
    wb = openpyxl.load_workbook(out)
    assert wb["Current Event"].max_row == 1 + 3   # header + 3 GW locations
    assert wb["GW by Event"].max_row == 1 + 9     # header + 3 locs × 3 analytes
    assert wb["Soil by Depth"].max_row == 1 + 4   # header + 2 locs × 2 depths


def test_exceedance_cell_fill(tmp_path):
    out = tmp_path / "summary.xlsx"
    export_summary_tables(ALL_RECORDS, out)
    wb = openpyxl.load_workbook(out)
    ce = wb["Current Event"]
    # MW-1 → data row 2 (row_idx 0  →  ws row = 0 + 2 = 2)
    # Benzene → ws col 3  (col_idx 2 → ws col = 2 + 1 = 3)
    cell = ce.cell(row=2, column=3)
    assert cell.fill.fgColor.rgb.endswith("FFCCCC")


def test_no_soil_records_qa_warning(tmp_path):
    qa = QACollector()
    out = tmp_path / "no_soil.xlsx"
    export_summary_tables(GW_RECORDS, out, include_soil_by_depth=True, qa=qa)
    assert "soil_records_absent" in [r.category for r in qa.records]


def test_empty_records_qa_warning(tmp_path):
    qa = QACollector()
    out = tmp_path / "empty.xlsx"
    export_summary_tables([], out, qa=qa)
    assert "no_records_for_sheet" in [r.category for r in qa.records]
