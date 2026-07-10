"""Merge-gate regression tests (ADR-0075).

One focused test per analyte-pivoting consumer converted to read through the
canonical-read policy (drop QC-flagged rows; resolve a Total/Dissolved pair on
the same grain to the preferred fraction). Each test feeds input where a QC row
and/or a Total+Dissolved pair on the SAME grain would corrupt the output, and
pins the corrected result — i.e. it would FAIL if the canonicalization call were
removed from the consumer.

arcpy-free: builds AnalyticalResultRecord via the shared _full_record factory.
"""
from __future__ import annotations

import datetime as dt

from openpyxl import load_workbook

from autogis.core.common.qa import QACollector
from autogis.core.common.records_csv import write_records_csv
from autogis.core.envmon.gdb_schema import AnalyticalResultRecord
from tests.envmon.test_gdb_schema_keys import _full_record

from autogis.core.envmon import (
    compare_events,
    history_report,
    export_summary_tables,
    export_geojson,
    schedule_vs_actual,
    data_gaps,
    export_summary,
    draft_plume_boundary,
    event_changelog,
    generate_event_report,
)

_D = dt.date(2026, 6, 26)


def test_compare_events_uses_total_not_qc_or_dissolved():
    # One grain, three rows. Order deliberately puts the WRONG rows first so
    # that without canonicalization compare_events would read the Dissolved
    # value as "current"; with it, only the Total survives.
    results = [
        _full_record(ResultFraction="Dissolved", ResultNumeric=1.0,
                     ResultRawText="1.0", IsDetected=1),
        _full_record(QCType="FIELD_BLANK", ResultFraction="Total",
                     ResultNumeric=9.0, ResultRawText="9.0", IsDetected=1),
        _full_record(ResultFraction="Total", ResultNumeric=5.0,
                     ResultRawText="5.0", IsDetected=1),
    ]
    qa = QACollector()
    out = compare_events.compare_events(results, qa, current_event_date=_D)
    assert len(out) == 1
    assert out[0].CurrentResultNumeric == 5.0   # Total, not Dissolved(1)/QC(9)


def test_history_report_counts_only_total_row():
    results = [
        _full_record(ResultFraction="Dissolved", ResultNumeric=1.0),
        _full_record(QCType="FIELD_BLANK", ResultFraction="Total",
                     ResultNumeric=99.0),
        _full_record(ResultFraction="Total", ResultNumeric=5.0),
    ]
    qa = QACollector()
    rows = history_report.build_history_report(results, qa=qa)
    assert len(rows) == 1
    assert rows[0].NTotal == 1                   # not 3
    assert rows[0].MinResult == 5.0
    assert rows[0].MaxResult == 5.0
    assert rows[0].MeanResult == 5.0             # Total only, not (1+5+99)/3


def test_export_summary_tables_cell_shows_total_not_qc(tmp_path):
    # Current-Event pivot writes one cell per (location, analyte); a raw dict
    # overwrite means the LAST row iterated wins. Order QC last so that without
    # canonicalization the cell would show the QC blank's DisplayText.
    results = [
        _full_record(ResultFraction="Total", DisplayText="5.0 [Total]"),
        _full_record(ResultFraction="Dissolved", DisplayText="1.0 [Dissolved]"),
        _full_record(QCType="FIELD_BLANK", DisplayText="QC-BLANK"),
    ]
    out = tmp_path / "summary.xlsx"
    export_summary_tables.export_summary_tables(results, out, qa=QACollector())

    ws = load_workbook(out)["Current Event"]
    headers = [c.value for c in ws[1]]
    col = headers.index("Benzene")
    cell = next(r[col] for r in ws.iter_rows(min_row=2, values_only=True)
                if r[0] == "MW-1")
    assert cell == "5.0 [Total]"                 # QC/Dissolved did not overwrite


def test_export_geojson_property_uses_total_not_qc():
    results = [
        _full_record(ResultFraction="Total", DisplayText="5.0 [Total]"),
        _full_record(ResultFraction="Dissolved", DisplayText="1.0 [Dissolved]"),
        _full_record(QCType="FIELD_BLANK", DisplayText="QC-BLANK"),
    ]
    qa = QACollector()
    fc = export_geojson.build_geojson(results, {"MW-1": (10.0, 20.0)}, qa=qa)
    props = fc["features"][0]["properties"]
    assert props["Benzene_value"] == "5.0 [Total]"   # not QC-BLANK / Dissolved


def test_schedule_vs_actual_qc_only_analyte_is_missing():
    # The analyte exists ONLY as a QC-flagged row; the canonical read drops it,
    # so the schedule must report it MISSING (a QC row is not a real sample).
    results = [_full_record(AnalyteCanonicalName="Benzene", QCType="FIELD_BLANK")]
    schedule = {"site_id": "S1", "wells": ["MW-1"],
                "required_analytes": ["Benzene"]}
    qa = QACollector()
    rows = schedule_vs_actual.compare_schedule_vs_actual(
        results, schedule, qa=qa, event_date=_D)
    rec = next(r for r in rows
               if r.LocationID == "MW-1" and r.AnalyteName == "Benzene")
    assert rec.Status == "MISSING"               # not SAMPLED


def test_data_gaps_qc_only_analyte_is_missed():
    # MW-1 has a real Toluene result but Benzene only as a QC row. After the
    # canonical read Benzene is absent -> a MISSED_ANALYTE gap must be emitted.
    results = [
        _full_record(AnalyteCanonicalName="Toluene"),
        _full_record(AnalyteCanonicalName="Benzene", QCType="FIELD_BLANK"),
    ]
    schedule = {"site_id": "S1", "wells": ["MW-1"],
                "required_analytes": ["Benzene", "Toluene"]}
    qa = QACollector()
    gaps = data_gaps.identify_data_gaps(
        results, schedule, event_date=_D, window_days=30, dry_wells={}, qa=qa)
    assert any(g.GapType == "MISSED_ANALYTE"
               and g.AnalyteCanonicalName == "Benzene" for g in gaps)


def test_export_summary_total_count_is_one(tmp_path):
    results = [
        _full_record(ResultFraction="Total", ResultNumeric=5.0, IsDetected=1),
        _full_record(ResultFraction="Dissolved", ResultNumeric=1.0, IsDetected=1),
        _full_record(QCType="FIELD_BLANK", ResultNumeric=9.0, IsDetected=1),
    ]
    out = tmp_path / "analytical_summary.xlsx"
    export_summary.export_analytical_summary(
        [], results, out, site_id="S1", qa=QACollector())

    ws = load_workbook(out)["Summary by Analyte"]
    total = next(r[1] for r in ws.iter_rows(min_row=2, values_only=True)
                 if r[0] == "Benzene")
    assert total == 1                            # not 3


def test_generate_event_report_summary_counts_are_canonical(tmp_path):
    # Executive-summary counts must not tally QC rows or both fractions. One
    # grain: Total (not exceeding) + Dissolved + a QC row stamped as exceeding.
    # Canonical => 1 result, 0 exceedances; without it => 3 results, 1 exceedance.
    results_path = tmp_path / "results.csv"
    write_records_csv(
        [_full_record(ResultFraction="Total", ExceedsScreeningLevel=0),
         _full_record(ResultFraction="Dissolved", ExceedsScreeningLevel=0),
         _full_record(QCType="FIELD_BLANK", ExceedsScreeningLevel=1)],
        results_path, record_class=AnalyticalResultRecord)

    md = generate_event_report.generate_event_report(
        "S1", "2026Q2", results_csv=results_path, qa=QACollector())
    assert "| Total analytical results | 1 |" in md          # not 3
    assert "| Screening level exceedances | 0 |" in md        # QC exceedance dropped


def test_event_changelog_qc_row_does_not_flip_classification():
    # current has the real unchanged row + a QC row (value 99, exceeds) on the
    # same grain. Without canonicalization the last-wins map would pick the QC
    # row and misclassify the pair as an exceedance/value change; with it the
    # QC row is dropped and the pair is NO_CHANGE.
    import dataclasses as _dc
    prior = [_dc.asdict(_full_record(ResultNumeric=5.0, ExceedsScreeningLevel=0))]
    current = [
        _dc.asdict(_full_record(ResultNumeric=5.0, ExceedsScreeningLevel=0)),
        _dc.asdict(_full_record(QCType="FIELD_BLANK", ResultNumeric=99.0,
                                ExceedsScreeningLevel=1)),
    ]
    res = event_changelog.generate_event_changelog(prior, current)
    rec = next(c for c in res.changes
               if c.location_id == "MW-1" and c.analyte_name == "Benzene")
    assert rec.change_type == "NO_CHANGE"        # QC row dropped, 5.0 unchanged


def test_draft_plume_boundary_qc_only_exceedance_drops_well(tmp_path):
    # A well whose ONLY exceedance sits on a QC row must not seed a plume point.
    results_path = tmp_path / "results.csv"
    coords_path = tmp_path / "coords.csv"
    write_records_csv(
        [_full_record(LocationID="MW-1", ExceedsScreeningLevel=1,
                      QCType="FIELD_BLANK")],
        results_path, record_class=AnalyticalResultRecord)
    coords_path.write_text("location_id,x,y\nMW-1,10.0,20.0\n", encoding="utf-8")

    qa = QACollector()
    pts = draft_plume_boundary.filter_results_to_exceedance_points(
        results_path, coords_path, qa=qa)
    assert pts == []                             # QC exceedance dropped
