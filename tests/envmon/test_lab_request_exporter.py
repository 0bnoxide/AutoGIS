from pathlib import Path
import csv
import pytest
import openpyxl
from autogis.core.envmon.lab_request_exporter import (
    LabRequestRow, LabRequestResult,
    build_lab_request_rows, write_lab_request_workbook, write_lab_request_csv,
)

_PLAN = [
    {"SampleID": "H281-MW01-20260615-GW", "LocationID": "MW-01",
     "AnalyteGroup": "GW_VOC", "Matrix": "GW",
     "Container": "40mL VOA", "Preservative": "HCl",
     "HoldTimeDays": "14", "Notes": ""},
    {"SampleID": "H281-MW02-20260615-GW", "LocationID": "MW-02",
     "AnalyteGroup": "GW_METALS", "Matrix": "GW",
     "Container": "250mL HDPE", "Preservative": "HNO3",
     "HoldTimeDays": "180", "Notes": ""},
]
_GROUPS = {
    "GW_VOC": {
        "analytes": ["Benzene", "Toluene", "Ethylbenzene", "Xylenes"],
        "container": "40mL VOA", "preservative": "HCl", "hold_time_days": 14,
    },
    "GW_METALS": {
        "analytes": ["Arsenic", "Cadmium", "Lead"],
        "container": "250mL HDPE", "preservative": "HNO3", "hold_time_days": 180,
    },
}


def test_build_rows_count():
    rows = build_lab_request_rows(_PLAN, _GROUPS)
    assert len(rows) == 2


def test_analyte_list_populated():
    rows = build_lab_request_rows(_PLAN, _GROUPS)
    voc_row = next(r for r in rows if r.analyte_group == "GW_VOC")
    assert "Benzene" in voc_row.analyte_list
    assert "Toluene" in voc_row.analyte_list


def test_hold_time_from_plan():
    rows = build_lab_request_rows(_PLAN, _GROUPS)
    voc_row = next(r for r in rows if r.analyte_group == "GW_VOC")
    assert voc_row.hold_time_days == 14


def test_turnaround_days_default():
    rows = build_lab_request_rows(_PLAN, _GROUPS)
    assert all(r.turnaround_days == 5 for r in rows)


def test_write_workbook_produces_xlsx(tmp_path):
    rows = build_lab_request_rows(_PLAN, _GROUPS)
    out = tmp_path / "lab_request.xlsx"
    result = write_lab_request_workbook(rows, out, site_id="H281",
                                         event_date="2026-06-15")
    assert out.exists()


def test_sheet1_has_sampleid_column(tmp_path):
    rows = build_lab_request_rows(_PLAN, _GROUPS)
    out = tmp_path / "lab_request.xlsx"
    write_lab_request_workbook(rows, out)
    wb = openpyxl.load_workbook(str(out))
    ws = wb.worksheets[0]
    headers = [cell.value for cell in ws[1]]
    assert "SampleID" in headers


def test_sheet2_analyte_list(tmp_path):
    rows = build_lab_request_rows(_PLAN, _GROUPS)
    out = tmp_path / "lab_request.xlsx"
    write_lab_request_workbook(rows, out)
    wb = openpyxl.load_workbook(str(out))
    assert len(wb.sheetnames) == 2


def test_write_csv(tmp_path):
    rows = build_lab_request_rows(_PLAN, _GROUPS)
    out = tmp_path / "lab_request.csv"
    write_lab_request_csv(rows, out)
    assert out.exists()
    with out.open() as fh:
        rows_read = list(csv.DictReader(fh))
    assert len(rows_read) == 2
    assert "SampleID" in rows_read[0]


def test_column_map_renames(tmp_path):
    rows = build_lab_request_rows(_PLAN, _GROUPS)
    out = tmp_path / "lab_request.xlsx"
    write_lab_request_workbook(rows, out, column_map={"SampleID": "Lab Sample ID"})
    wb = openpyxl.load_workbook(str(out))
    ws = wb.worksheets[0]
    headers = [cell.value for cell in ws[1]]
    assert "Lab Sample ID" in headers


def test_export_lab_request_in_help():
    from click.testing import CliRunner
    from autogis.adapters.cli import autogis
    result = CliRunner().invoke(autogis, ["envmon", "--help"])
    assert "export-lab-request" in result.output
