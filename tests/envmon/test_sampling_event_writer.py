"""Tests for sampling_event_writer — arcpy-free, openpyxl only."""
import pytest

from autogis.core.envmon.create_sampling_event import (
    build_sampling_event_plan,
)
from autogis.core.envmon.sampling_event_writer import (
    write_sampling_event_workbook,
)

SITE = {"site_id": "H281", "site_name": "H281 Glasgow"}
EVENT = {
    "event_name": "2026-Q2",
    "event_date": "2026-07-15",
    "coc_prefix": "H281-COC",
    "lab_name": "TestAmerica Seattle",
    "matrices": ["GW"],
    "location_ids": ["MW-1", "MW-2"],
    "crew_list": ["Alice Smith"],
    "dup_frequency": 0,
    "analyte_groups": {
        "VOCs": ["Benzene", "Toluene"],
        "Metals": ["Arsenic"],
    },
    "group_sampling": {
        "VOCs": {"container": "40mL VOA", "preservative": "HCl",
                 "hold_time_hr": 14, "bottles": 1},
        "Metals": {"container": "250mL PP", "preservative": "HNO3",
                   "hold_time_hr": 180, "bottles": 1},
    },
}
ADICT = {
    "Benzene": {"abbreviation": "B", "display_order": 10,
                "default_units_by_matrix": {"GW": "ug/L"}},
    "Toluene": {"abbreviation": "T", "display_order": 20,
                "default_units_by_matrix": {"GW": "ug/L"}},
    "Arsenic": {"abbreviation": "As", "display_order": 30,
                "default_units_by_matrix": {"GW": "ug/L"}},
}


@pytest.fixture
def plan():
    return build_sampling_event_plan(SITE, EVENT, ADICT, run_id="writer-test-001")


@pytest.fixture
def wb_path(plan, tmp_path):
    out = tmp_path / "H281_2026-Q2_sampling_plan.xlsx"
    write_sampling_event_workbook(plan, out)
    return out


def test_file_is_created(wb_path):
    assert wb_path.exists()
    assert wb_path.stat().st_size > 0


def test_workbook_has_three_sheets(wb_path):
    import openpyxl
    wb = openpyxl.load_workbook(wb_path)
    assert set(wb.sheetnames) == {"Expected_Samples", "Crew_Assignment", "COC_Draft"}


def test_expected_samples_header(wb_path):
    import openpyxl
    ws = openpyxl.load_workbook(wb_path)["Expected_Samples"]
    headers = [ws.cell(1, c).value for c in range(1, 13)]
    assert "SampleID" in headers
    assert "AnalyteGroup" in headers
    assert "COCNumber" in headers


def test_expected_samples_row_count(wb_path):
    """2 wells × 2 groups = 4 primary rows + 1 header = 5 rows total."""
    import openpyxl
    ws = openpyxl.load_workbook(wb_path)["Expected_Samples"]
    # max_row includes header
    assert ws.max_row == 5


def test_crew_assignment_header(wb_path):
    import openpyxl
    ws = openpyxl.load_workbook(wb_path)["Crew_Assignment"]
    headers = [ws.cell(1, c).value for c in range(1, 4)]
    assert "LocationID" in headers
    assert "AssignedTo" in headers
    assert "BottleCount" in headers


def test_crew_assignment_row_count(wb_path):
    """2 wells = 2 data rows + 1 header = 3 rows."""
    import openpyxl
    ws = openpyxl.load_workbook(wb_path)["Crew_Assignment"]
    assert ws.max_row == 3


def test_coc_draft_header(wb_path):
    import openpyxl
    ws = openpyxl.load_workbook(wb_path)["COC_Draft"]
    headers = [ws.cell(1, c).value for c in range(1, 16)]
    assert "COCNumber" in headers
    assert "SampleID" in headers
    assert "SamplerSignature" in headers
    assert "DateTimeSampled" in headers


def test_coc_draft_row_count(wb_path):
    """Same row count as Expected_Samples (1-to-1 map)."""
    import openpyxl
    wb = openpyxl.load_workbook(wb_path)
    es_rows = wb["Expected_Samples"].max_row
    coc_rows = wb["COC_Draft"].max_row
    assert coc_rows == es_rows


def test_returns_written_path(plan, tmp_path):
    out = tmp_path / "plan.xlsx"
    returned = write_sampling_event_workbook(plan, out)
    assert returned == out


def test_sample_id_in_expected_samples_sheet(wb_path):
    import openpyxl
    ws = openpyxl.load_workbook(wb_path)["Expected_Samples"]
    col_a = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
    assert "MW-1-20260715-GW" in col_a
