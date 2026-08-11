from pathlib import Path
import pytest
import openpyxl
from autogis.core.envmon.qc_sample_summary import (
    QC_TYPES, QCRecord, classify_qc_rows, compute_rpd,
    write_qc_summary_workbook, QCSummaryResult,
)

_ROWS = [
    {"SampleID": "S1-GW-VOC",    "SampleType": "primary",   "LocationID": "MW-01",
     "AnalyteName": "Benzene", "ResultValue": "5.0",  "ResultQualifier": "",
     "ReportedUnits": "ug/L", "SampleDate": "2026-06-15"},
    {"SampleID": "S1-MB",        "SampleType": "method_blank", "LocationID": "",
     "AnalyteName": "Benzene", "ResultValue": "ND",   "ResultQualifier": "ND",
     "ReportedUnits": "ug/L", "SampleDate": "2026-06-15"},
    {"SampleID": "S1-FB",        "SampleType": "field_blank", "LocationID": "",
     "AnalyteName": "Benzene", "ResultValue": "0.5",  "ResultQualifier": "",
     "ReportedUnits": "ug/L", "SampleDate": "2026-06-15"},
    {"SampleID": "S1-LD-A",      "SampleType": "lab_duplicate", "LocationID": "MW-01",
     "AnalyteName": "Benzene", "ResultValue": "5.0",  "ResultQualifier": "",
     "ReportedUnits": "ug/L", "SampleDate": "2026-06-15"},
    {"SampleID": "S1-LD-B",      "SampleType": "lab_duplicate", "LocationID": "MW-01",
     "AnalyteName": "Benzene", "ResultValue": "4.5",  "ResultQualifier": "",
     "ReportedUnits": "ug/L", "SampleDate": "2026-06-15"},
]


def test_classify_primary_excluded():
    records = classify_qc_rows(_ROWS)
    assert not any(r.qc_type == "primary" for r in records)


def test_classify_method_blank():
    records = classify_qc_rows(_ROWS)
    mb = [r for r in records if r.qc_type == "method_blank"]
    assert len(mb) == 1


def test_classify_field_blank():
    records = classify_qc_rows(_ROWS)
    fb = [r for r in records if r.qc_type == "field_blank"]
    assert len(fb) == 1


def test_blank_detection_warning():
    result = _make_result(_ROWS)
    # S1-FB has 0.5 — a detection in a blank
    assert result.blank_detections >= 1


def _make_result(rows):
    records = classify_qc_rows(rows)
    from autogis.core.envmon.qc_sample_summary import QCSummaryResult
    from autogis.core.common.qa import QACollector
    blank_count = sum(
        1 for r in records
        if r.qc_type in ("method_blank", "field_blank", "trip_blank")
        and r.result_value is not None
    )
    return QCSummaryResult(
        records=records, blank_detections=blank_count,
        spike_failures=0, duplicate_failures=0, qa=QACollector(),
    )


def test_compute_rpd():
    assert compute_rpd(5.0, 4.5) == pytest.approx(10.526, rel=1e-3)


def test_compute_rpd_zero_denominator():
    """#436 -- a 0/0 pair is not calculable, not 'perfect agreement'.

    Must agree with the sibling implementations, which return None; reporting
    0.0 here would show a 0/0 duplicate pair as a passing 0.0% RPD.
    """
    from autogis.core.envmon.evaluate_rpd_qa import _rpd as _rpd_eval
    from autogis.core.envmon.normalize_rpd import _rpd as _rpd_norm

    assert compute_rpd(0.0, 0.0) is None
    assert _rpd_eval(0.0, 0.0) is None and _rpd_norm(0.0, 0.0) is None


def test_suffix_inference_mb():
    rows = [{"SampleID": "ABC-MB", "SampleType": "", "LocationID": "",
             "AnalyteName": "Benzene", "ResultValue": "ND", "ResultQualifier": "ND",
             "ReportedUnits": "ug/L", "SampleDate": "2026-06-15"}]
    records = classify_qc_rows(rows)
    assert records[0].qc_type == "method_blank"


def test_write_workbook_produces_xlsx(tmp_path):
    from autogis.core.envmon.qc_sample_summary import QCSummaryResult
    from autogis.core.common.qa import QACollector
    records = classify_qc_rows(_ROWS)
    result = QCSummaryResult(records=records, blank_detections=1,
                              spike_failures=0, duplicate_failures=0,
                              qa=QACollector())
    out = tmp_path / "qc_summary.xlsx"
    write_qc_summary_workbook(result, out)
    assert out.exists()
    wb = openpyxl.load_workbook(str(out))
    assert len(wb.sheetnames) > 1  # at least one QC sheet + Summary


def test_summary_sheet_present(tmp_path):
    from autogis.core.envmon.qc_sample_summary import QCSummaryResult
    from autogis.core.common.qa import QACollector
    records = classify_qc_rows(_ROWS)
    result = QCSummaryResult(records=records, blank_detections=0,
                              spike_failures=0, duplicate_failures=0,
                              qa=QACollector())
    out = tmp_path / "qc_summary.xlsx"
    write_qc_summary_workbook(result, out)
    wb = openpyxl.load_workbook(str(out))
    assert "Summary" in wb.sheetnames


def test_infer_qc_type_field_duplicate_suffix():
    from autogis.core.envmon.qc_sample_summary import _infer_qc_type
    assert _infer_qc_type("MW-1-20260715-GW-FD", "") == "field_duplicate"
