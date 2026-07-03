"""CLI smoke tests for the §2/§7 intake + planning batch (5 headless tools).

Each test:
1. Builds minimal valid inputs under tmp_path.
2. Invokes the CLI via CliRunner (catch_exceptions=False).
3. Asserts exit_code == 0 and output file(s) exist with expected content.

No arcpy, no arcgis. Only touches tests/envmon/.
"""
import csv
import json
from pathlib import Path

import openpyxl
import pytest
import yaml
from click.testing import CliRunner

from autogis.adapters.cli import autogis


def _run(*args):
    return CliRunner().invoke(autogis, list(args), catch_exceptions=False)


def _write_csv(path: Path, rows: list) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        w.writeheader()
        for r in rows:
            w.writerow(r)


# ---------------------------------------------------------------------------
# 1. draft-parser-profile (Tool 2.1)
# ---------------------------------------------------------------------------

class TestDraftParserProfile:
    def test_creates_draft_yaml(self, tmp_path):
        """A single-sheet workbook produces a YAML draft profile."""
        wb_path = tmp_path / "lab.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(["LocationID", "SampleDate", "Benzene", "Toluene"])
        ws.append(["MW-01", "2024-06-01", "12.3", "ND"])
        wb.save(str(wb_path))

        out = tmp_path / "draft.yaml"
        r = _run("envmon", "draft-parser-profile", str(wb_path),
                 "--output", str(out), "--profile-id", "TEST_DRAFT")
        assert r.exit_code == 0, r.output
        assert out.exists()
        profile = yaml.safe_load(out.read_text(encoding="utf-8"))
        assert profile["profile_id"] == "TEST_DRAFT"
        assert "sheets" in profile
        assert len(profile["sheets"]) >= 1
        assert "_TODO" in profile["sheets"][0]

    def test_multi_sheet_workbook(self, tmp_path):
        wb_path = tmp_path / "multi.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        wb.active.append(["ID", "Date", "Analyte"])
        wb.create_sheet("Sheet2")
        wb["Sheet2"].append(["ID", "Date", "Analyte"])
        wb.save(str(wb_path))

        out = tmp_path / "multi_draft.yaml"
        r = _run("envmon", "draft-parser-profile", str(wb_path),
                 "--output", str(out))
        assert r.exit_code == 0, r.output
        profile = yaml.safe_load(out.read_text(encoding="utf-8"))
        assert len(profile["sheets"]) == 2


# ---------------------------------------------------------------------------
# 2. batch-import-workbooks (Tool 2.2)
# ---------------------------------------------------------------------------

class TestBatchImportWorkbooks:
    def _make_edd_profile(self, path: Path) -> None:
        import textwrap
        profile_yaml = textwrap.dedent("""
            profile_id: TEST_PROFILE
            lab_name: Test Lab
            format: flat_csv
            date_format: "%Y-%m-%d"
            encoding: utf-8
            columns:
              sample_id:       SampleID
              location_id:     LocationID
              event_date:      SampleDate
              matrix:          Matrix
              analyte:         Analyte
              result:          Result
              units:           Units
              qualifier:       Qualifier
              reporting_limit: RL
            matrix_map:
              GW: GW
            nondetect_qualifiers:
              - U
        """).strip()
        path.write_text(profile_yaml, encoding="utf-8")

    def _make_edd_csv(self, path: Path) -> None:
        rows = [
            {"SampleID": "S-001", "LocationID": "MW-01", "SampleDate": "2024-06-01",
             "Matrix": "GW", "Analyte": "Benzene", "Result": "12.3",
             "Units": "ug/L", "Qualifier": "", "RL": "1.0"},
        ]
        _write_csv(path, rows)

    def test_batch_processes_one_workbook(self, tmp_path):
        prof_path = tmp_path / "profile.yaml"
        wb_path = tmp_path / "lab_edd.csv"
        self._make_edd_profile(prof_path)
        self._make_edd_csv(wb_path)

        manifest_path = tmp_path / "manifest.csv"
        _write_csv(manifest_path, [{"workbook_path": str(wb_path),
                                    "profile_path": str(prof_path),
                                    "site_id": "SITE-A"}])
        out_dir = tmp_path / "out"

        r = _run("envmon", "batch-import-workbooks",
                 "--manifest", str(manifest_path),
                 "--output-dir", str(out_dir))
        assert r.exit_code == 0, r.output
        manifest_out = out_dir / "batch_manifest.csv"
        assert manifest_out.exists()
        rows = list(csv.DictReader(manifest_out.open(encoding="utf-8")))
        assert len(rows) == 1
        assert rows[0]["Status"] == "OK"

    def test_missing_workbook_marked_skip(self, tmp_path):
        prof_path = tmp_path / "profile.yaml"
        self._make_edd_profile(prof_path)
        manifest_path = tmp_path / "manifest.csv"
        _write_csv(manifest_path, [{"workbook_path": "/no/such/file.csv",
                                    "profile_path": str(prof_path),
                                    "site_id": "SITE-A"}])
        out_dir = tmp_path / "out"
        r = _run("envmon", "batch-import-workbooks",
                 "--manifest", str(manifest_path),
                 "--output-dir", str(out_dir))
        assert r.exit_code == 0, r.output
        rows = list(csv.DictReader((out_dir / "batch_manifest.csv").open(encoding="utf-8")))
        assert rows[0]["Status"] == "SKIP"

    # --- --edd-dir alternate input mode (BatchEDDImport fold, ADR-0048) ---

    def test_edd_dir_mode_imports_directory(self, tmp_path):
        """Directory of EDD CSVs + one shared profile/site imports every file."""
        prof_path = tmp_path / "profile.yaml"
        self._make_edd_profile(prof_path)
        edd_dir = tmp_path / "edds"
        edd_dir.mkdir()
        self._make_edd_csv(edd_dir / "a.csv")
        self._make_edd_csv(edd_dir / "b.csv")
        out_dir = tmp_path / "out"

        r = _run("envmon", "batch-import-workbooks",
                 "--edd-dir", str(edd_dir),
                 "--profile", str(prof_path),
                 "--site", "SITE-A",
                 "--output-dir", str(out_dir))
        assert r.exit_code == 0, r.output
        rows = list(csv.DictReader((out_dir / "batch_manifest.csv").open(encoding="utf-8")))
        assert len(rows) == 2
        assert all(row["Status"] == "OK" for row in rows)
        assert all(row["SiteID"] == "SITE-A" for row in rows)
        samples = list(csv.DictReader((out_dir / "sample_records.csv").open(encoding="utf-8")))
        assert len(samples) == 2

    def test_edd_dir_per_file_failure_does_not_abort(self, tmp_path):
        """One unreadable file is marked ERROR; the rest still import."""
        prof_path = tmp_path / "profile.yaml"
        self._make_edd_profile(prof_path)
        edd_dir = tmp_path / "edds"
        edd_dir.mkdir()
        self._make_edd_csv(edd_dir / "good.csv")
        (edd_dir / "bad.csv").write_bytes(b"\xff\xfe\x00\x00not utf-8\xff")
        out_dir = tmp_path / "out"

        r = _run("envmon", "batch-import-workbooks",
                 "--edd-dir", str(edd_dir),
                 "--profile", str(prof_path),
                 "--site", "SITE-A",
                 "--output-dir", str(out_dir))
        assert r.exit_code == 1  # SEV_ERROR recorded -> fail-on=error default
        rows = list(csv.DictReader((out_dir / "batch_manifest.csv").open(encoding="utf-8")))
        statuses = {Path(row["WorkbookPath"]).name: row["Status"] for row in rows}
        assert statuses == {"good.csv": "OK", "bad.csv": "ERROR"}
        samples = list(csv.DictReader((out_dir / "sample_records.csv").open(encoding="utf-8")))
        assert len(samples) == 1  # good.csv still imported

    def test_manifest_and_edd_dir_mutually_exclusive(self, tmp_path):
        prof_path = tmp_path / "profile.yaml"
        self._make_edd_profile(prof_path)
        wb_path = tmp_path / "lab_edd.csv"
        self._make_edd_csv(wb_path)
        manifest_path = tmp_path / "manifest.csv"
        _write_csv(manifest_path, [{"workbook_path": str(wb_path),
                                    "profile_path": str(prof_path),
                                    "site_id": "SITE-A"}])
        r = _run("envmon", "batch-import-workbooks",
                 "--manifest", str(manifest_path),
                 "--edd-dir", str(tmp_path),
                 "--profile", str(prof_path),
                 "--site", "SITE-A",
                 "--output-dir", str(tmp_path / "out"))
        assert r.exit_code == 2  # UsageError: both input modes supplied

    def test_input_mode_required(self, tmp_path):
        r = _run("envmon", "batch-import-workbooks",
                 "--output-dir", str(tmp_path / "out"))
        assert r.exit_code == 2  # UsageError: neither --manifest nor --edd-dir

    def test_edd_dir_no_matching_files_errors(self, tmp_path):
        prof_path = tmp_path / "profile.yaml"
        self._make_edd_profile(prof_path)
        edd_dir = tmp_path / "empty"
        edd_dir.mkdir()
        r = _run("envmon", "batch-import-workbooks",
                 "--edd-dir", str(edd_dir),
                 "--profile", str(prof_path),
                 "--site", "SITE-A",
                 "--output-dir", str(tmp_path / "out"))
        assert r.exit_code == 1  # ClickException: nothing matched

    def test_manifest_rows_from_dir_xlsx_fallback(self, tmp_path):
        """Default pattern *.csv falls back to *.xlsx; explicit pattern doesn't."""
        from autogis.core.envmon.batch_workbook_importer import manifest_rows_from_dir
        (tmp_path / "a.xlsx").write_bytes(b"")
        rows = manifest_rows_from_dir(tmp_path, Path("p.yaml"), "SITE-A")
        assert [Path(r["workbook_path"]).name for r in rows] == ["a.xlsx"]
        assert rows[0] == {"workbook_path": str(tmp_path / "a.xlsx"),
                           "profile_path": "p.yaml", "site_id": "SITE-A"}
        assert manifest_rows_from_dir(tmp_path, Path("p.yaml"), "SITE-A",
                                      pattern="*.csv") == []


# ---------------------------------------------------------------------------
# 3. migrate-legacy-data (Tool 2.4)
# ---------------------------------------------------------------------------

class TestMigrateLegacyData:
    def test_wide_to_long_basic(self, tmp_path):
        wide = tmp_path / "wide.csv"
        _write_csv(wide, [
            {"LocationID": "MW-01", "SampleDate": "2024-06-01",
             "Benzene": "12.3", "Toluene": "<5", "Xylenes": ""},
            {"LocationID": "MW-02", "SampleDate": "2024-06-01",
             "Benzene": "ND", "Toluene": "8.1", "Xylenes": ""},
        ])
        out = tmp_path / "long.csv"
        r = _run("envmon", "migrate-legacy-data",
                 "--input-csv", str(wide),
                 "--output", str(out),
                 "--site-id", "SITE-A")
        assert r.exit_code == 0, r.output
        assert out.exists()
        rows = list(csv.DictReader(out.open(encoding="utf-8")))
        # MW-01: Benzene + Toluene (Xylenes blank = skipped); MW-02: Toluene + Benzene
        assert len(rows) == 4
        analytes = {r["AnalyteCanonicalName"] for r in rows}
        assert "Benzene" in analytes and "Toluene" in analytes

    def test_nondetect_flagged(self, tmp_path):
        wide = tmp_path / "wide.csv"
        _write_csv(wide, [
            {"LocationID": "MW-01", "SampleDate": "2024-06-01",
             "Benzene": "<5"},
        ])
        out = tmp_path / "long.csv"
        r = _run("envmon", "migrate-legacy-data",
                 "--input-csv", str(wide), "--output", str(out),
                 "--nondetect-prefix", "<")
        assert r.exit_code == 0, r.output
        rows = list(csv.DictReader(out.open(encoding="utf-8")))
        assert rows[0]["IsNondetect"] == "True"

    def test_missing_location_skipped(self, tmp_path):
        wide = tmp_path / "wide.csv"
        _write_csv(wide, [
            {"LocationID": "", "SampleDate": "2024-06-01", "Benzene": "5.0"},
            {"LocationID": "MW-01", "SampleDate": "2024-06-01", "Benzene": "3.0"},
        ])
        out = tmp_path / "long.csv"
        r = _run("envmon", "migrate-legacy-data",
                 "--input-csv", str(wide), "--output", str(out))
        assert r.exit_code == 0, r.output
        rows = list(csv.DictReader(out.open(encoding="utf-8")))
        assert all(row["LocationID"] == "MW-01" for row in rows)


# ---------------------------------------------------------------------------
# 4. create-sampling-plan (Tool 7.2)
# ---------------------------------------------------------------------------

class TestCreateSamplingPlan:
    def _make_wells(self, path: Path, extra_col: bool = False) -> None:
        rows = [{"location_id": "MW-01", "matrix": "GW"},
                {"location_id": "MW-02", "matrix": "GW"}]
        if extra_col:
            for r in rows:
                r["analyte_groups"] = "voc,metals"
        _write_csv(path, rows)

    def _make_groups(self, path: Path) -> None:
        groups = {
            "voc": {"bottles": 2, "bottle_size_ml": 40,
                    "preservation": "HCl pH<2"},
            "metals": {"bottles": 1, "bottle_size_ml": 500,
                       "preservation": "HNO3 pH<2"},
        }
        path.write_text(yaml.dump(groups), encoding="utf-8")

    def test_plan_created_with_two_groups(self, tmp_path):
        wells = tmp_path / "wells.csv"
        groups = tmp_path / "groups.yaml"
        self._make_wells(wells)
        self._make_groups(groups)
        samples_out = tmp_path / "samples.csv"
        bottles_out = tmp_path / "bottles.csv"

        r = _run("envmon", "create-sampling-plan",
                 "--wells-csv", str(wells),
                 "--analyte-groups", str(groups),
                 "--event-date", "2024-06-01",
                 "--site-id", "SITE-A",
                 "--samples-output", str(samples_out),
                 "--bottles-output", str(bottles_out))
        assert r.exit_code == 0, r.output
        samples = list(csv.DictReader(samples_out.open(encoding="utf-8")))
        # 2 wells × 2 groups = 4 planned samples
        assert len(samples) == 4
        bottles = list(csv.DictReader(bottles_out.open(encoding="utf-8")))
        assert len(bottles) == 2
        voc_row = next(b for b in bottles if b["AnalyteGroup"] == "voc")
        assert int(voc_row["BottleCount"]) == 4  # 2 wells × 2 bottles each

    def test_empty_wells_produces_empty_plan(self, tmp_path):
        wells = tmp_path / "wells.csv"
        wells.write_text("location_id,matrix\n", encoding="utf-8")
        groups = tmp_path / "groups.yaml"
        self._make_groups(groups)
        r = _run("envmon", "create-sampling-plan",
                 "--wells-csv", str(wells),
                 "--analyte-groups", str(groups),
                 "--event-date", "2024-06-01",
                 "--samples-output", str(tmp_path / "s.csv"),
                 "--bottles-output", str(tmp_path / "b.csv"))
        assert r.exit_code == 0, r.output

    def test_per_well_analyte_group_restriction(self, tmp_path):
        """Wells with analyte_groups column only get those groups, not all."""
        wells = tmp_path / "wells.csv"
        # MW-01 gets only voc; MW-02 gets all groups (no restriction)
        _write_csv(wells, [
            {"location_id": "MW-01", "matrix": "GW", "analyte_groups": "voc"},
            {"location_id": "MW-02", "matrix": "GW", "analyte_groups": ""},
        ])
        groups = tmp_path / "groups.yaml"
        self._make_groups(groups)
        samples_out = tmp_path / "samples.csv"
        bottles_out = tmp_path / "bottles.csv"
        r = _run("envmon", "create-sampling-plan",
                 "--wells-csv", str(wells),
                 "--analyte-groups", str(groups),
                 "--event-date", "2024-06-01",
                 "--samples-output", str(samples_out),
                 "--bottles-output", str(bottles_out))
        assert r.exit_code == 0, r.output
        samples = list(csv.DictReader(samples_out.open(encoding="utf-8")))
        # MW-01: only voc (1); MW-02: voc + metals (2) → 3 total
        assert len(samples) == 3
        mw01 = [s for s in samples if s["LocationID"] == "MW-01"]
        assert all(s["AnalyteGroup"] == "voc" for s in mw01)


# ---------------------------------------------------------------------------
# 5. reconcile-field-lab (Tool 7.3)
# ---------------------------------------------------------------------------

class TestReconcileFieldLab:
    def _field_rows(self):
        return [
            {"sample_id": "S-001", "location_id": "MW-01",
             "collection_date": "2024-06-01", "matrix": "GW"},
            {"sample_id": "S-002", "location_id": "MW-02",
             "collection_date": "2024-06-01", "matrix": "GW"},
        ]

    def _lab_rows(self):
        return [
            {"sample_id": "S-001", "location_id": "MW-01",
             "analysis_date": "2024-06-01", "analyte_name": "Benzene",
             "matrix": "GW"},
            {"sample_id": "S-001", "location_id": "MW-01",
             "analysis_date": "2024-06-01", "analyte_name": "Toluene",
             "matrix": "GW"},
        ]

    def test_missing_lab_result_flagged(self, tmp_path):
        # FIELD_MISSING_FROM_LAB is an ERROR → exit 1 with fail-on=error (default)
        field = tmp_path / "field.csv"
        lab = tmp_path / "lab.csv"
        _write_csv(field, self._field_rows())
        _write_csv(lab, self._lab_rows())
        out = tmp_path / "flags.csv"
        r = CliRunner().invoke(autogis, [
            "envmon", "reconcile-field-lab",
            "--field-csv", str(field), "--lab-csv", str(lab),
            "--output", str(out),
        ], catch_exceptions=False)
        assert r.exit_code == 1  # errors present → non-zero
        assert out.exists()
        flags = list(csv.DictReader(out.open(encoding="utf-8")))
        types = {f["FlagType"] for f in flags}
        assert "FIELD_MISSING_FROM_LAB" in types  # S-002 has no lab result

    def test_extra_lab_sample_flagged(self, tmp_path):
        field = tmp_path / "field.csv"
        lab = tmp_path / "lab.csv"
        _write_csv(field, self._field_rows()[:1])  # only S-001
        lab_rows = self._lab_rows() + [
            {"sample_id": "S-EXTRA", "location_id": "MW-99",
             "analysis_date": "2024-06-01", "analyte_name": "Benzene",
             "matrix": "GW"},
        ]
        _write_csv(lab, lab_rows)
        out = tmp_path / "flags.csv"
        r = CliRunner().invoke(autogis, [
            "envmon", "reconcile-field-lab",
            "--field-csv", str(field), "--lab-csv", str(lab),
            "--output", str(out),
        ], catch_exceptions=False)
        assert r.exit_code == 1  # errors present
        assert out.exists()
        flags = list(csv.DictReader(out.open(encoding="utf-8")))
        types = {f["FlagType"] for f in flags}
        assert "LAB_MISSING_FROM_FIELD" in types

    def test_date_mismatch_flagged(self, tmp_path):
        # DATE_MISMATCH is a WARNING; fail-on=warning → exit 1
        field = tmp_path / "field.csv"
        lab = tmp_path / "lab.csv"
        _write_csv(field, [{"sample_id": "S-001", "location_id": "MW-01",
                             "collection_date": "2024-06-01", "matrix": "GW"}])
        _write_csv(lab, [{"sample_id": "S-001", "location_id": "MW-01",
                           "analysis_date": "2024-06-10",  # 9-day gap
                           "analyte_name": "Benzene", "matrix": "GW"}])
        out = tmp_path / "flags.csv"
        r = CliRunner().invoke(autogis, [
            "envmon", "reconcile-field-lab",
            "--field-csv", str(field), "--lab-csv", str(lab),
            "--output", str(out),
            "--date-tolerance", "1", "--fail-on", "warning",
        ], catch_exceptions=False)
        assert r.exit_code == 1  # warning present and fail-on=warning
        assert out.exists()
        flags = list(csv.DictReader(out.open(encoding="utf-8")))
        assert any(f["FlagType"] == "DATE_MISMATCH" for f in flags)

    def test_clean_data_no_flags(self, tmp_path):
        field = tmp_path / "field.csv"
        lab = tmp_path / "lab.csv"
        _write_csv(field, [{"sample_id": "S-001", "location_id": "MW-01",
                             "collection_date": "2024-06-01", "matrix": "GW"}])
        _write_csv(lab, [{"sample_id": "S-001", "location_id": "MW-01",
                           "analysis_date": "2024-06-01",
                           "analyte_name": "Benzene", "matrix": "GW"}])
        out = tmp_path / "flags.csv"
        r = _run("envmon", "reconcile-field-lab",
                 "--field-csv", str(field),
                 "--lab-csv", str(lab),
                 "--output", str(out))
        assert r.exit_code == 0, r.output
        flags = list(csv.DictReader(out.open(encoding="utf-8")))
        assert len(flags) == 0
