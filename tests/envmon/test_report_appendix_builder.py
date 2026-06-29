"""Tests for report_appendix_builder (Tool 9.2) — arcpy-free.

Checklist mirrors the Approved design spec
docs/superpowers/specs/2026-06-28-build-monitoring-report-appendix-design.md
"""
from pathlib import Path

from openpyxl import load_workbook

from autogis.core.envmon.report_appendix_builder import (
    AppendixSheetSpec,
    build_appendix_sheet_specs,
    write_appendix_workbook,
    YELLOW_FILL,
    RED_FILL,
)


# Two analyte groups (VOC, Metals), 2 wells, 2 events.
_ROWS = [
    # MW-01 Benzene (VOC): detected both events, 2026-04 exceeds SL=5
    {"LocationID": "MW-01", "AnalyteName": "Benzene", "ResultValue": "5.2",
     "ResultQualifier": "", "ReportedUnits": "ug/L", "SampleDate": "2026-01-15"},
    {"LocationID": "MW-01", "AnalyteName": "Benzene", "ResultValue": "12.0",
     "ResultQualifier": "", "ReportedUnits": "ug/L", "SampleDate": "2026-04-15"},
    # MW-01 Toluene (VOC): ND then detected
    {"LocationID": "MW-01", "AnalyteName": "Toluene", "ResultValue": "ND",
     "ResultQualifier": "ND", "ReportedUnits": "ug/L", "SampleDate": "2026-01-15"},
    {"LocationID": "MW-01", "AnalyteName": "Toluene", "ResultValue": "8.5",
     "ResultQualifier": "", "ReportedUnits": "ug/L", "SampleDate": "2026-04-15"},
    # MW-02 Benzene (VOC): ND both
    {"LocationID": "MW-02", "AnalyteName": "Benzene", "ResultValue": "ND",
     "ResultQualifier": "ND", "ReportedUnits": "ug/L", "SampleDate": "2026-01-15"},
    {"LocationID": "MW-02", "AnalyteName": "Benzene", "ResultValue": "ND",
     "ResultQualifier": "ND", "ReportedUnits": "ug/L", "SampleDate": "2026-04-15"},
    # MW-02 Lead (Metals): detected, exceeds SL=15
    {"LocationID": "MW-02", "AnalyteName": "Lead", "ResultValue": "40.0",
     "ResultQualifier": "", "ReportedUnits": "ug/L", "SampleDate": "2026-01-15"},
    {"LocationID": "MW-02", "AnalyteName": "Lead", "ResultValue": "10.0",
     "ResultQualifier": "", "ReportedUnits": "ug/L", "SampleDate": "2026-04-15"},
]

_SL = {"Benzene": 5.0, "Toluene": 1000.0, "Lead": 15.0}
_GROUP_MAP = {"Benzene": "VOC", "Toluene": "VOC", "Lead": "Metals"}


def test_build_specs_groups_by_group_map():
    specs = build_appendix_sheet_specs(_ROWS, screening_levels=_SL,
                                       group_map=_GROUP_MAP)
    by_group = {s.analyte_group: s for s in specs}
    assert set(by_group) == {"VOC", "Metals"}
    assert sorted(by_group["VOC"].analytes) == ["Benzene", "Toluene"]
    assert by_group["Metals"].analytes == ["Lead"]
    assert by_group["VOC"].screening_levels["Benzene"] == 5.0


def test_build_specs_default_group_when_no_map():
    specs = build_appendix_sheet_specs(_ROWS, screening_levels=_SL)
    # Without a group map every analyte falls into one default group.
    assert len(specs) == 1


def test_workbook_one_sheet_per_group(tmp_path):
    specs = build_appendix_sheet_specs(_ROWS, screening_levels=_SL,
                                       group_map=_GROUP_MAP)
    out = tmp_path / "appendix.xlsx"
    res = write_appendix_workbook(_ROWS, specs, out, site_id="SITE1")
    assert out.exists()
    wb = load_workbook(out)
    assert "VOC" in wb.sheetnames and "Metals" in wb.sheetnames
    assert res.sheet_count == 2


def test_well_and_event_counts(tmp_path):
    specs = build_appendix_sheet_specs(_ROWS, screening_levels=_SL,
                                       group_map=_GROUP_MAP)
    out = tmp_path / "appendix.xlsx"
    res = write_appendix_workbook(_ROWS, specs, out)
    assert res.well_count == 2
    assert res.event_count == 2


def _find_cell(ws, analyte):
    """Return (row, header_col_offset) for the data row whose first col == analyte."""
    for row in ws.iter_rows():
        if row[0].value == analyte:
            return row
    raise AssertionError(f"analyte row {analyte!r} not found")


def test_nd_cell_no_fill(tmp_path):
    specs = build_appendix_sheet_specs(_ROWS, screening_levels=_SL,
                                       group_map=_GROUP_MAP)
    out = tmp_path / "appendix.xlsx"
    write_appendix_workbook(_ROWS, specs, out)
    ws = load_workbook(out)["VOC"]
    row = _find_cell(ws, "Benzene")
    # MW-02 columns are ND -> value "ND", no solid fill.
    nd_cells = [c for c in row[2:] if c.value == "ND"]
    assert nd_cells, "expected at least one ND cell"
    for c in nd_cells:
        assert c.fill.fill_type != "solid"


def test_detected_cell_yellow(tmp_path):
    specs = build_appendix_sheet_specs(_ROWS, screening_levels=_SL,
                                       group_map=_GROUP_MAP)
    out = tmp_path / "appendix.xlsx"
    write_appendix_workbook(_ROWS, specs, out)
    ws = load_workbook(out)["VOC"]
    row = _find_cell(ws, "Toluene")
    # Toluene 8.5 (below SL 1000) -> detected, yellow.
    yellow = [c for c in row[2:]
              if c.fill.fill_type == "solid"
              and c.fill.fgColor.rgb.endswith("FFFF99")]
    assert yellow, "expected a yellow detected cell for Toluene"


def test_exceedance_cell_red(tmp_path):
    specs = build_appendix_sheet_specs(_ROWS, screening_levels=_SL,
                                       group_map=_GROUP_MAP)
    out = tmp_path / "appendix.xlsx"
    write_appendix_workbook(_ROWS, specs, out)
    ws = load_workbook(out)["VOC"]
    row = _find_cell(ws, "Benzene")
    # Benzene 12.0 exceeds SL 5.0 -> red.
    red = [c for c in row[2:]
           if c.fill.fill_type == "solid"
           and c.fill.fgColor.rgb.endswith("FF9999")]
    assert red, "expected a red exceedance cell for Benzene"


def test_summary_detection_frequency_format(tmp_path):
    specs = build_appendix_sheet_specs(_ROWS, screening_levels=_SL,
                                       group_map=_GROUP_MAP)
    out = tmp_path / "appendix.xlsx"
    write_appendix_workbook(_ROWS, specs, out)
    ws = load_workbook(out)["VOC"]
    detect_row = _find_cell(ws, "Detects")
    freqs = [c.value for c in detect_row[2:] if isinstance(c.value, str)
             and "/" in str(c.value)]
    assert freqs, "expected d/n detection-frequency cells in summary"
