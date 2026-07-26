"""Tests for survey_schema — XLSForm reader, validator, differ. Arcpy-free."""
import openpyxl
import pytest

from autogis.core.envmon.survey_schema import (
    CLASS_DESTRUCTIVE, CLASS_REVIEW, CLASS_SAFE,
    FormSchema, SurveyQuestion, read_xlsform,
)


def make_form(tmp_path, survey_rows, choices_rows=None, settings=None,
              survey_headers=("type", "name", "label", "hint", "required",
                              "calculation", "appearance", "default"),
              name="form.xlsx"):
    """Write a minimal XLSForm workbook and return its path."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("survey")
    ws.append(list(survey_headers))
    for r in survey_rows:
        ws.append(list(r))
    cs = wb.create_sheet("choices")
    cs.append(["list_name", "name", "label"])
    for r in (choices_rows or []):
        cs.append(list(r))
    if settings is not None:
        st = wb.create_sheet("settings")
        st.append(list(settings.keys()))
        st.append(list(settings.values()))
    p = tmp_path / name
    wb.save(p)
    return p


def test_reader_maps_columns_by_header_not_position(tmp_path):
    p = make_form(
        tmp_path,
        [("Notes", "text", "Notes label")],         # data in header order
        survey_headers=("name", "type", "label"),   # shuffled headers
    )
    schema = read_xlsform(p)
    q = schema.questions[0]
    assert (q.type, q.name, q.label) == ("text", "Notes", "Notes label")
    assert q.calculation == ""   # absent column -> ""


def test_reader_choices_and_settings(tmp_path):
    p = make_form(
        tmp_path,
        [("select_one yn", "Q1", "Q")],
        choices_rows=[("yn", "yes", "Yes"), ("yn", "no", "No")],
        settings={"form_title": "T", "form_id": "t_form", "version": "1"},
    )
    s = read_xlsform(p)
    assert s.choices["yn"] == [("yes", "Yes"), ("no", "No")]
    assert s.settings["form_id"] == "t_form"


def test_reader_missing_sheets_yield_empty_structures(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "survey"
    wb.active.append(["type", "name"])
    p = tmp_path / "bare.xlsx"
    wb.save(p)
    s = read_xlsform(p)
    assert s.questions == [] and s.choices == {} and s.settings == {}


def test_reader_rejects_a_workbook_with_no_survey_sheet(tmp_path):
    """Not an XLSForm. Returning an empty schema made diff_forms report every
    question in the OTHER form as a safe addition, so a mistyped
    --baseline-form produced a clean bill of health on a publication gate."""
    wb = openpyxl.Workbook()
    wb.active.title = "Data"
    wb.active.append(["a", "b"])
    p = tmp_path / "not_a_form.xlsx"
    wb.save(p)
    with pytest.raises(ValueError, match="not an XLSForm"):
        read_xlsform(p)


def test_reader_accepts_capitalized_sheet_names(tmp_path):
    """Excel round-trips capitalize sheet names; such a form is still valid,
    and treating it as unreadable would be the same false-negative."""
    wb = openpyxl.Workbook()
    wb.active.title = "Survey"
    wb.active.append(["type", "name"])
    wb.active.append(["text", "A"])
    cs = wb.create_sheet("Choices")
    cs.append(["list_name", "name", "label"])
    cs.append(["l1", "c1", "C1"])
    st = wb.create_sheet("Settings")
    st.append(["form_id"]); st.append(["f1"])
    p = tmp_path / "caps.xlsx"
    wb.save(p)
    s = read_xlsform(p)
    assert [q.name for q in s.questions] == ["A"]
    assert s.choices == {"l1": [("c1", "C1")]}
    assert s.settings["form_id"] == "f1"


def test_reader_keeps_structural_rows_in_order(tmp_path):
    p = make_form(tmp_path, [
        ("begin_group", "g1", "G"),
        ("decimal", "d1", "D"),
        ("end_group", "", ""),
    ])
    s = read_xlsform(p)
    assert [q.type for q in s.questions] == ["begin_group", "decimal",
                                             "end_group"]


# ---------------------------------------------------------------- validate

from autogis.core.common.qa import QACollector
from autogis.core.envmon.survey_schema import validate_form
from autogis.core.envmon.sample_id import xform_sample_id_calc


def _validate(tmp_path, survey_rows, choices_rows=None, settings=None, **kw):
    p = make_form(tmp_path, survey_rows, choices_rows, settings)
    qa = QACollector()
    validate_form(read_xlsform(p), qa, **kw)
    return qa


def _cats(qa):
    return {r.category for r in qa.records}


def test_missing_choices_sheet_is_error(tmp_path):
    wb = openpyxl.Workbook(); wb.active.title = "survey"
    wb.active.append(["type", "name"]); wb.active.append(["text", "A"])
    p = tmp_path / "f.xlsx"; wb.save(p)
    qa = QACollector(); validate_form(read_xlsform(p), qa)
    assert "missing_sheet" in _cats(qa)          # choices missing -> ERROR
    assert any(r.category == "settings_incomplete" and r.severity == "WARNING"
               for r in qa.records)              # settings missing -> WARNING


def test_duplicate_and_invalid_names(tmp_path):
    qa = _validate(tmp_path, [("text", "A", ""), ("text", "a", ""),
                              ("text", "9bad", "")])
    assert "duplicate_name" in _cats(qa)
    assert "invalid_name" in _cats(qa)


def test_xlsform_legal_question_names_accepted(tmp_path):
    """xlsform.org: names may contain letters, digits, hyphens, underscores
    and periods after a leading letter/underscore."""
    qa = _validate(tmp_path, [("text", "Depth.ft", ""),
                              ("text", "Depth-to-water", ""),
                              ("text", "_private", "")])
    assert "invalid_name" not in _cats(qa)


@pytest.mark.parametrize("code", ["1", "2", "101-MW", "2A", "a b"])
def test_choice_values_are_not_identifiers(tmp_path, code):
    """xlsform.org constrains choice VALUES exactly once -- no spaces for
    select_multiple. Likert codes 1/2 and location ids like 101-MW are legal,
    and build_xlsform writes location_ids verbatim as choice codes, so the
    identifier rule made validate-survey-form reject its own builder."""
    qa = _validate(tmp_path,
                   [("select_one l1", "Q", "")],
                   [("l1", code, "Label")])
    assert "invalid_choice_name" not in _cats(qa)


def test_select_multiple_choice_with_space_is_error(tmp_path):
    """The one real constraint: spaces separate saved multi-select values."""
    qa = _validate(tmp_path,
                   [("select_multiple l1", "Q", "")],
                   [("l1", "has space", "Label")])
    assert "invalid_choice_name" in _cats(qa)


def test_empty_choice_name_is_still_an_error(tmp_path):
    qa = _validate(tmp_path,
                   [("select_one l1", "Q", "")],
                   [("l1", "", "Label"), ("l1", "ok", "OK")])
    assert "invalid_choice_name" in _cats(qa)


def test_allow_choice_duplicates_suppresses_duplicate_choice(tmp_path):
    """xlsform.org documents this setting for cascading selects."""
    rows = [("select_one l1", "Q", "")]
    dupes = [("l1", "c1", "A"), ("l1", "c1", "B")]
    assert "duplicate_choice" in _cats(_validate(tmp_path, rows, dupes))
    qa = _validate(tmp_path, rows, dupes,
                   settings={"form_id": "f", "form_title": "t",
                             "version": "1", "allow_choice_duplicates": "yes"})
    assert "duplicate_choice" not in _cats(qa)


def test_reference_to_repeat_group_and_note_resolves(tmp_path):
    """count(${rep}) and indexed-repeat are standard XLSForm; a note is a
    legal relevant target. Building the name set from field-type rows only
    made every one of these a false unresolved_reference."""
    qa = _validate(tmp_path, [
        ("begin_repeat", "rep", "R"),
        ("text", "Depth", ""),
        ("end_repeat", "", ""),
        ("note", "n1", "hello"),
        ("calculate", "C", "", "", "", "count(${rep})"),
        ("calculate", "D", "", "", "", "indexed-repeat(${Depth}, ${rep}, 1)"),
        ("text", "E", "", "", "", "", "", "", "${n1}"),
    ])
    assert "unresolved_reference" not in _cats(qa)


def test_unknown_reference_is_still_an_error(tmp_path):
    qa = _validate(tmp_path,
                   [("calculate", "C", "", "", "", "${nope}")])
    assert "unresolved_reference" in _cats(qa)


def test_skipped_cross_checks_are_reported(tmp_path):
    """Without a config, checks 11-13 do not run. Returning silently let a
    bare `validate-survey-form form.xlsx` report PASS as though the form had
    been checked against its configs."""
    qa = _validate(tmp_path, [("text", "A", "")])
    assert "cross_checks_skipped" in _cats(qa)


def test_cross_checks_not_reported_as_skipped_when_config_given(tmp_path):
    qa = _validate(tmp_path, [("text", "A", "")], event_config=EVENT)
    assert "cross_checks_skipped" not in _cats(qa)


def test_missing_name_and_unknown_type(tmp_path):
    qa = _validate(tmp_path, [("text", "", ""), ("wat", "W", "")])
    assert "missing_name" in _cats(qa)
    assert "unknown_type" in _cats(qa)


def test_select_list_checks(tmp_path):
    qa = _validate(
        tmp_path,
        [("select_one nolist", "Q1", ""), ("select_one empty", "Q2", ""),
         ("select_one", "Q3", "")],
        choices_rows=[("empty", "", ""),               # invalid choice name
                      ("dup", "x", "X"), ("dup", "x", "X2")],
    )
    for cat in ("unknown_choice_list", "empty_choice_list", "unknown_type",
                "duplicate_choice", "invalid_choice_name"):
        assert cat in _cats(qa), cat


def test_unresolved_reference_and_backward_ref_ok(tmp_path):
    qa = _validate(tmp_path, [
        ("text", "A", ""),
        ("calculate", "C1", "", "", "", "concat(${A}, ${Missing})"),
        ("calculate", "C2", "", "", "", "${Later}"),
        ("text", "Later", ""),
    ])
    recs = [r for r in qa.records if r.category == "unresolved_reference"]
    assert len(recs) == 1 and "Missing" in recs[0].message


def test_unbalanced_group_detected(tmp_path):
    qa = _validate(tmp_path, [("begin_group", "g", ""), ("text", "A", ""),
                              ("end_repeat", "", "")])
    assert "unbalanced_group" in _cats(qa)


def test_odd_required_value_warns(tmp_path):
    qa = _validate(tmp_path, [("text", "A", "", "", "maybe")])
    rec = next(r for r in qa.records if r.category == "odd_required_value")
    assert rec.severity == "WARNING"


def test_sample_id_missing_is_error(tmp_path):
    qa = _validate(tmp_path, [("text", "A", "")])
    assert "sample_id_missing" in _cats(qa)


def test_sample_id_contract_match_and_mismatch(tmp_path):
    good = [("select_one well_list", "WellID", "", "", "yes"),
            ("date", "SamplingDate", "", "", "yes"),
            ("select_one matrix_list", "Matrix", "", "", "yes"),
            ("select_multiple qa_flags", "QAFlags", ""),
            ("calculate", "SampleID", "", "", "", xform_sample_id_calc())]
    choices = [("well_list", "MW-1", "MW-1"), ("matrix_list", "GW", "GW"),
               ("qa_flags", "field_dup", "Field Duplicate")]
    qa = _validate(tmp_path, good, choices)
    assert "sample_id_contract_mismatch" not in _cats(qa)

    bad = list(good); bad[4] = ("calculate", "SampleID", "", "", "", "${WellID}")
    qa2 = _validate(tmp_path, bad, choices)
    rec = next(r for r in qa2.records
               if r.category == "sample_id_contract_mismatch")
    assert rec.severity == "ERROR"


def test_sample_id_dup_leg_requires_flag_choice(tmp_path):
    rows = [("select_one well_list", "WellID", "", "", "yes"),
            ("date", "SamplingDate", "", "", "yes"),
            ("select_one matrix_list", "Matrix", "", "", "yes"),
            ("select_multiple qa_flags", "QAFlags", ""),
            ("calculate", "SampleID", "", "", "", xform_sample_id_calc())]
    choices = [("well_list", "MW-1", "MW-1"), ("matrix_list", "GW", "GW"),
               ("qa_flags", "resampled", "Resampled")]   # no field_dup
    qa = _validate(tmp_path, rows, choices)
    assert "sample_id_dup_leg" in _cats(qa)


def test_validation_complete_summary_present(tmp_path):
    qa = _validate(tmp_path, [("text", "A", "")])
    assert any(r.category == "validation_complete" and r.severity == "INFO"
               for r in qa.records)


# ------------------------------------------------------------- cross-refs

SITE = {"site_id": "H281", "site_name": "H281 Glasgow"}
EVENT = {
    "analyte_groups": {"VOCs": ["Benzene", "Toluene"], "Metals": ["Arsenic"]},
    "crew_list": ["Alice Smith", "Bob Jones"],
    "coc_prefix": "H281-COC", "matrices": ["GW"],
    # "101-MW" is deliberate: build_xlsform writes location_ids verbatim as
    # choice codes, and a leading digit is legal for a choice value. With only
    # MW-* names the round-trip test passed while validate-survey-form still
    # rejected build-survey-form's own output for real sites.
    "location_ids": ["MW-1", "MW-2", "MW-3", "101-MW"],
}
ADICT = {"analytes": {
    "Benzene": {"abbreviation": "B",
                "default_units_by_matrix": {"GW": "ug/L"}},
    "Toluene": {"abbreviation": "T",
                "default_units_by_matrix": {"GW": "ug/L"}},
    "Arsenic": {"abbreviation": "As",
                "default_units_by_matrix": {"GW": "ug/L"}},
}}


def test_generated_form_validates_clean_against_its_configs(tmp_path):
    """Round-trip lockstep: builder output + same configs -> zero findings."""
    from autogis.core.envmon.survey123_form_builder import build_xlsform
    wb = build_xlsform(SITE, EVENT, ADICT)
    p = tmp_path / "gen.xlsx"
    wb.save(p)
    qa = QACollector()
    validate_form(read_xlsform(p), qa, event_config=EVENT,
                  site_config=SITE, analyte_dict=ADICT)
    bad = [r for r in qa.records if r.severity in ("ERROR", "WARNING")]
    assert bad == [], [f"{r.category}: {r.message}" for r in bad]


def test_missing_location_and_extra_crew_flagged(tmp_path):
    from autogis.core.envmon.survey123_form_builder import build_xlsform
    smaller = {**EVENT, "location_ids": ["MW-1"],
               "crew_list": ["Alice Smith", "Bob Jones", "Cara Doe"]}
    wb = build_xlsform(SITE, smaller, ADICT)
    p = tmp_path / "gen2.xlsx"
    wb.save(p)
    qa = QACollector()
    validate_form(read_xlsform(p), qa, event_config=EVENT,
                  analyte_dict=ADICT)   # validate against the BIGGER event
    cats = _cats(qa)
    assert "missing_location_choice" in cats     # MW-2/MW-3 not in form
    assert "extra_crew_choice" in cats           # cara_doe extra


def test_missing_analyte_question_flagged(tmp_path):
    from autogis.core.envmon.survey123_form_builder import build_xlsform
    fewer = {**EVENT, "analyte_groups": {"VOCs": ["Benzene"]}}
    wb = build_xlsform(SITE, fewer, ADICT)
    p = tmp_path / "gen3.xlsx"
    wb.save(p)
    qa = QACollector()
    validate_form(read_xlsform(p), qa, event_config=EVENT,
                  analyte_dict=ADICT)
    assert "missing_analyte_question" in _cats(qa)   # Toluene, Arsenic


def test_unknown_analyte_warns_with_dict(tmp_path):
    from autogis.core.envmon.survey123_form_builder import build_xlsform
    wb = build_xlsform(SITE, EVENT, ADICT)
    p = tmp_path / "gen4.xlsx"
    wb.save(p)
    event = {**EVENT,
             "analyte_groups": {**EVENT["analyte_groups"],
                                "VOCs": ["Benzene", "Toluene", "Xylene"]}}
    qa = QACollector()
    validate_form(read_xlsform(p), qa, event_config=event,
                  analyte_dict=ADICT)
    assert "unknown_analyte" in _cats(qa)


# -------------------------------------------------------------- diff_forms

from autogis.core.envmon.survey_schema import (
    diff_forms, worst_classification, SchemaChange,
)


BASE_ROWS = [
    ("text", "A", "Label A", "", "yes"),
    ("select_one lst", "S", "Sel"),
    ("begin_repeat", "rep", "R"),
    ("decimal", "InRep", "d"),
    ("end_repeat", "", ""),
    ("calculate", "C", "", "", "", "${A}"),
    ("calculate", "SampleID", "", "", "", "x"),
]
BASE_CHOICES = [("lst", "one", "One"), ("lst", "two", "Two")]
BASE_SETTINGS = {"form_title": "T", "form_id": "t_id", "version": "1"}


def _diff(tmp_path, new_rows, new_choices=None, new_settings=None):
    old = read_xlsform(make_form(tmp_path, BASE_ROWS, BASE_CHOICES,
                                 dict(BASE_SETTINGS), name="old.xlsx"))
    new = read_xlsform(make_form(
        tmp_path, new_rows,
        BASE_CHOICES if new_choices is None else new_choices,
        dict(BASE_SETTINGS) if new_settings is None else new_settings,
        name="new.xlsx"))
    return diff_forms(old, new)


def _kinds(changes):
    return {(c.kind, c.classification) for c in changes}


def test_diff_taxonomy_rows(tmp_path):
    rows = list(BASE_ROWS)

    # unchanged -> no changes
    assert _diff(tmp_path, rows) == []

    # added optional / added required
    ch = _kinds(_diff(tmp_path, rows + [("text", "NewOpt", "")]))
    assert ("question_added", CLASS_SAFE) in ch
    ch = _kinds(_diff(tmp_path, rows + [("text", "NewReq", "", "", "yes")]))
    assert ("question_added_required", CLASS_REVIEW) in ch

    # removed -> destructive
    ch = _kinds(_diff(tmp_path, [r for r in rows if r[1] != "A"]))
    assert ("question_removed", CLASS_DESTRUCTIVE) in ch

    # type change -> destructive
    ch = _kinds(_diff(tmp_path,
                      [("integer", *r[1:]) if r[1] == "A" else r for r in rows]))
    assert ("type_changed", CLASS_DESTRUCTIVE) in ch

    # required flips
    ch = _kinds(_diff(tmp_path,
                      [("text", "A", "Label A", "", "") if r[1] == "A" else r
                       for r in rows]))
    assert ("required_relaxed", CLASS_SAFE) in ch
    rows_opt = [("text", "A", "Label A", "", "") if r[1] == "A" else r
                for r in rows]
    old_opt = read_xlsform(make_form(tmp_path, rows_opt, BASE_CHOICES,
                                     dict(BASE_SETTINGS), name="o2.xlsx"))
    new_req = read_xlsform(make_form(tmp_path, rows, BASE_CHOICES,
                                     dict(BASE_SETTINGS), name="n2.xlsx"))
    assert ("required_tightened", CLASS_REVIEW) in _kinds(
        diff_forms(old_opt, new_req))

    # calculation changes
    ch = _kinds(_diff(tmp_path,
                      [("calculate", "C", "", "", "", "${A} + 1")
                       if r[1] == "C" else r for r in rows]))
    assert ("calculation_changed", CLASS_REVIEW) in ch
    ch = _kinds(_diff(tmp_path,
                      [("calculate", "SampleID", "", "", "", "y")
                       if r[1] == "SampleID" else r for r in rows]))
    assert ("sample_id_calculation_changed", CLASS_DESTRUCTIVE) in ch

    # list re-pointed
    ch = _kinds(_diff(tmp_path,
                      [("select_one other", "S", "Sel") if r[1] == "S" else r
                       for r in rows],
                      new_choices=BASE_CHOICES + [("other", "x", "X")]))
    assert ("list_repointed", CLASS_REVIEW) in ch

    # repeat scope move -> destructive
    moved = [("text", "A", "Label A", "", "yes"),
             ("select_one lst", "S", "Sel"),
             ("begin_repeat", "rep", "R"),
             ("end_repeat", "", ""),
             ("decimal", "InRep", "d"),
             ("calculate", "C", "", "", "", "${A}"),
             ("calculate", "SampleID", "", "", "", "x")]
    assert ("repeat_scope_changed", CLASS_DESTRUCTIVE) in _kinds(
        _diff(tmp_path, moved))

    # repeat added / removed -> destructive
    ch = _kinds(_diff(tmp_path, rows + [("begin_repeat", "rep2", "R2"),
                                        ("end_repeat", "", "")]))
    assert ("repeat_added", CLASS_DESTRUCTIVE) in ch
    norep = [r for r in rows if r[1] not in ("rep", "InRep")
             and r[0] != "end_repeat"]
    assert ("repeat_removed", CLASS_DESTRUCTIVE) in _kinds(
        _diff(tmp_path, norep))

    # cosmetic -> safe
    ch = _kinds(_diff(tmp_path,
                      [("text", "A", "Renamed label", "", "yes")
                       if r[1] == "A" else r for r in rows]))
    assert ("cosmetic_changed", CLASS_SAFE) in ch


def test_diff_choice_rows(tmp_path):
    rows = list(BASE_ROWS)
    # choice added / label changed -> safe
    ch = _kinds(_diff(tmp_path, rows,
                      new_choices=BASE_CHOICES + [("lst", "three", "Three")]))
    assert ("choice_added", CLASS_SAFE) in ch
    ch = _kinds(_diff(tmp_path, rows,
                      new_choices=[("lst", "one", "Uno"),
                                   ("lst", "two", "Two")]))
    assert ("choice_label_changed", CLASS_SAFE) in ch
    # choice removed -> review
    ch = _kinds(_diff(tmp_path, rows, new_choices=[("lst", "one", "One")]))
    assert ("choice_removed", CLASS_REVIEW) in ch
    # code change (same label, new code) -> destructive
    ch = _kinds(_diff(tmp_path, rows,
                      new_choices=[("lst", "uno", "One"),
                                   ("lst", "two", "Two")]))
    assert ("choice_code_changed", CLASS_DESTRUCTIVE) in ch


def test_diff_settings_and_worst(tmp_path):
    rows = list(BASE_ROWS)
    ch = _diff(tmp_path, rows,
               new_settings={"form_title": "T", "form_id": "other",
                             "version": "2"})
    kinds = _kinds(ch)
    assert ("form_id_changed", CLASS_DESTRUCTIVE) in kinds
    assert ("settings_changed", CLASS_SAFE) in kinds
    assert worst_classification(ch) == CLASS_DESTRUCTIVE
    assert worst_classification([]) is None


# --------------------------------------------------------- form vs layer

from autogis.core.envmon.survey_schema import (
    diff_form_vs_layer, form_layer_fields,
)

LAYER_SPEC = {
    "layer_name": "MonitoringWells",
    "fields": [
        {"name": "WellID", "type": "esriFieldTypeString"},
        {"name": "SamplingDate", "type": "esriFieldTypeDate"},
        {"name": "Matrix", "type": "esriFieldTypeString",
         "domain": {"name": "MatrixDom",
                    "coded_values": [{"code": "GW", "name": "Groundwater"}]}},
        {"name": "Sampled", "type": "esriFieldTypeSmallInteger"},
    ],
}

FORM_ROWS = [
    ("select_one well_list", "WellID", "Well"),
    ("date", "SamplingDate", "Date"),
    ("select_one matrix_list", "Matrix", "Matrix"),
    ("decimal", "DepthToWater_ft", "DTW"),
    ("note", "n1", "note rows map to no field"),
]
FORM_CHOICES = [("well_list", "MW-1", "MW-1"),
                ("matrix_list", "GW", "Groundwater")]


def test_form_layer_fields_mapping(tmp_path):
    s = read_xlsform(make_form(tmp_path, FORM_ROWS, FORM_CHOICES))
    fields = {f["name"]: f for f in form_layer_fields(s)}
    assert fields["WellID"]["type"] == "esriFieldTypeString"
    assert fields["SamplingDate"]["type"] == "esriFieldTypeDate"
    assert fields["DepthToWater_ft"]["type"] == "esriFieldTypeDouble"
    assert "n1" not in fields                       # note -> no field
    assert fields["Matrix"]["domain"]["codedValues"] == [
        {"code": "GW", "name": "Groundwater"}]


def test_diff_form_vs_layer_classifications(tmp_path):
    s = read_xlsform(make_form(tmp_path, FORM_ROWS, FORM_CHOICES))
    changes = diff_form_vs_layer(s, LAYER_SPEC)
    by_kind = {(c.kind, c.classification) for c in changes}
    # form-only field -> review-required
    assert ("extra_field", CLASS_REVIEW) in by_kind          # DepthToWater_ft
    # layer-only field -> safe
    assert ("missing_field", CLASS_SAFE) in by_kind          # Sampled
    # matching domains normalized -> no DOMAIN_DRIFT for Matrix
    assert not any(c.kind == "domain_drift" and c.name == "Matrix"
                   for c in changes)


def test_diff_form_vs_layer_type_mismatch_destructive(tmp_path):
    rows = [("text", "SamplingDate", "now text")]
    s = read_xlsform(make_form(tmp_path, rows))
    changes = diff_form_vs_layer(s, LAYER_SPEC)
    assert any(c.kind == "type_mismatch" and
               c.classification == CLASS_DESTRUCTIVE and
               c.name == "SamplingDate" for c in changes)


# ---------------------------------------------------------------- CLI

def _write_yaml(tmp_path, name, data):
    import yaml
    p = tmp_path / name
    p.write_text(yaml.safe_dump(data), encoding="utf-8")
    return p


def test_both_commands_in_help():
    from click.testing import CliRunner
    from autogis.adapters.cli import autogis
    out = CliRunner().invoke(autogis, ["envmon", "--help"]).output
    assert "validate-survey-form" in out
    assert "diff-survey-schema" in out


def test_validate_cli_pass_and_fail(tmp_path):
    from click.testing import CliRunner
    from autogis.core.envmon.survey123_form_builder import build_xlsform
    from autogis.adapters.cli import autogis
    good = tmp_path / "good.xlsx"
    build_xlsform(SITE, EVENT, ADICT).save(good)
    r = CliRunner().invoke(autogis, ["envmon", "validate-survey-form",
                                     str(good)])
    assert r.exit_code == 0 and "Status: PASS" in r.output

    bad = make_form(tmp_path, [("text", "A", "")], name="bad.xlsx")
    r = CliRunner().invoke(autogis, ["envmon", "validate-survey-form",
                                     str(bad)])
    assert r.exit_code == 1 and "Status: FAIL" in r.output


def test_diff_cli_requires_a_baseline(tmp_path):
    from click.testing import CliRunner
    from autogis.adapters.cli import autogis
    f = make_form(tmp_path, [("text", "A", "")])
    r = CliRunner().invoke(autogis, ["envmon", "diff-survey-schema", str(f)])
    assert r.exit_code != 0
    assert "baseline" in r.output.lower() or "layer" in r.output.lower()


def test_diff_cli_semantic_exit_codes(tmp_path):
    from click.testing import CliRunner
    from autogis.adapters.cli import autogis
    runner = CliRunner()
    old = make_form(tmp_path, [("text", "A", "L")], name="o.xlsx")
    same = make_form(tmp_path, [("text", "A", "L")], name="s.xlsx")
    cosmetic = make_form(tmp_path, [("text", "A", "L2")], name="c.xlsx")
    newreq = make_form(tmp_path, [("text", "A", "L"),
                                  ("text", "B", "", "", "yes")],
                       name="r.xlsx")
    removed = make_form(tmp_path, [("text", "Z", "L")], name="d.xlsx")

    assert runner.invoke(autogis, ["envmon", "diff-survey-schema", str(same),
                                   "--baseline-form",
                                   str(old)]).exit_code == 0
    assert runner.invoke(autogis, ["envmon", "diff-survey-schema",
                                   str(cosmetic), "--baseline-form",
                                   str(old)]).exit_code == 0     # safe only
    assert runner.invoke(autogis, ["envmon", "diff-survey-schema",
                                   str(newreq), "--baseline-form",
                                   str(old)]).exit_code == 2
    assert runner.invoke(autogis, ["envmon", "diff-survey-schema",
                                   str(removed), "--baseline-form",
                                   str(old)]).exit_code == 3


def test_diff_cli_layer_spec_and_json_report(tmp_path):
    import json
    from click.testing import CliRunner
    from autogis.adapters.cli import autogis
    f = make_form(tmp_path, [("text", "OnlyInForm", "")])
    spec = _write_yaml(tmp_path, "spec.yaml", {
        "layer_name": "L",
        "fields": [{"name": "OnlyInForm", "type": "esriFieldTypeString"}],
    })
    rpt = tmp_path / "out.json"
    r = CliRunner().invoke(autogis, ["envmon", "diff-survey-schema", str(f),
                                     "--layer-spec", str(spec),
                                     "--report", str(rpt)])
    assert r.exit_code == 0
    assert json.loads(rpt.read_text(encoding="utf-8")) == []


def test_new_commands_registered_in_capabilities():
    from autogis.runtime.capabilities import TOOLS, Runtime, TOOL_REGISTRY
    assert TOOLS["validate-survey-form"] is Runtime.CLOUD
    assert TOOLS["diff-survey-schema"] is Runtime.CLOUD
    names = {c.command for c in TOOL_REGISTRY}
    assert {"validate-survey-form", "diff-survey-schema"} <= names


# ------------------------------------------------- review-fix regressions

def test_semantic_exit_logs_run_history_success(tmp_path, monkeypatch):
    """Blocker fix: a destructive diff (exit 3) records status=success in
    run history, not error — a drift finding is not a tool failure."""
    import csv as csv_mod
    from click.testing import CliRunner
    from autogis.adapters.cli import autogis
    hist = tmp_path / "run_history.csv"
    monkeypatch.setenv("AUTOGIS_RUN_HISTORY", str(hist))
    old = make_form(tmp_path, [("text", "A", "L")], name="rh_o.xlsx")
    removed = make_form(tmp_path, [("text", "Z", "L")], name="rh_n.xlsx")
    r = CliRunner().invoke(autogis, ["envmon", "diff-survey-schema",
                                     str(removed), "--baseline-form",
                                     str(old)])
    assert r.exit_code == 3
    with hist.open(newline="", encoding="utf-8") as fh:
        rows = [row for row in csv_mod.DictReader(fh)
                if row["tool_name"] == "diff-survey-schema"]
    assert rows and rows[-1]["status"] == "success"


def test_diff_usage_errors_exit_1_not_2(tmp_path):
    """Should-fix: usage/IO errors exit 1; 2 stays reserved for
    review-required drift."""
    from click.testing import CliRunner
    from autogis.adapters.cli import autogis
    runner = CliRunner()
    f = make_form(tmp_path, [("text", "A", "")])
    # no baseline at all
    assert runner.invoke(autogis, ["envmon", "diff-survey-schema",
                                   str(f)]).exit_code == 1
    # nonexistent baseline file
    assert runner.invoke(autogis, ["envmon", "diff-survey-schema", str(f),
                                   "--baseline-form",
                                   str(tmp_path / "nope.xlsx")]).exit_code == 1
    # nonexistent form file
    assert runner.invoke(autogis, ["envmon", "diff-survey-schema",
                                   str(tmp_path / "gone.xlsx"),
                                   "--baseline-form",
                                   str(f)]).exit_code == 1


def test_missing_select_question_is_error_not_silence(tmp_path):
    """Should-fix: config supplies crew but the form has no SampledBy
    question at all — must ERROR, not pass silently."""
    qa = _validate(tmp_path, [("text", "A", "")],
                   event_config={"crew_list": ["Alice Smith"],
                                 "location_ids": [], "matrices": [],
                                 "analyte_groups": {}})
    recs = [r for r in qa.records if r.category == "missing_crew_choice"]
    assert recs and recs[0].severity == "ERROR"
    assert "SampledBy" in recs[0].message


def test_domain_drift_positive(tmp_path):
    """A form choice absent from the spec's coded values -> DOMAIN_DRIFT,
    review-required."""
    rows = [("select_one matrix_list", "Matrix", "Matrix")]
    choices = [("matrix_list", "GW", "Groundwater"),
               ("matrix_list", "SW", "Surface Water")]   # SW not in spec
    s = read_xlsform(make_form(tmp_path, rows, choices))
    changes = diff_form_vs_layer(s, LAYER_SPEC)
    assert any(c.kind == "domain_drift" and c.name == "Matrix" and
               c.classification == CLASS_REVIEW for c in changes)


def test_group_move_is_review_required(tmp_path):
    old_rows = [("begin_group", "g1", "G1"), ("text", "A", ""),
                ("end_group", "", "")]
    new_rows = [("begin_group", "g2", "G2"), ("text", "A", ""),
                ("end_group", "", "")]
    old = read_xlsform(make_form(tmp_path, old_rows, name="g_o.xlsx"))
    new = read_xlsform(make_form(tmp_path, new_rows, name="g_n.xlsx"))
    assert ("group_changed", CLASS_REVIEW) in _kinds(diff_forms(old, new))


def test_settings_shape_warnings(tmp_path):
    qa = _validate(tmp_path, [("text", "A", "")],
                   settings={"form_title": "T", "form_id": "Not A Slug",
                             "version": ""})
    msgs = [r.message for r in qa.records
            if r.category == "settings_incomplete"]
    assert any("form_id" in m for m in msgs)
    assert any("version" in m for m in msgs)


def test_matrix_missing_and_extra_choices(tmp_path):
    rows = [("select_one matrix_list", "Matrix", "Matrix")]
    choices = [("matrix_list", "GW", "GW"), ("matrix_list", "SOIL", "Soil")]
    qa = _validate(tmp_path, rows, choices,
                   event_config={"matrices": ["GW", "SW"],
                                 "location_ids": [], "crew_list": [],
                                 "analyte_groups": {}})
    cats = _cats(qa)
    assert "missing_matrix_choice" in cats     # SW planned, not in form
    assert "extra_matrix_choice" in cats       # SOIL in form, not planned


def test_unexpected_analyte_question_warns(tmp_path):
    rows = [("begin_group", "grp_vocs", "VOCs"),
            ("decimal", "Benzene", "B"),
            ("decimal", "Stray", "not an analyte"),
            ("end_group", "", "")]
    qa = _validate(tmp_path, rows,
                   event_config={"analyte_groups": {"VOCs": ["Benzene"]},
                                 "location_ids": [], "matrices": [],
                                 "crew_list": []})
    rec = next(r for r in qa.records
               if r.category == "unexpected_analyte_question")
    assert rec.severity == "WARNING" and "Stray" in rec.message


def test_diff_cli_markdown_report(tmp_path):
    from click.testing import CliRunner
    from autogis.adapters.cli import autogis
    old = make_form(tmp_path, [("text", "A", "L")], name="md_o.xlsx")
    new = make_form(tmp_path, [("text", "A", "L2")], name="md_n.xlsx")
    rpt = tmp_path / "out.md"
    r = CliRunner().invoke(autogis, ["envmon", "diff-survey-schema",
                                     str(new), "--baseline-form", str(old),
                                     "--report", str(rpt)])
    assert r.exit_code == 0
    text = rpt.read_text(encoding="utf-8")
    assert "| class | kind | name | detail |" in text
    assert "cosmetic_changed" in text
