"""CLI test for the headless gen-synthetic-workbook command (Tool 10.6)."""
import openpyxl
from click.testing import CliRunner

from autogis.adapters.cli import autogis


def test_gen_synthetic_workbook_in_help():
    result = CliRunner().invoke(autogis, ["envmon", "--help"])
    assert result.exit_code == 0
    assert "gen-synthetic-workbook" in result.output


def test_gen_synthetic_workbook_end_to_end(tmp_path):
    out = tmp_path / "synth.xlsx"
    result = CliRunner().invoke(autogis, [
        "envmon", "gen-synthetic-workbook",
        "--site-id", "TEST01", "--wells", "3", "--events", "2",
        "--features", "nondetects,rpd_sheet", "--seed", "1",
        "--out", str(out),
    ])
    assert result.exit_code == 0, result.output
    assert out.exists()
    wb = openpyxl.load_workbook(out)
    assert "Results" in wb.sheetnames
    assert "RPD" in wb.sheetnames


def test_gen_synthetic_workbook_rejects_unknown_feature(tmp_path):
    out = tmp_path / "synth.xlsx"
    result = CliRunner().invoke(autogis, [
        "envmon", "gen-synthetic-workbook",
        "--features", "not_a_feature", "--out", str(out),
    ])
    assert result.exit_code != 0
