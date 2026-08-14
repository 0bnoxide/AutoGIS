"""Unit + CLI tests for export_comparison_excel."""
import csv
import zipfile
from pathlib import Path

from click.testing import CliRunner
from openpyxl import load_workbook

from autogis.adapters.cli import autogis
from autogis.core.common.qa import QACollector
from autogis.core.envmon.export_comparison_excel import export_comparison_excel


# Fixture uses the REAL vocabulary emitted by compare_events.py:
#   TrendClass from _classify(): INCREASED/DECREASED/STABLE/NEW_DETECTION/...
#   CurrentExceedance from _exc(): "Y"/"N"/""
def _row(trend="STABLE", exceed="N"):
    return {
        "SiteID": "S", "LocationID": "MW-1", "AnalyteCanonicalName": "Benzene",
        "Matrix": "GW", "TrendClass": trend, "CurrentExceedance": exceed,
        "PreviousExceedance": "N", "CurrentResultRaw": "5.0",
        "PreviousResultRaw": "5.0", "Delta": "0", "PercentChange": "0",
    }


def test_basic_export(tmp_path):
    out = tmp_path / "result.xlsx"
    qa = QACollector()
    result = export_comparison_excel([_row()], out, qa=qa)
    assert result == out
    assert out.exists()
    with zipfile.ZipFile(out) as zf:
        names = zf.namelist()
    assert any("sheet1" in n.lower() or "AllResults" in n for n in names)


def test_exceedance_sheet(tmp_path):
    rows = [_row("INCREASED", "Y"), _row("STABLE", "N")]
    out = tmp_path / "r.xlsx"
    qa = QACollector()
    export_comparison_excel(rows, out, qa=qa)
    assert any(r.category == "export_complete" for r in qa.records)


def test_overwrite_false_errors(tmp_path):
    out = tmp_path / "r.xlsx"
    out.write_bytes(b"existing")
    qa = QACollector()
    export_comparison_excel([_row()], out, overwrite=False, qa=qa)
    assert any(r.category == "output_exists" for r in qa.records)


def test_no_records_warns(tmp_path):
    out = tmp_path / "r.xlsx"
    qa = QACollector()
    export_comparison_excel([], out, qa=qa)
    assert any(r.category == "no_records" for r in qa.records)


def test_trend_fill_applied(tmp_path):
    out = tmp_path / "r.xlsx"
    qa = QACollector()
    export_comparison_excel([_row("INCREASED", "N")], out, qa=qa)
    wb = load_workbook(out)
    ws = wb["AllResults"]
    headers = [c.value for c in ws[1]]
    trend_col = headers.index("TrendClass") + 1
    fill = ws.cell(row=2, column=trend_col).fill
    assert fill.fill_type == "solid"
    assert fill.fgColor.rgb.endswith("FFCCCC")  # INCREASE -> red tint


def test_exceedance_sheet_only_exceedances(tmp_path):
    out = tmp_path / "r.xlsx"
    qa = QACollector()
    export_comparison_excel([_row("INCREASED", "Y"), _row("STABLE", "N")], out, qa=qa)
    ws = load_workbook(out)["Exceedances"]
    # header + exactly one exceedance data row
    assert ws.max_row == 2


# --- CLI --------------------------------------------------------------------

def _write_comparison_csv(path, rows):
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)


def test_export_comparison_excel_in_help():
    result = CliRunner().invoke(autogis, ["envmon", "--help"])
    assert "export-comparison-excel" in result.output


def test_cli_export_comparison_excel_end_to_end(tmp_path):
    src = tmp_path / "comparison.csv"
    _write_comparison_csv(src, [_row("INCREASED", "Y"), _row("STABLE", "N")])
    out = tmp_path / "out.xlsx"
    result = CliRunner().invoke(autogis, [
        "envmon", "export-comparison-excel",
        "--comparison-csv", str(src), "--output", str(out),
    ])
    assert result.exit_code == 0, result.output
    assert out.exists()
    assert "AllResults" in load_workbook(out).sheetnames
