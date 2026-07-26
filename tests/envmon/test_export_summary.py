import dataclasses
import csv as _csv
from datetime import date
from pathlib import Path

import openpyxl
import pytest
from autogis.core.envmon.gdb_schema import SampleRecord, AnalyticalResultRecord
from autogis.core.envmon.export_summary import export_analytical_summary


def _r(sample_id, analyte, numeric, exceeds=0, is_det=1, units="ug/L"):
    return AnalyticalResultRecord(
        ImportBatchID="B1", SiteID="H281", Matrix="GROUNDWATER",
        LocationID="MW-1", SampleID=sample_id, ParentSampleID="",
        SampleDate=date(2026, 1, 1),
        DepthTop_ft=None, DepthBottom_ft=None, DepthIntervalText="",
        AnalyticalGroup="VOC", MethodGroup="",
        AnalyteName=analyte, AnalyteCanonicalName=analyte,
        AnalyteAbbreviation=analyte[:8],
        ResultRawText=str(numeric) if numeric else "<1.0",
        ResultNumeric=numeric, ReportingLimit=1.0, DetectionLimit=None,
        Units=units, Qualifier="",
        IsNonDetect=0 if is_det else 1, IsDetected=is_det,
        IsEstimated=0, IsDiluted=0, IsNotAnalyzed=0,
        IsNotSampled=0, IsNotMeasured=0,
        ScreeningLevel=5.0, ScreeningLevelSource="config",
        ExceedsScreeningLevel=exceeds,
        DisplayText=str(numeric), DisplayColorClass="EXCEEDANCE" if exceeds else "DETECTED",
        SourceWorkbook="edd.csv", SourceSheet="", SourceRow=0,
        SourceColumn="", SourceCell="")


def test_export_summary_creates_four_sheets(tmp_path):
    """With no site_id/event_id/samples, only the four result sheets appear
    (no site_id passed -- the fourth positional arg is required, so pass "")."""
    results = [_r("S1", "Benzene", 10.0, exceeds=1),
               _r("S1", "Toluene", 2.0, exceeds=0),
               _r("S2", "Benzene", 0.5, exceeds=0)]
    path = export_analytical_summary([], results, tmp_path / "out.xlsx", "")
    wb = openpyxl.load_workbook(path)
    assert set(wb.sheetnames) == {"All Results", "Detections", "Exceedances",
                                  "Summary by Analyte"}


def test_export_summary_site_and_event_id_produce_metadata_sheet(tmp_path):
    """Issue #343: site_id/event_id were accepted but silently never appeared
    anywhere in the output."""
    results = [_r("S1", "Benzene", 10.0, exceeds=1)]
    path = export_analytical_summary([], results, tmp_path / "out.xlsx",
                                     "H281", event_id="2026Q2")
    wb = openpyxl.load_workbook(path)
    assert "Metadata" in wb.sheetnames
    ws = wb["Metadata"]
    rows = {r[0]: r[1] for r in ws.iter_rows(min_row=2, values_only=True)}
    assert rows["SiteID"] == "H281"
    assert rows["EventID"] == "2026Q2"


def test_export_summary_samples_produce_samples_sheet(tmp_path):
    """Issue #343: samples were accepted but silently never appeared anywhere
    in the output -- an empty list (as every existing test passed) hid this."""
    from datetime import date as _date
    sample = SampleRecord(
        ImportBatchID="B1", SiteID="H281", Matrix="GROUNDWATER",
        LocationID="MW-1", SampleID="S1", ParentSampleID="",
        SampleDate=_date(2026, 1, 1), SampleDateRaw="1/1/2026",
        DepthTop_ft=None, DepthBottom_ft=None, DepthIntervalText="",
        IsDuplicate=0, DuplicateType="", LabSampleID="LAB-1",
        SourceWorkbook="edd.xlsx", SourceSheet="Samples", SourceRow=2)
    results = [_r("S1", "Benzene", 10.0, exceeds=1)]
    path = export_analytical_summary([sample], results, tmp_path / "out.xlsx", "H281")
    wb = openpyxl.load_workbook(path)
    assert "Samples" in wb.sheetnames
    ws = wb["Samples"]
    header = [c.value for c in ws[1]]
    row = dict(zip(header, next(ws.iter_rows(min_row=2, values_only=True))))
    assert row["SampleID"] == "S1"
    assert row["LabSampleID"] == "LAB-1"


def test_exceedances_sheet_filters_correctly(tmp_path):
    results = [_r("S1", "Benzene", 10.0, exceeds=1),
               _r("S1", "Toluene", 2.0, exceeds=0)]
    path = export_analytical_summary([], results, tmp_path / "out.xlsx", "H281")
    wb = openpyxl.load_workbook(path)
    ws = wb["Exceedances"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    analyte_col = [c.value for c in ws[1]].index("AnalyteName")
    assert rows[0][analyte_col] == "Benzene"


def test_summary_sheet_counts_by_analyte(tmp_path):
    results = [_r("S1", "Benzene", 10.0, exceeds=1),
               _r("S2", "Benzene", 2.0, exceeds=0),
               _r("S1", "Toluene", 0.5, exceeds=0)]
    path = export_analytical_summary([], results, tmp_path / "out.xlsx", "H281")
    wb = openpyxl.load_workbook(path)
    ws = wb["Summary by Analyte"]
    header = [c.value for c in ws[1]]
    rows = {r[0]: r for r in ws.iter_rows(min_row=2, values_only=True)}
    assert "Benzene" in rows
    exc_col = header.index("ExceedanceCount")
    det_col = header.index("DetectionCount")
    assert rows["Benzene"][exc_col] == 1
    assert rows["Benzene"][det_col] == 2


# --- Task 4b: CLI ---
from click.testing import CliRunner
from autogis.adapters.cli import autogis as _cli


def _write(path, rc_class, recs):
    fnames = [f.name for f in dataclasses.fields(rc_class)]
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = _csv.DictWriter(fh, fieldnames=fnames)
        w.writeheader()
        for r in recs:
            w.writerow(dataclasses.asdict(r))


def test_export_summary_cli(tmp_path):
    results = [_r("S1", "Benzene", 10.0, exceeds=1)]
    rc = tmp_path / "results.csv"
    out = tmp_path / "summary.xlsx"
    _write(rc, AnalyticalResultRecord, results)
    r = CliRunner().invoke(_cli, [
        "envmon", "export-summary",
        "--results-csv", str(rc), "--output", str(out)])
    assert r.exit_code == 0, r.output
    assert out.exists()
    wb = openpyxl.load_workbook(out)
    assert "Exceedances" in wb.sheetnames
